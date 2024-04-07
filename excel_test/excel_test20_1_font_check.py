"""
セルのスタイルをチェックする
"""


import excel_data
from excel_data import ExcelSheetDataUtil
from pathlib import Path
import pandas as pd


# print_('*書き込みシートをすべてクリア')
file_name = 'myworkbook.xlsx'
sheet_name = '日付Count'
# sheet_name = 'Copy'
ex_data = ExcelSheetDataUtil(file_name, sheet_name, data_only=True)

from pathlib import Path
_W_PATH = str(Path(__file__)).replace('.py', '.txt')
def print_(value):
    print(value)
    encoding = 'sjis'
    encoding = 'utf-8'
    with open(_W_PATH, 'a', encoding=encoding)as f:
        f.write(value + '\n')

print_('データの用意')
address_list = ['B6','B7','B8']
address_list = ['B7']
# address_list = ['B42']


from openpyxl.styles.fonts import Font
from openpyxl.styles.colors import Color
from openpyxl.styles.borders import Side
from openpyxl.styles import PatternFill
# from openpyxl.styles.borders import _SideStyle
print_('*セルを取得して値を出力する')
for address in address_list:
    buf_cell = ex_data._get_cell(ex_data.sheet, address)
    print_('#############################')
    print_('address = {}'.format(buf_cell.coordinate))
    print_('value = {}'.format(buf_cell.value))
    # フォント属性
    print_('============ font')
    font:Font = buf_cell.font
    print_('buf_cell.font.name = {}'.format(font.name))
    print_('buf_cell.font.size = {}'.format(font.size))
    print_('buf_cell.font.bold = {}'.format(font.bold))
    print_('buf_cell.font.italic = {}'.format(font.italic))
    print_('buf_cell.font.underline = {}'.format(font.underline))
    print_('buf_cell.font.strike = {}'.format(font.strike))
    print_('============ color')
    color:Color = font.color
    print_('buf_cell.font.color = {}'.format(color))
    print_('color.value = {}'.format(color.value))
    print_('color.rgb = {}'.format(color.rgb))
    print_('color.indexed = {}'.format(color.indexed))
    print_('color.type = {}'.format(color.type))
    print_('color.auto = {}'.format(color.auto))
    print_('color.theme = {}'.format(color.theme))
    print_('color.tint = {}'.format(color.tint))

    print_('============ fill')
    # 塗りつぶし（背景色）
    # fill = PatternFill(patternType='lightHorizontal', fgColor='d3d3d3', bgColor='ff1493')
    fill:PatternFill = buf_cell.fill
    print_('buf_cell.fill.fill_type = {}'.format(fill))
    print_('buf_cell.fill.fill_type = {}'.format(fill.fill_type))
    print_('buf_cell.fill.start_color = {}'.format(fill.start_color.value))
    print_('buf_cell.fill.end_color = {}'.format(fill.end_color.value))
    print_('----')
    print_('buf_cell.fill.patternType = {}'.format(fill.patternType))
    fgColor:Color = fill.fgColor
    print_('buf_cell.fill.gfcolor = {}'.format(fgColor.value))
    bgColor:Color = fill.bgColor
    print_('buf_cell.fill.bgColor = {}'.format(bgColor.value))

    print_('============ border')
    # 境界線
    print_('buf_cell.border.left = {}'.format(buf_cell.border.left))
    side:Side = buf_cell.border.left
    # print_('side.border_style = {}'.format(side.border_style))
    print_('side.color = {}'.format(side.color))
    print_('side.color = {}'.format(side.style))
    print_('buf_cell.border.right = {}'.format(buf_cell.border.right))
    print_('buf_cell.border.top = {}'.format(buf_cell.border.top))
    print_('buf_cell.border.bottom = {}'.format(buf_cell.border.bottom))

    print_('============ etc')
    # 数値フォーマット
    print_('buf_cell.number_format = {}'.format(buf_cell.number_format))

    # セルの水平方向の配置
    print_('buf_cell.alignment.horizontal = {}'.format(buf_cell.alignment.horizontal))

    # セルの垂直方向の配置
    print_('buf_cell.alignment.vertical = {}'.format(buf_cell.alignment.vertical))

    # セルのテキストの折り返し
    print_('buf_cell.alignment.wrap_text = {}'.format(buf_cell.alignment.wrap_text))

    # セルの縮小して全体を表示
    print_('buf_cell.alignment.shrink_to_fit = {}'.format(buf_cell.alignment.shrink_to_fit))

    # セルのインデント
    print_('buf_cell.alignment.indent = {}'.format(buf_cell.alignment.indent))

    # セルのテキストの向き
    print_('buf_cell.alignment.text_rotation = {}'.format(buf_cell.alignment.text_rotation))
    print_('#############################')





# try:
#     ex_data.save_book()
# except PermissionError as e:
#     msg = '[!]ERROR: ファイルアクセスエラーです。ファイルが開いていたらファイルを閉じてください。'
#     print_(msg)
#     # raise e