"""
指定したセルから入力or空白が連続したセルを取得する
"""
# https://ramunememo.hatenablog.com/entry/2021/09/23/182202
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
import re
from typing import Union


# def get_cell_value(sheet:Worksheet, address):
#     """
#     セルのデータを取得（文字列）
#     """
#     try:
#         val = sheet[address].value
#         if val == None:
#             val = ''
#         return val
#     except Exception as e:
#         # 文字列に変換できないエラーの時はスキップ 予定（未対応）
#         # 何のエラーが発生するか確認してから対応する
#         raise e
    
# def _get_cell_value(cell_val:Cell):
#     """
#     セルのデータを取得（文字列）
#     """
#     try:
#         return str(cell_val.value)
#     except Exception as e:
#         # 文字列に変換できないエラーの時はスキップ 予定（未対応）
#         # 何のエラーが発生するか確認してから対応する
#         raise e

# def _get_cells_address(cell_val:Cell):
#     """
#     セルの番地を取得
#     """
#     cell_address = get_column_letter(cell_val.column) +  str(cell_val.row)
#     return cell_address

# def _is_match_patterns(patterns:Union[str, list[str]], value:str):
#     if not isinstance(patterns, list):
#         patterns = [str(patterns)]
#     for pattern in patterns:
#         ret = re.search(pattern, value)
#         if ret!=None:
#             return True
#     return False

# def _get_rectangle(worksheet_val:Worksheet, begin_address:str, end_address:str):
#     """ ワークシートと開始・終了アドレスから対象の範囲の セルリスト list[list[Cell]] を取得する """
#     rectangle = worksheet_val[begin_address:end_address]
#     return rectangle

# def _search_keyword_in_rectangle(rectangle:list[list[Cell]] ,pattern:Union[str, list[str]]):
#     """
#     特定の範囲を検索（セルのリストは取得済みであること）

#     Args:
#         rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
#         keyword: 検索対象の文字列、または、パターン
#     """
#     result = []
#     for col in rectangle:
#         for cell in col:
#             value = _get_cell_value(cell)
#             if _is_match_patterns(pattern, value):
#                 cell_address = _get_cells_address(cell)
#                 result.append(cell_address)
#     return result

# def search_rectangle_in_sheet(
#     worksheet_val:Worksheet,
#     begin_address:str,
#     end_address:str,
#     keyword:Union[str, list[str]]):
#     """
#     指定したワークシートの、特定の範囲を検索
#      (ワークシートと開始・終了アドレスを指定するVer)

#     Args
#         ws: worksheet object
#         range_address: address (ex) 'A1:C5'
#     """
#     rectangle = _get_rectangle(worksheet_val,begin_address,end_address)
#     result = _search_keyword_in_rectangle(rectangle, keyword)
#     return result

# def search_entire_sheet(worksheet_val:Worksheet, keyword):
#     """
#     シート全体を検索する

#     Args:
#         rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
#         keyword: 検索対象の文字列、または、パターン
#     """
#     rectangle = worksheet_val.columns
#     result = _search_keyword_in_rectangle(rectangle, keyword)
#     return result

# #########
# def get_a1_address_from_cell(cell:Cell):
#     """ A1形式のアドレスを取得する """
#     return cell.coordinate

# def get_row_and_col_from_a1_address(address:Union[str, Cell]):
#     """ A1形式のアドレスを行と列に分割する """
#     if isinstance(address, Cell):
#         address = address.coordinate
#     column_letter = ''.join(filter(str.isalpha, address))
#     row_number = ''.join(filter(str.isdigit, address))
#     # 列のアルファベットを数値に変換
#     column_number = column_index_from_string(column_letter)
#     # row, colを返す
#     return int(row_number), int(column_number)

# def get_a1_address_from_row_and_col(row:int ,col:int):
#     """ R1C1形式のアドレスを行と列に変換する """
#     # 列番号をアルファベットに変換
#     column_letter = get_column_letter(col)
#     # A1形式でアドレスを組み立て
#     return f"{column_letter}{row}"

# def get_cells(sheet:Worksheet, *args):
#     """
#     セルを取得する
#     """
#     # 引数の最初が数値ならR1C1形式を採用、
#     # この場合2つ目も読み取り、それぞれをrow,colとして扱う
#     if isinstance(args[0], int):
#         row = args[0]
#         col = args[1]
#         cell = sheet.cell(row=row, column=col)
#     elif isinstance(args[0], str):
#         if args[0].isdigit():
#             row = args[0]
#             col = args[1]
#             cell = sheet.cell(row=row, column=col)
#         else:
#             a1 = args[0]
#             cell = sheet[a1]
#     else:
#         cell = None
#     return cell

# def get_offset_cell(sheet:Worksheet, cell:Cell, row, col):
#     """ 特定のセルからrow, col を移動したセルを取得する """
#     now_row, now_col = get_row_and_col_from_a1_address(cell)
#     new_row = now_row + row
#     new_column = now_col + col
#     return sheet.cell(row=new_row, column=new_column)


# def get_offset_address(address:str, row, col):
#     """ 特定のセルaddressからrow, col を移動したセルを取得する """
#     now_row, now_col = get_row_and_col_from_a1_address(address)
#     new_row = now_row + row
#     new_column = now_col + col
#     ret_address = get_a1_address_from_row_and_col(new_row, new_column)
#     return ret_address

# def _value_is_blank(value):
#     if value == '':
#         is_blank = True
#     else:
#         is_blank = False
#     if value == None:
#         is_blank = True
#     return is_blank

# class Direction():
#     LEFT = 1 << 0  # 0000000001 in binary
#     TOP = 1 << 1  # 0000000010 in binary
#     RIGHT = 1 << 2  # 0000000100 in binary
#     BOTTOM = 1 << 3 # 0000001000 in binary

# def _set_dirction(direction):
#     """ 方向をセットする、何もない場合はRIGHT,BOTTOMとする """
#     horizon = None
#     if Direction.RIGHT & direction:
#         horizon = Direction.RIGHT
#     else:
#         horizon = Direction.LEFT
#     vertical = None
#     if Direction.BOTTOM & direction:
#         vertical = Direction.BOTTOM
#     else:
#         vertical = Direction.TOP
#     return horizon | vertical

# def _set_begin_col(begin_address, direction, loop_max_col):
#     # target_row, target_col = get_row_and_col_from_a1_address(begin_address)
#     direction = _set_dirction(direction)
#     if Direction.LEFT & direction:
#         end_col = 1
#     else:
#         #Direction.RIGHT & direction:
#         end_col = loop_max_col
#     return end_col

# def _set_begin_row(begin_address, direction, loop_max_row):
#     # target_row, target_col = get_row_and_col_from_a1_address(begin_address)
#     if Direction.TOP & direction:
#         end_row = 1
#     else:
#         #Direction.BOTTOM & direction:
#         end_row = loop_max_row
#     return end_row

# def _get_for_range(begin_num, end_num):
#     if begin_num <= end_num:
#         return range(begin_num, end_num)
#     else:
#         # return reversed(range(begin_num, end_num))
#         return range(begin_num, end_num, -1)

# def _get_last_add_value(direction):
#     if Direction.LEFT & direction:
#         add_col = +1
#     else:
#         #Direction.RIGHT & direction:
#         add_col = -1
#     if Direction.BOTTOM & direction:
#         add_row = -1
#     else:
#         #Direction.RIGHT & direction:
#         add_row = +1
#     return add_row, add_col


# _MAX = 32767
# def get_end_address_to_end_vertical(sheet:Worksheet, begin_address:str, direction, loop_max=_MAX):
#     """ 連続した空白or入力値のセルの最終のアドレスを取得する 縦方向-垂直"""
#     begin_row, begin_col = get_row_and_col_from_a1_address(begin_address)
#     begin_val = sheet[begin_address]
#     begin_is_blank = _value_is_blank(begin_val)
#     direction_ = _set_dirction(direction)
#     end_row = _set_begin_row(
#         begin_address, direction_, loop_max)
#     last_add_row, last_add_col = _get_last_add_value(direction)
#     for_range = _get_for_range(begin_row, end_row)
#     for row in for_range:
#         now_val = sheet.cell(row=row, column=begin_col).value
#         now_is_blank = _value_is_blank(now_val)
#         # print((now_val, now_is_blank, get_a1_address_from_row_and_col(row, begin_col)))
#         if begin_is_blank == now_is_blank:
#             continue
#         else:
#             row += last_add_row
#             break
#     end_address = get_a1_address_from_row_and_col(row, begin_col)
#     return end_address


# def get_end_address_to_end_horizon(sheet:Worksheet, begin_address:str, direction, loop_max=_MAX):
#     """ 連続した空白or入力値のセルの最終のアドレスを取得する 横方向-水平"""
#     begin_row, begin_col = get_row_and_col_from_a1_address(begin_address)
#     begin_val = sheet[begin_address]
#     begin_is_blank = _value_is_blank(begin_val)
#     direction_ = _set_dirction(direction)
#     end_col = _set_begin_col(
#         begin_address, direction_, loop_max)
#     last_add_row, last_add_col = _get_last_add_value(direction)
#     for_range = _get_for_range(begin_col, end_col)
#     for col in for_range:
#         now_val = sheet.cell(row=begin_row, column=col).value
#         now_is_blank = _value_is_blank(now_val)
#         # print((now_val, now_is_blank, get_a1_address_from_row_and_col(begin_row, col)))
#         if begin_is_blank == now_is_blank:
#             continue
#         else:
#             col += last_add_col
#             break
#     end_address = get_a1_address_from_row_and_col(begin_row, col)
#     return end_address

# def get_table_address(address_list:'list[str]'):
#     """ 複数のA1形式アドレスのリストから矩形のアドレスを取得する """
#     min_row, min_col, max_row, max_col = None, None, None, None
#     for address in address_list:
#         row, col = get_row_and_col_from_a1_address(address)
#         if min_row==None:
#             min_row, min_col = row, col
#             max_row, max_col = row, col
#             continue
#         min_row, max_row = _update_min_max_by_compare(min_row, max_row, row)
#         min_col, max_col = _update_min_max_by_compare(min_col, max_col, col)
#     begin_address = get_a1_address_from_row_and_col(min_row, min_col)
#     end_address = get_a1_address_from_row_and_col(max_row, max_col)
#     return begin_address, end_address

# def _update_min_max_by_compare(min, max, value):
#     """ value が min より小さければminを更新、maxより大きければmaxを更新して返す """
#     if value < min:
#         min = value
#     if max < value:
#         max = value
#     return min, max

import excel_test3_2_3 as ex_mod
from excel_test3_2_3 import CellUtil, Direction


def main():
    # filename = 'FileIO.xlsm'
    filename = 'myworkbook.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['TestData']

    result = ex_mod.search_rectangle_in_sheet(ws, 'A1', 'F38', '■TableA')
    print(result)
    result = ex_mod.search_entire_sheet(ws, '■TableA')
    print(result)
    #### 下方向に向かって連続した空白か入力済みセルの最終アドレスを取得する
    sheet = ws
    begin_address = result[0]
    _MAX = 32767
    begin_row, begin_col = ex_mod.get_row_and_col_from_a1_address(begin_address)
    begin_val = sheet[begin_address]
    begin_is_blank = ex_mod._value_is_blank(begin_val)
    end_row = begin_row
    for offset_row in range(_MAX):
        row = begin_row + offset_row
        now_val = sheet.cell(row=row, column=begin_col).value
        now_is_blank = ex_mod._value_is_blank(now_val)
        # print((now_val, now_is_blank))
        if begin_is_blank == now_is_blank:
            continue
        else:
            end_row = begin_row + offset_row - 1
            
            break
    end_address = ex_mod.get_a1_address_from_row_and_col(end_row, begin_col)
    print(end_address)
    val = sheet[end_address].value
    print(val)
    ###
    print('###')
    #
    end_address = ex_mod.get_end_address_to_end_vertical(
        sheet, begin_address, Direction.TOP, _MAX)
    print(end_address)
    val = ex_mod.get_cell_value(sheet, end_address)
    print(val)
    #
    end_address = ex_mod.get_end_address_to_end_vertical(
        sheet, begin_address, Direction.BOTTOM, _MAX)
    print(end_address)
    val = ex_mod.get_cell_value(sheet, end_address)
    print(val)
    #
    begin_address_b = ex_mod.get_offset_address(begin_address,2,0)
    table_begin_address = begin_address_b
    #
    end_address = ex_mod.get_end_address_to_end_horizon(
        sheet, begin_address_b, Direction.LEFT, _MAX)
    print(end_address)
    val = ex_mod.get_cell_value(sheet, end_address)
    print(val)
    #
    end_address = ex_mod.get_end_address_to_end_horizon(
        sheet, begin_address_b, Direction.RIGHT, _MAX)
    print(end_address)
    val = ex_mod.get_cell_value(sheet, end_address)
    print(val)
    table_end_address = end_address
    #
    print('***')
    table_address = ex_mod.get_table_address([table_begin_address, table_end_address])
    print(table_address)

if __name__ == '__main__':
    main()