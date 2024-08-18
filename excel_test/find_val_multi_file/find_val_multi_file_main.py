
import sys
from pathlib import Path
import pandas as pd

path = Path(__file__).parent
target_name = 'Repository4_python'
while not Path(path).name == '':
    if Path(path).name == target_name:
        if not str(path) in sys.path:
            sys.path.append(str(path))
            break
    path = path.parent
else:
    print('### directory is not found [{}]'.format(target_name))

from excel_test.excel_data import ExcelSheetDataUtil
from openpyxl.cell import Cell
import glob
import re

def _is_match_patterns(patterns:list[str], value):
    for pattern in patterns:
        ret = re.search(pattern, value)
        if ret!=None:
            return True
    return False

def _is_match_conditions(
    include_condition_patterns:list[str],
    ignore_conditions_patterns:list[str],
    path:str):
    if not _is_match_patterns(include_condition_patterns, path):
        return False
    # ignoreリストの方を優先する
    if _is_match_patterns(ignore_conditions_patterns, path):
        return False
    else:
        return True

def execute_find_main(find_value_list = None):
    #/
    if find_value_list==None:
        find_value_list = [
            'test'
        ]
    #/
    excel_file_dir = r'C:\ZMyFolder\after to base\disk_240800'
    include_conditions_regix_pattern = [r'__test_list_.*.xlsx']
    # ignore_conditions_regix_pattern = []
    ignore_conditions_regix_pattern = [r'.*_dir.xlsx', r'BAKUP2306', r'.*_ig.xlsx']
    # check_columna_names = ['name', 'path']
    check_columna_names = ['path']
    output_dir = Path(r'C:\Users\OK\Desktop')
    output_filename = 'find_value_exel_result.csv'
    output_file_path = output_dir.joinpath(output_filename)
    #/
    logger = DummyLogger()
    logger.info('# find_value_list = {}'.format(', '.join(find_value_list)))
    all_paths = glob.glob(excel_file_dir +'/*')
    logger.info('excel_file_dir = {}'.format(excel_file_dir))
    logger.info('len(all_paths) = {}'.format(len(all_paths)))
    excel_paths = []
    for path in all_paths:
        if _is_match_conditions(
            include_conditions_regix_pattern,
            ignore_conditions_regix_pattern,
            path):
            excel_paths.append(path)
    logger.info('len(excel_paths) = {}'.format(len(excel_paths)))
    # #/
    # # debug
    # ループ用のファイルを変更1
    # if 0<len(excel_paths):
    #     target_path = list(excel_paths)[0]
    #/
    # ループ用のファイルを変更2
    # excel_paths = [Path(excel_file_dir).joinpath('__test_list_DISK0__HDD_VIDEO_MEDIA.xlsx')]
    # #/
    all_result_df_list = []
    result_data_list = []
    for i, target_path in enumerate(excel_paths):
        logger.info('########## [{}] ##########'.format(i))
        logger.info('target_path.name = {}'.format(Path(target_path).name))
        value_finder = ValueFinderInExcel(logger)
        value_finder.init_setting(
            target_path, check_columna_names)
        value_finder.open_excel()
        # 例外が発生しました: InvalidFileException
        #openpyxl does not support binary format .xlsb, please convert this file to .xlsx format if you want to open it with openpyxl
        # xlsbファイルはopenpyxlでは読み込めない
        value_finder.set_column_numbers_by_columna_names()
        value_finder.execute_find_value(find_value_list)
        value_finder.get_info_to_dict_list()
        value_finder.align_data(value_finder.result_df_list)
        value_finder.output_data()
        all_result_df_list.extend(value_finder.result_df_list)
        result_data_list.append(value_finder.result_data)
    #/
    logger.info('######## close proc ##########')
    ret_df = pd.DataFrame({})
    result_data:ResultData=None
    for result_data in result_data_list:
        if ret_df.size!=0:
            ret_df = pd.concat([ret_df, result_data.df], axis=0, ignore_index=True)
        else:
            ret_df = result_data.df
        
    logger.info('len(all_result_df_list) = {}'.format(len(all_result_df_list)))
    logger.info('len(ret_df) = {}'.format(ret_df.shape[0]))
    # print(ret_df)
    buf_ret = ResultData()
    buf_ret.df = ret_df
    value_finder._output_data(buf_ret, output_file_path)
    logger.info('Done.')
    print()

########################################################################
########################################################################

class DummyLogger():
    def info(self, value):
        print(value)

class ResultData():
    def __init__(self) -> None:
        self.df = None
    def set_data(self, df:pd.DataFrame):
        self.df = df

# CSV型を想定（1行目にColumnがある連続した表データ）
# 上から下に向かって（指定列）探す。（左から右などは未対応）
class ValueFinderInExcel():
    def __init__(self, logger:DummyLogger) -> None:
        self.logger = logger
    def init_setting(self, path, column_names:list[str], sheet_name:list[str]=''):
        self.path = path
        self.column_names = column_names
        self.column_numbers = []
        self.start_columns = []
        self.file_name = Path(path).name
        self.sheet_name = sheet_name
        self.ex_data:ExcelSheetDataUtil = None
        self.match_cells: list[Cell] = []
        self.result_dict_list = []
        self.df_columns = []
        self.result_df_list = []
        self.result_data = ResultData()
    
    def open_excel(self):
        # init_settingの後に実行する
        self.logger.info('execute open_excel')
        # 1番目のシートを対象とする
        # ex_data = ExcelSheetDataUtil(None, None, data_only=True)
        self.ex_data = ExcelSheetDataUtil(str(self.path), None, data_only=True)
        # ex_data.set_workbook()
        # 1番最初のシートを取得する
        first_sheet = self.ex_data.book.worksheets[0]
        self.ex_data.set_sheet(first_sheet.title)
        self.ex_data._init_param_after_set_sheet()


    def set_column_numbers_by_columna_names(self):
        # open_excel の後に実行する
        self.logger.info('execute set_column_numbers_by_columna_names')
        # 手動で self.column_numbers を指定してもよい
        # ※検索行数はシートの有効範囲を自動的に読み取り設定される execute_find_value内で
        #/
        # self.address を 'A1:E1'形式にしなければならない
        find_address = 'A1:G1' # 暫定対応
        memo_ad = self.ex_data.address
        for column_name in self.column_names:
            self.ex_data.address = find_address
            # buf_cell = self.ex_data.find_entire_row_in_range(1, column_name)
            buf_cell = self.ex_data.find_entire_row_in_range(0, column_name)
            if buf_cell==None:
                msg = 'column_name[{}] が見つからない, file_name={}'.format(column_name, Path(self.path).name)
                raise Exception(msg)
            self.column_numbers.append(buf_cell.column)
            self.start_columns.append(buf_cell)
        self.ex_data.set_address_a1(memo_ad)

    def execute_find_value(self, find_values:list[str]):
        # set_column_numbers_by_columna_names の後に実行する
        self.logger.info('execute execute_find_value')
        # エクセルのCell値検索の、ループ処理の関係で、検索文字列リストループを外側に
        for find_value in find_values:
            for i, column_number in enumerate(self.column_numbers):
                self._execute_fine_value_one_column(column_number, find_value)
    
    def _execute_fine_value_one_column(self, column_number, find_value:str):
        self.logger.info('execute _execute_fine_value_one_column')
        # begin_row = 1
        # end_row = self.ex_data.sheet.max_row
        #/
        # 'B1:B12345'形式で設定しなければならない
        begin_cell = self.ex_data.get_cell_r1c1(1, column_number)
        end_cell = self.ex_data.get_cell_r1c1(self.ex_data.sheet.max_row, column_number)
        find_address = begin_cell.coordinate + ':' + end_cell.coordinate
        memo_ad = self.ex_data.address
        self.ex_data.address = find_address
        self.logger.info('find_address = {}'.format(find_address))
        #/
        buf_cell = self.ex_data.find_entire_col_in_range(column_number, find_value)
        if buf_cell!=None:
            self.match_cells.append(buf_cell)
            buf_cell = self._execute_fine_value_one_column_continue(
                column_number, find_address, buf_cell, find_value)
        if buf_cell==None:
            self.logger.info('find end.')

    def _execute_fine_value_one_column_continue(self, column_number:int,  find_address:str, now_found_cell:Cell, find_value:str):
        """
        上記関数では、最初に一致したCellしか返さないので
        見つかったセルの次のセルから継続して検索する
        今回は1列のみの検索想定とする、複数列は考慮しない（
        なので、次のセルの指定は now_found_cell に row + 1 したところから、下方向に探す
        """
        # self.logger.info('execute _execute_fine_value_one_column')
        # begin_row = 1
        # end_row = self.ex_data.sheet.max_row
        #/
        # 'B1:B12345'形式で設定しなければならない
        begin_cell = self.ex_data.get_cell_r1c1(now_found_cell.row + 1, now_found_cell.column)
        end_address = self.ex_data._get_end_address(find_address)
        end_cell = self.ex_data._get_cell(self.ex_data.sheet, end_address)
        find_address = begin_cell.coordinate + ':' + end_cell.coordinate
        memo_ad = self.ex_data.address
        self.ex_data.address = find_address
        self.logger.info('find_address = {}'.format(find_address))
        #/
        buf_cell = self.ex_data.find_entire_col_in_range(column_number, find_value)
        if buf_cell!=None:
            self.match_cells.append(buf_cell)
            buf_cell = self._execute_fine_value_one_column_continue(
                column_number, find_address, buf_cell, find_value)
        return buf_cell
    
    def get_info_to_dict_list(self):
        # execute_find_value の後に実行する
        self.logger.info('execute get_info_to_dict_list')
        self.df_columns = ['id','name','size','path','is_file','dir','size_b']
        # マッチした行をdictにして、それを一致行分リストにする
        # dictにする予定だったが、DataFrameにしている（dictにはしていない）
        # 以下のFor文完了後、self.result_df_list.append(df)に格納される
        for match_cell in self.match_cells:
            #/
            # ファイルデータ構成は以下の通り
            # "A1は空",name,size,path,is_file,dir,size_b
            # A列は、番号が記載される
            # 2行目以降はレコードがある、nameは空のことがある
            #/
            found_row = self._get_info_to_dict_one_line(match_cell)
            # self.logger.info('found_row = {}'.format(found_row))
    
    def _get_info_to_dict_one_line(self, cell:Cell):
        self.logger.info('execute _get_info_to_dict_one_line')
        self.ex_data.set_cell(cell)
        # self.logger.info('begin_address = {}'.format(self.ex_data.address))
        #/
        col_end = self.ex_data.valid_cells.end_cell.column
        #/
        # buf_right_address = self.ex_data.get_end_address_to_end_horizon(ex_data.Direction.RIGHT)
        buf_right_cell = self.ex_data.get_cell_r1c1(
            cell.row, col_end)
        buf_right_address = buf_right_cell.coordinate
        # self.logger.info('buf_right_address = {}'.format(buf_right_address))
        #/
        # buf_bottom_address = ex_data.get_end_address_to_end_vertical(ex_data.Direction.DOWN)
        buf_bottom_cell = self.ex_data.get_cell_r1c1(
            cell.row, col_end)
        buf_bottom_address = buf_bottom_cell.coordinate
        # self.logger.info('buf_bottom_address = {}'.format(buf_bottom_address))
        #/
        self.ex_data.set_range_address([buf_right_address, buf_bottom_address])
        self.logger.info('range_address = {}'.format(self.ex_data.range_address))

        df = self.ex_data.get_values_from_range_address_pd(
            self.ex_data.range_address, columns=None)
    
        # self.logger.info('df = ')
        # self.logger.info(df.columns)
        # self.logger.info(df.values)
        # DataFrameを1行ずつprintする処理
        # for index, row in df.iterrows():
        #     row_str = ','.join(map(str, row.values))
        #     self.logger.info(f"Index: {index}, Row: {row_str}")
        found_rows = df.shape[0]
        self.logger.info('found_rows = {}'.format(found_rows))
        #/
        # ファイル名を追加
        # 追加する新しい行のすべての列に同じ文字列を設定
        df['db_file_name'] = self.file_name
        #/
        self.result_df_list.append(df)
        return found_rows
    
    def align_data(self, arg_df_list:list[pd.DataFrame]):
        """ 取得後のデータ整形をする """
        # output_dataの前、get_info_to_dict_listの後に実行する
        #/
        # dataframeリストを、1つのdatafrmaeにする
        # self.df_columns = ['id','name','size','path','is_file','dir','size_b']
        self.df_columns = ['name','size','is_file','dir','size_b']
        self.df_columns = ['name','size','is_file','dir','size_b','db_file_name']
        df = pd.DataFrame({}, columns=self.df_columns)
        for buf_df in arg_df_list:
            # if df.shape[0]==0:
            #     df = buf_df
            # else:
            #     df = pd.concat([df, buf_df], axis=0, ignore_index=True)
            # print(df.index.is_unique)  # Trueなら重複なし、Falseなら重複あり
            # print(buf_df.index.is_unique)  # Trueなら重複なし、Falseなら重複あり
            # buf_df = buf_df.reset_index(drop=True)# インデックス
            # df = df.reset_index(drop=True, ignore_index=True)# インデックス
            # buf_df = buf_df.reset_index(drop=True, ignore_index=True)# インデックス
            # buf_df = buf_df.reset_index(drop=True)# インデックス
            # df = df.reset_index(drop=True)# インデックス
            # print('buf_df.size = {}'.format(buf_df.size))
            # print('df.size = {}'.format(df.size))
            # df = pd.concat([df, buf_df], axis=0, ignore_index=True)
            if df.size!=0:
                df = pd.concat([df, buf_df], axis=0, ignore_index=True)
            else:
                df = buf_df
                # df.columns = self.df_columns 
                # 上記のcolumnas設定で以下のエラーが発生する
                # ERROR1  ValueError: Length mismatch: Expected axis has 5 elements, new values have 7 elements  
                # ERROR2  pandas.errors.InvalidIndexError: Reindexing only valid with uniquely valued Index object#
            # df = df.reset_index(drop=True)# インデックス
            # buf_df.reset_index(drop=True)# インデックス
        df.columns = self.df_columns
        self.result_data.set_data(df)
        return df


    def output_data(self):
        # get_info_to_dict_list の後に実行する
        self.logger.info('execute output_data')
        self.logger.info('len(self.result_df_lisg) = {}'.format(len(self.result_df_list)))
        df = self.result_data.df
        for index, row in df.iterrows():
            row_str = ','.join(map(str, row.values))
            self.logger.info(f"Index: {index}, Row: {row_str}")

    def _output_data(self, arg_result_data:ResultData, output_file_path):
        # get_info_to_dict_list の後に実行する
        self.logger.info('execute output_data')
        self.logger.info('len(self.result_df_lisg) = {}'.format(len(self.result_df_list)))
        df = arg_result_data.df
        for index, row in df.iterrows():
            row_str = ', '.join(map(str, row.values))
            self.logger.info(f"Index: {index}, Row: {row_str}")
        ret = df.to_csv(output_file_path, encoding='shift-jis')
        # print(ret)#None
        self.logger.info('output_file_path = {}'.format(output_file_path))

def _get_find_value_list_from_text_file():
    text_file_path = Path(__file__).parent.joinpath('__test_setting_value.txt')
    with open(str(text_file_path), 'r', errors='ignore', encoding='utf-8')as f:
        lines = f.readlines()
    lines = list(lines)
    lines = [x for x in lines if not x.startswith('#')]
    last_line = lines[-1].strip()
    buf_list = last_line.split('/')
    for i, buf in enumerate(buf_list):
        buf_list[i] = buf_list[i].strip()
    ret_list = [x for x in buf_list if not x=='']
    return ret_list

import datetime
if __name__ == '__main__':
    print('*****\n')
    # find_value_list = ['find_value']
    find_value_list = _get_find_value_list_from_text_file()
    start_time = datetime.datetime.now()
    execute_find_main(find_value_list)
    pass_time = datetime.datetime.now() - start_time
    print('pass_time = {}'.format(pass_time))