"""
Excelの行列から特定の文字列を検索し、そこから入力連続したセルを縦・横に取得して、矩形アドレスを取得する
クラス化
"""
# https://ramunememo.hatenablog.com/entry/2021/09/23/182202
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
import re
from typing import Union


def get_cell_value(sheet:Worksheet, address):
    """
    セルのデータを取得（文字列）
    """
    try:
        val = sheet[address].value
        if val == None:
            val = ''
        return val
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        raise e
    
def _get_cell_value(cell_val:Cell):
    """
    セルのデータを取得（文字列）
    """
    try:
        return str(cell_val.value)
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        raise e

def _get_cell_address(value:'Union[str,Cell]'):
    """
    セルの番地を取得(str,Cell)
    """
    if isinstance(value, Cell):
        # cell_address = get_column_letter(cell_val.column) +  str(cell_val.row)
        address_str = value.coordinate
    elif isinstance(value, str):
        if re.match(r'[a-zA-Z]{1,3}[1-9]{1,8}', value)!=None:
            address_str = value
        else:
            raise ValueError('value is not Address')
    else:
        raise TypeError(type(value))
    return address_str

def _is_match_patterns(patterns:Union[str, list[str]], value:str):
    if not isinstance(patterns, list):
        patterns = [str(patterns)]
    for pattern in patterns:
        ret = re.search(pattern, value)
        if ret!=None:
            return True
    return False


def _search_cell_list(cell_list:'list[Cell]', keyword:str):
    """
    セルのリストの中からキーワードに合致するアドレスを取得する

    get_address_list_to_match_keyword
    """
    result = []
    for cell in cell_list:
        value = _get_cell_value(cell)
        if value == keyword:
            cell_address = _get_cell_address(cell)
            result.append(cell_address)
    return result


# # 特定の列を検索 >>> _search_cell_list
# def search_column(column, keyword):
#     result = []
#     for cell in column:
#         value = _get_cell_value(cell)
#         if value == keyword:
#             cell_address = _get_cells_address(cell)
#             result.append(cell_address)
#     return result

# # 特定の行を検索 >>> _search_cell_list
# def search_row(row, keyword):
#     result = []
#     for cell in row:
#         # セルのデータを文字列に変換
#         try:
#             value = str(cell.value)
#         # 文字列に変換できないデータはスキップ
#         except:
#             continue
#         # キーワードに一致するセルの番地を取得
#         if value == keyword:
#             cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
#             result.append(cell_address)
            
#     return result

def _get_address_a1_str(value:'Union[str, Cell]'):
    """
    A1アドレスを取得する
     文字列の場合はそのまま、Cellの場合はCell.coordinateを取得する
    """
    if isinstance(value, str):
        return value
    elif isinstance(value, Cell):
        return str(value.coordinate)
    else:
        return value


def _get_cells_iterator(
        worksheet_val:Worksheet,
        begin_address:'Union[str, Cell]',
        end_address:'Union[str, Cell]'):
    """ 指定した範囲のCellの Iterator を取得する """
    begin_address = _get_address_a1_str(begin_address)
    end_address = _get_address_a1_str(end_address)
    rectangle = worksheet_val[begin_address:end_address]
    return rectangle

# def _get_rectangle(worksheet_val:Worksheet, begin_address:str, end_address:str):
#     rectangle = worksheet_val[begin_address:end_address]
#     return rectangle

def _get_cells_by_search_keyword_in_rectangle(rectangle:list[list[Cell]] ,keyword:str):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列
    
    Returns:
        {list[openxl.cell.Cell]} : 見つかったCellのリスト
    """
    result_cells = []
    for col in rectangle:
        for cell in col:
            value = _get_cell_value(cell)
            if value == keyword:
                # cell_address = _get_cells_address(cell)
                # result.append(cell_address)
                result_cells.append(cell)
    return result_cells


def _get_rectangle(worksheet_val:Worksheet, begin_address:'Union[str,Cell]', end_address:'Union[str,Cell]'):
    """ ワークシートと開始・終了アドレスから対象の範囲の セルリスト list[list[Cell]] を取得する """
    begin_address = _get_cell_address(begin_address)
    rectangle = worksheet_val[begin_address:end_address]
    return rectangle

def _search_keyword_in_rectangle(rectangle:list[list[Cell]] ,pattern:Union[str, list[str]]):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列、または、パターン
    """
    result = []
    for col in rectangle:
        for cell in col:
            value = _get_cell_value(cell)
            if _is_match_patterns(pattern, value):
                cell_address = _get_cell_address(cell)
                result.append(cell_address)
    return result


# def _get_cells_iterator(
#         worksheet_val:Worksheet,
#         begin_address:'Union[str, Cell]',
#         end_address:'Union[str, Cell]'):
#     """ 指定した範囲のCellの Iterator を取得する """
#     begin_address = _get_address_a1_str(begin_address)
#     end_address = _get_address_a1_str(end_address)
#     rectangle = worksheet_val[begin_address:end_address]
#     return rectangle


def _get_cells_by_search_keyword_in_rectangle(rectangle:list[list[Cell]] ,keyword:str):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列
    
    Returns:
        {list[openxl.cell.Cell]} : 見つかったCellのリスト
    """
    result_cells = []
    for col in rectangle:
        for cell in col:
            value = _get_cell_value(cell)
            if value == keyword:
                # cell_address = _get_cells_address(cell)
                # result.append(cell_address)
                result_cells.append(cell)
    return result_cells

def search_rectangle_in_sheet(
    worksheet_val:Worksheet,
    begin_address:str,
    end_address:str,
    keyword:Union[str, list[str]]):
    """
    指定したワークシートの、特定の範囲を検索
     (ワークシートと開始・終了アドレスを指定するVer)

    Args
        ws: worksheet object
        range_address: address (ex) 'A1:C5'
    """
    rectangle = _get_rectangle(worksheet_val,begin_address,end_address)
    result = _search_keyword_in_rectangle(rectangle, keyword)
    return result

def search_entire_sheet(worksheet_val:Worksheet, keyword):
    """
    シート全体を検索する

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列、または、パターン
    """
    rectangle = worksheet_val.columns
    result = _search_keyword_in_rectangle(rectangle, keyword)
    return result

def _get_cells_of_entire_column_by_search(worksheet:Worksheet, keyword):
    """
    シート全体を検索して、マッチした列のlist[Cell]を取得する
    """
    rectangle = worksheet.columns
    # result = _search_keyword_in_rectangle(rectangle, keyword)
    result_cell = _get_cells_by_search_keyword_in_rectangle(rectangle, keyword)
    return result_cell

#########
def get_a1_address_from_cell(cell:Cell):
    """ A1形式のアドレスを取得する """
    return cell.coordinate

def get_row_and_col_from_a1_address(address:Union[str, Cell]):
    """
    A1形式のアドレスを行と列に分割する

    Memo: 'A1:B5' は A1(row=1,col=1) を返す
    """
    address_ = address
    if isinstance(address_, Cell):
        address_ = address_.coordinate
    if ':' in address_:
        address_ = address_[:address_.find(':')]
    # Memo 'A1:B5'は'AB15'と解釈される
    column_letter = ''.join(filter(str.isalpha, address_))
    row_number = ''.join(filter(str.isdigit, address_))
    # 列のアルファベットを数値に変換
    column_number = column_index_from_string(column_letter)
    # row, colを返す
    return int(row_number), int(column_number)

def get_a1_address_from_row_and_col(row:int ,col:int):
    """ R1C1形式のアドレスを行と列に変換する """
    # 列番号をアルファベットに変換
    column_letter = get_column_letter(col)
    # A1形式でアドレスを組み立て
    return f"{column_letter}{row}"

def get_cells(sheet:Worksheet, *args):
    """
    WorkSheetからCellを取得する

    Args:
        *args:
            第1引数がstrならA1形式
            第1引数、第2引数がintならR1C1形式
    """
    # 引数の最初が数値ならR1C1形式を採用、
    # この場合2つ目も読み取り、それぞれをrow,colとして扱う
    if isinstance(args[0], int):
        row = args[0]
        col = args[1]
        cell = sheet.cell(row=row, column=col)
    elif isinstance(args[0], str):
        if args[0].isdigit():
            row = args[0]
            col = args[1]
            cell = sheet.cell(row=row, column=col)
        else:
            a1 = args[0]
            cell:Cell = sheet[a1]
    else:
        cell = None
    return cell

def get_offset_cell(sheet:Worksheet, cell:Cell, row, col):
    """ 特定のセルからrow, col を移動したセルを取得する """
    now_row, now_col = get_row_and_col_from_a1_address(cell)
    new_row = now_row + row
    new_column = now_col + col
    return sheet.cell(row=new_row, column=new_column)

def get_offset_address(address:str, row, col):
    """
    特定のセルaddressからrow, col を移動したセルを取得する
     (Sheetは必要ない)
    """
    now_row, now_col = get_row_and_col_from_a1_address(address)
    new_row = now_row + row
    new_column = now_col + col
    ret_address = get_a1_address_from_row_and_col(new_row, new_column)
    return ret_address

def _value_is_blank(value):
    if value == '':
        is_blank = True
    else:
        is_blank = False
    if value == None:
        is_blank = True
    return is_blank

class Direction():
    LEFT = 1 << 0  # 0000000001 in binary
    TOP = 1 << 1  # 0000000010 in binary
    UP = 1 << 1  # 0000000010 in binary #TOP==UP
    RIGHT = 1 << 2  # 0000000100 in binary
    BOTTOM = 1 << 3 # 0000001000 in binary
    DOWN = 1 << 3 # 0000001000 in binary #BOTTOM==DOWN
    @classmethod
    def _print_direction(cls, direction_flags):
        if Direction.RIGHT & direction_flags:
            print('Direction.RIGHT')
        else:
            print('Direction.LEFT')
        if Direction.DOWN & direction_flags:
            print('Direction.DOWN')
        else:
            print('Direction.UP')

def _set_dirction(direction):
    """ 方向をセットする、何もない場合はRIGHT,BOTTOMとする """
    horizon = None
    if Direction.RIGHT & direction:
        horizon = Direction.RIGHT
    else:
        horizon = Direction.LEFT
    vertical = None
    if Direction.BOTTOM & direction:
        vertical = Direction.BOTTOM
    else:
        vertical = Direction.TOP
    return horizon | vertical

def _set_begin_col(begin_address, direction, loop_max_col):
    # target_row, target_col = get_row_and_col_from_a1_address(begin_address)
    direction = _set_dirction(direction)
    if Direction.LEFT & direction:
        end_col = 1
    else:
        #Direction.RIGHT & direction:
        end_col = loop_max_col
    return end_col

def _set_begin_row(begin_address, direction, loop_max_row):
    # target_row, target_col = get_row_and_col_from_a1_address(begin_address)
    if Direction.TOP & direction:
        end_row = 1
    else:
        #Direction.BOTTOM & direction:
        end_row = loop_max_row
    return end_row

def _get_for_range(begin_num, end_num):
    if begin_num <= end_num:
        return range(begin_num, end_num)
    else:
        # return reversed(range(begin_num, end_num))
        return range(begin_num, end_num, -1)

def _get_last_add_value(direction):
    if Direction.LEFT & direction:
        add_col = +1
    else:
        #Direction.RIGHT & direction:
        add_col = -1
    if Direction.BOTTOM & direction:
        add_row = -1
    else:
        #Direction.RIGHT & direction:
        add_row = +1
    return add_row, add_col


_MAX = 32767
def get_end_address_to_end_vertical(sheet:Worksheet, begin_address:str, direction, loop_max=_MAX):
    """ 連続した空白or入力値のセルの最終のアドレスを取得する 縦方向-垂直"""
    begin_row, begin_col = get_row_and_col_from_a1_address(begin_address)
    begin_val = sheet[begin_address]
    begin_is_blank = _value_is_blank(begin_val)
    direction_ = _set_dirction(direction)
    Direction._print_direction(direction)
    end_row = _set_begin_row(
        begin_address, direction_, loop_max)
    last_add_row, last_add_col = _get_last_add_value(direction)
    for_range = _get_for_range(begin_row, end_row)
    for row in for_range:
        now_val = sheet.cell(row=row, column=begin_col).value
        now_is_blank = _value_is_blank(now_val)
        # print((now_val, now_is_blank, get_a1_address_from_row_and_col(row, begin_col)))
        if begin_is_blank == now_is_blank:
            continue
        else:
            row += last_add_row
            break
    end_address = get_a1_address_from_row_and_col(row, begin_col)
    return end_address

def get_end_cell_to_end_vertical(sheet:Worksheet, begin_address:'Union[str,Cell]', direction, loop_max=_MAX):
    """ 連続した空白or入力値のセルの最終のCellを取得する 縦方向-垂直"""
    end_address = get_end_address_to_end_vertical(
        sheet, begin_address, direction, loop_max)
    end_cell = get_cells(end_address)
    return end_cell

def get_end_address_to_end_horizon(sheet:Worksheet, begin_address:str, direction, loop_max=_MAX):
    """ 連続した空白or入力値のセルの最終のアドレスを取得する 横方向-水平"""
    begin_row, begin_col = get_row_and_col_from_a1_address(begin_address)
    begin_val = sheet[begin_address]
    begin_is_blank = _value_is_blank(begin_val)
    direction_ = _set_dirction(direction)
    end_col = _set_begin_col(
        begin_address, direction_, loop_max)
    last_add_row, last_add_col = _get_last_add_value(direction)
    for_range = _get_for_range(begin_col, end_col)
    for col in for_range:
        now_val = sheet.cell(row=begin_row, column=col).value
        now_is_blank = _value_is_blank(now_val)
        # print((now_val, now_is_blank, get_a1_address_from_row_and_col(begin_row, col)))
        if begin_is_blank == now_is_blank:
            continue
        else:
            col += last_add_col
            break
    end_address = get_a1_address_from_row_and_col(begin_row, col)
    return end_address

def get_end_cell_to_end_horizon(sheet:Worksheet, begin_address:'Union[str,Cell]', direction, loop_max=_MAX):
    """ 連続した空白or入力値のセルの最終のCellを取得する 横方向-水平"""
    end_address = get_end_address_to_end_horizon(
        sheet, begin_address, direction, loop_max)
    end_cell = get_cells(end_address)
    return end_cell

def get_range_cell(sheet:Worksheet, value_list:'list[Union[str,Cell]]')->Cell:
    """ 複数のA1形式アドレスorCellのリストから矩形のCellを取得する """
    range_address = get_table_address(value_list)
    return get_cells(sheet, range_address)

def get_table_address(value_list:'list[Union[str,Cell]]'):
    """
    複数のA1形式アドレスのリストから矩形のアドレスを取得する

    Returns:
        {str} : address 'A1:B2'
    """
    min_row, min_col, max_row, max_col = None, None, None, None
    for value in value_list:
        address = _get_cell_address(value)
        row, col = get_row_and_col_from_a1_address(address)
        if min_row==None:
            min_row, min_col = row, col
            max_row, max_col = row, col
            continue
        min_row, max_row = _update_min_max_by_compare(min_row, max_row, row)
        min_col, max_col = _update_min_max_by_compare(min_col, max_col, col)
    begin_address = get_a1_address_from_row_and_col(min_row, min_col)
    end_address = get_a1_address_from_row_and_col(max_row, max_col)
    return begin_address + ':' + end_address

def get_begin_and_end_cell_in_range(sheet:Worksheet, value_list:'list[Union[str,Cell]]'):
    """ 複数のA1形式アドレスのリストから矩形のアドレスを探して、そのbeginとendのCell取得する """
    begin_address, end_address = get_begin_and_end_address_in_range(value_list)
    begin_cell = get_cells(sheet, begin_address)
    end_cell = get_cells(sheet, end_address)
    return begin_cell, end_cell

def get_begin_and_end_address_in_range(value_list:'list[Union[str,Cell]]'):
    """ 複数のA1形式アドレスのリストから矩形のアドレスを探して、そのbeginとendをアドレスを文字列で取得する """
    min_row, min_col, max_row, max_col = None, None, None, None
    for value in value_list:
        address = _get_cell_address(value)
        row, col = get_row_and_col_from_a1_address(address)
        if min_row==None:
            min_row, min_col = row, col
            max_row, max_col = row, col
            continue
        min_row, max_row = _update_min_max_by_compare(min_row, max_row, row)
        min_col, max_col = _update_min_max_by_compare(min_col, max_col, col)
    begin_address = get_a1_address_from_row_and_col(min_row, min_col)
    end_address = get_a1_address_from_row_and_col(max_row, max_col)
    return begin_address, end_address

def _update_min_max_by_compare(min, max, value):
    """ value が min より小さければminを更新、maxより大きければmaxを更新して返す """
    if value < min:
        min = value
    if max < value:
        max = value
    return min, max


class CellUtil():
    class Direction(Direction):
        pass

    def __init__(self, sheet:Worksheet) -> None:
        self.sheet = sheet
        self.address = ''
        self.range_address = ''
        self.loop_max_row = _MAX
        self.loop_max_col = _MAX
        self.cell = ''

    def set_address_by_find(
            self,
            keyword:str,
            find_begin_address:'Union[str,Cell]'=None,
            find_end_address:'Union[str,Cell]'=None,
            found_number:int=0):
        """
        文字列を検索して、存在したらセルのアドレスをセットする

        Args:
            keyword : regix pattern
            found_number : 複数発見したときのindex
        """
        address_list = CellUtil.get_address_list_by_find_keyword(self.sheet, keyword, find_begin_address, find_end_address)
        if 0<len(address_list):
            address = address_list[found_number]
        self.address = address
        self.cell = get_cell_value(self.sheet, address)
        return address

    def set_cell_by_find(self, keyword:str, find_begin_address:str=None, find_end_address:str=None, found_number:int=0):
        """
        文字列を検索して、存在したらCellをセットする

        Args:
            keyword : regix pattern
            found_number : 複数発見したときのindex
        """
        address = self.set_address_by_find(keyword, find_begin_address, find_end_address, found_number)
        self.cell = get_cells(self.sheet, address)
        return self.cell

    @classmethod
    def get_cells_by_find_keyword(
        cls,
        sheet:Worksheet,
        keyword:str,
        find_begin:'Union[str,Cell]'=None,
        find_end:'Union[str,Cell]'=None):
        """
        文字列を検索して、keywordが含まれるセルのリスト取得する

        Args:
            keyword : regix pattern
        """
        # TODO:
        # 発見したら終了させる
        # Rowを優先、Colを優先を実装
        if find_begin==None or find_end==None:
            # address = search_entire_sheet(sheet, keyword)
            cell_list = _get_cells_of_entire_column_by_search(sheet, keyword)
        else:
            # address = search_rectangle_in_sheet(
            #     sheet, find_begin_address, find_end_address, keyword)
            earch_cells = _get_cells_iterator(sheet, find_begin, find_end)
            cell_list = _get_cells_by_search_keyword_in_rectangle(
                sheet, find_begin, find_end, keyword)
        return cell_list

    # rename find_value
    @classmethod
    def get_address_list_by_find_keyword(cls, sheet:Worksheet, keyword:str, find_begin_address:str=None, find_end_address:str=None):
        """
        文字列を検索して、存在したらセルのアドレスをセットする

        Args:
            keyword : regix pattern
        """
        if find_begin_address==None or find_end_address==None:
            address_list = search_entire_sheet(sheet, keyword)
        else:
            address_list = search_rectangle_in_sheet(
                sheet, find_begin_address, find_end_address, keyword)
        return address_list
    
    # >get_address_by_find_keyword
    @classmethod
    def find_value(cls, sheet:Worksheet, keyword:str, find_begin_address:str=None, find_end_address:str=None):
        """
        文字列を検索して、存在したらセルのアドレスをセットする
         deprecated method

        Args:
            keyword : regix pattern
        """
        if find_begin_address==None or find_end_address==None:
            address = search_entire_sheet(sheet, keyword)
        else:
            address = search_rectangle_in_sheet(
                sheet, find_begin_address, find_end_address, keyword)
        return address
    
    def value_is_blank(self):
        row, col = get_row_and_col_from_a1_address(self.address)
        value = self.sheet.cell(row=row, column=col)
        return _value_is_blank(value)

    def get_end_cell_to_end_vertical(self, direction):
        """ 連続した空白or連続した入力済みのセルの最終Cellを取得する 縦方向-垂直"""
        return get_end_cell_to_end_vertical(
            self.sheet, self.address, direction, self.loop_max_row)

    def get_end_address_to_end_vertical(self, direction):
        """ 連続した空白or連続した入力済みのセルの最終のアドレスを取得する 縦方向-垂直"""
        return get_end_address_to_end_vertical(
            self.sheet, self.address, direction, self.loop_max_row)
    
    def get_end_cell_to_end_horizon(self,direction):
        """ 連続した空白or連続した入力済みのセルの最終Cellを取得する 横方向-水平"""
        return get_end_cell_to_end_horizon(
            self.sheet, self.address, direction, self.loop_max_col)

    def get_end_address_to_end_horizon(self,direction):
        """ 連続した空白or連続した入力済みのセルの最終のアドレスを取得する 横方向-水平"""
        return get_end_address_to_end_horizon(
            self.sheet, self.address, direction, self.loop_max_col)

    def get_table_address(self, direction=Direction.RIGHT|Direction.DOWN):
        """ 現在のCellから縦横のDirection方向に連続した入力セルをたどり、矩形アドレス文字列を取得する """
        horizon_end_address = self.get_end_address_to_end_horizon(direction)
        vertical_end_address = self.get_end_address_to_end_vertical(direction)
        table_address = get_table_address([horizon_end_address, vertical_end_address])
        self.range_address = table_address
        return table_address
    
    def get_range_cells(self, direction=Direction.RIGHT|Direction.DOWN):
        """ 現在のCellから縦横のDirection方向に連続した入力セルをたどり、矩形のCellを取得する """
        horizon_end_cell = self.get_end_cell_to_end_horizon(direction)
        vertical_end_cell = self.get_end_cell_to_end_vertical(direction)
        table_cell:Cell = get_range_cell([horizon_end_cell, vertical_end_cell])
        self.range_address = table_cell.coordinate
        return table_cell

    # def move_address(self, offset_row, offset_col):
    #     """ self.addressから offset_row, offset_col の分を移動する """
    #     now_row, now_col = get_row_and_col_from_a1_address(self.address)
    #     new_row = now_row + offset_row
    #     new_column = now_col + offset_col
    #     new_address = get_a1_address_from_row_and_col(new_row, new_column)
    #     self.address = new_address
    #     return new_address
    
    def move_address(self, offset_row, offset_col):
        """ self.cellから offset_row, offset_col の分を移動する """
        self.cell = get_offset_cell(self.sheet, self.cell, offset_row, offset_col)
        return self.cell
    
    def get_value(self, address=None):
        if address==None:
            address = self.address
        return get_cell_value(self.sheet, address)

    @classmethod
    def _cnv_row_col_from_a1_address(cls, a1_address):
        """ A1 > (1,1)  """
        return get_row_and_col_from_a1_address(a1_address)
    
    @classmethod
    def _cnv_a1_address_from_row_col(cls, row, col):
        """ (1,1) > A1 """
        return get_a1_address_from_row_and_col(row, col)

def main():
    # filename = 'FileIO.xlsm'
    filename = 'myworkbook.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['TestData']

    cellu = CellUtil(ws)
    # result = search_rectangle_in_sheet(ws, 'A1', 'F38', '■TableA')
    result  = cellu.set_address_by_find('■TableA','A1', 'F38')
    print(result)
    # result = search_entire_sheet(ws, '■TableA')
    result  = cellu.set_address_by_find('■TableA')
    print(result)

    result  = cellu.set_cell_by_find('■TableA')
    print(result)
    #### 下方向に向かって連続した空白か入力済みセルの最終アドレスを取得する
    sheet = ws
    # begin_address = result[0]
    begin_address = cellu.address
    _MAX = 32767
    begin_row, begin_col = get_row_and_col_from_a1_address(begin_address)
    begin_val = sheet[begin_address]
    begin_is_blank = _value_is_blank(begin_val)
    end_row = begin_row
    for offset_row in range(_MAX):
        row = begin_row + offset_row
        now_val = sheet.cell(row=row, column=begin_col).value
        now_is_blank = _value_is_blank(now_val)
        # print((now_val, now_is_blank))
        if begin_is_blank == now_is_blank:
            continue
        else:
            end_row = begin_row + offset_row - 1
            
            break
    end_address = get_a1_address_from_row_and_col(end_row, begin_col)
    print(end_address)
    val = sheet[end_address].value
    print(val)
    ###
    print('###')
    #
    # end_address = get_end_address_to_end_vertical(
    #     sheet, begin_address, Direction.TOP, _MAX)
    end_address = cellu.get_end_address_to_end_vertical(Direction.UP)
    print('begin to UP = {}'.format(end_address))
    val = get_cell_value(sheet, end_address)
    print(val)
    #
    # end_address = get_end_address_to_end_vertical(
    #     sheet, begin_address, Direction.BOTTOM, _MAX)
    end_address = cellu.get_end_address_to_end_vertical(Direction.DOWN)
    print('begin to DOWN = {}'.format(end_address))
    val = get_cell_value(sheet, end_address)
    print(val)
    #
    # begin_address_b = get_offset_address(begin_address,2,0)
    begin_address_b = cellu.move_address(2,0)
    table_begin_address = begin_address_b
    #
    # end_address = get_end_address_to_end_horizon(
    #     sheet, begin_address_b, Direction.LEFT, _MAX)
    end_address = cellu.get_end_address_to_end_horizon(Direction.LEFT)
    print('begin to LEFT = {}'.format(end_address))
    val = get_cell_value(sheet, end_address)
    print(val)
    #
    # end_address = get_end_address_to_end_horizon(
    #     sheet, begin_address_b, Direction.RIGHT, _MAX)
    end_address = cellu.get_end_address_to_end_horizon(Direction.RIGHT)
    print('begin to RIGHT = {}'.format(end_address))
    val = get_cell_value(sheet, end_address)
    print(val)
    table_end_address = end_address
    #
    print('***')
    table_address = get_table_address([table_begin_address, table_end_address])
    print('table = {}'.format(table_address))
    # 
    print('=========')
    address = 'A1:B5'
    row, col = get_row_and_col_from_a1_address(address)
    print((address, row, col))

if __name__ == '__main__':
    main()