"""
Excelの行列から特定の文字列を検索し、そこから入力連続したセルを縦・横に取得して、矩形アドレスを取得する
"""
# https://ramunememo.hatenablog.com/entry/2021/09/23/182202
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
import re
import copy
from typing import Union
from pathlib import Path
from datetime import datetime, timedelta
import numpy as np
# for password
import io
import msoffcrypto #pip install msoffcrypto-tool
from typing import Union

_DEBUG = False
# https://office-hack.com/excel/maximum-number-of-lines/
"""
Excel 2007
最大行数	最大列数
1,048,576	16,384
Excel 2010
最大行数	最大列数
1,048,576	16,384
Excel 2013
最大行数	最大列数
1,048,576	16,384
Excel 2016
最大行数	最大列数
1,048,576	16,384
Excel 2019
最大行数	最大列数
1,048,576	16,384
"""

class ConstExcel():
    #1つのSheetの最大列、最大行を決める
    MAX = 32767
    EXCEL_ROW_MAX = 1048576
    EXCEL_COLUMN_MAX = 16384
    # pandasでデータを取得するときに、strに格納するか、StrCellに格納するかを指定するときに使用する
    MODE_VALUE_STR_CELL = 1
    MODE_VALUE_STR = 2
    MODE_VALUE_STR_CELL_EX = 2
    # 何か処理の時に、ROW、COLを分岐させるときに使用する
    ROW = 1
    COLUMN = 2
    COL = COLUMN
    STYLE_NORMAL = '標準'

_MAX = 32767
_EXCEL_ROW_MAX = 1048576
_EXCEL_COLUMN_MAX = 16384
def get_cell_value(sheet:Worksheet, value:'Union[str,Cell]'):
    """
    セルのデータを取得（文字列）
    """
    try:
        if isinstance(value, Cell):
            # val = value.value
            # valueがsheet=Noneでも取得できるようにしている
            val = sheet[value.coordinate]
        else:
            val = sheet[value].value
        if val == None:
            val = ''
        return val
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        # address==Noneの時は TypeError が発生する
        # TypeError: expected string or bytes-like object
        raise e

def get_cell_value_r1c1(sheet:Worksheet, row:int, col:int):
    """
    セルのデータを取得（R1C1形式）

    Returns: Cell
    """
    try:
        val = sheet.cell(row=row, column=col).value
        if val == None:
            val = ''
        return val
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        raise e

def set_cell_value(sheet:Worksheet, address, value):
    """
    セルのデータを入力（文字列）
    """
    try:
        sheet[address].value = value
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        # address==Noneの時は TypeError が発生する
        # TypeError: expected string or bytes-like object
        raise e

def set_cell_value_r1c1(sheet:Worksheet, row:int, col:int, value):
    """
    セルのデータを入力（R1C1形式）
    """
    try:
        sheet.cell(row=row, column=col).value = value
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        raise e

def _get_cell_value(cell_val:Cell):
    """
    セルのデータを取得（文字列）

    Memo:
        セルの値が空のときはNoneが返る
    """
    try:
        return str(cell_val.value)
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        raise e

def _get_address(value:'Union[str,Cell]'):
    """
    セルの番地をA1形式で取得する
    Args:
        value {str,Cell}
    """
    return _get_cell_address(value)

def _get_cell_address(value:'Union[str,Cell]'):
    """
    セルの番地をA1形式で取得する
    Args:
        value {str,Cell}
    """
    if isinstance(value, Cell):
        # cell_address = get_column_letter(cell_val.column) +  str(cell_val.row)
        address_str = value.coordinate
    elif isinstance(value, str):
        if re.match(r'[a-zA-Z]{1,3}[1-9]{1,8}', value)!=None:
            address_str = value
        else:
            raise ValueError('value is not Address')
    else:
        raise TypeError(type(value))
    return address_str

def _is_match_patterns(patterns:Union[str, list[str]], value:str):
    if not isinstance(patterns, list):
        patterns = [str(patterns)]
    for pattern in patterns:
        ret = re.search(pattern, value)
        if ret!=None:
            return True
    return False

def _is_include_address(cell_a:'Union[Cell, str]', cell_b:'Union[Cell,str]'):
    """
    cell_a と cell_b の範囲が重なっている部分があるか判定する

    Returns: {bool}
    """
    a_begin_cell, a_end_cell = get_row_and_col_from_a1_address_range(
        cell_a)
    b_begin_cell, b_end_cell = get_row_and_col_from_a1_address_range(
        cell_b)
    if a_begin_cell.row <= b_begin_cell.row <= a_end_cell.row\
        or a_begin_cell.row <= b_end_cell.row <= a_end_cell.row:
        if a_begin_cell.column <= b_begin_cell.column <= a_end_cell.column\
            or a_begin_cell.column <= b_end_cell.column <= a_end_cell.column:
            return True
    return False


def get_row_and_col_from_a1_address_range(cell_address):
    """
    A1形式のアドレスを行と列に分割する 範囲用 （例:'A1:B2'）

    Memo:
      未対応(対応予定) 'A1:B2, C3, D4:E5:F6'
        ['A1:B2', 'C3', 'D4:E5:F6']
    
    Returns:
        list[SimpleCellInfo, SimpleCellInfo]  :  
          list[begin_cell, bend_cell]
    """
    cell_address = _get_cell_address(cell_address)
    if ':' in cell_address:
        buf = cell_address.split(':')
        begin_address = buf[0]
        end_address = buf[1]
    else:
        begin_address = cell_address
        end_address = cell_address
    row, col = get_row_and_col_from_a1_address(begin_address)
    # begin_cell = SimpleCellInfo()
    begin_cell = Cell(worksheet=None, row=row, column=col)
    # openxl.~.Cell は インスタンス引数にworksheetがあるので、上記クラスで扱う
    row, col = get_row_and_col_from_a1_address(end_address)
    end_cell = Cell(worksheet=None, row=row, column=col)
    return begin_cell, end_cell

def _is_a1_address(value):
    """ check by regix """
    ret = re.search(r'[a-zA-Z]{1,3}[0-9]{1,8}', value)
    if ret!=None:
        return True
    else:
        return False

def _is_single_cell(cell_address:str):
    cell_address = _get_cell_address(cell_address)
    if not ':' in cell_address:
        if not ',' in cell_address:
            return True
    
    row, col = get_row_and_col_from_a1_address(cell_address)
    if row==1 and col==1:
        return True
    else:
        return False

def _is_match_patterns(patterns:Union[str, list[str]], value:str):
    if not isinstance(patterns, list):
        patterns = [str(patterns)]
    for pattern in patterns:
        ret = re.search(pattern, value)
        if ret!=None:
            return True
    return False


def _search_cell_list(cell_list:'list[Cell]', keyword:str):
    """
    セルのリストの中からキーワードに合致するアドレスを取得する

    get_address_list_to_match_keyword
    """
    result = []
    for cell in cell_list:
        value = _get_cell_value(cell)
        if value == keyword:
            cell_address = _get_cell_address(cell)
            result.append(cell_address)
    return result

def _get_address_a1_str(value:'Union[str, Cell]'):
    """
    A1アドレスを取得する
     文字列の場合はそのまま、Cellの場合はCell.coordinateを取得する
    """
    if isinstance(value, str):
        return value
    elif isinstance(value, Cell):
        return str(value.coordinate)
    else:
        return value

def _get_cells_iterator(
        worksheet_val:Worksheet,
        begin_address:'Union[str, Cell]',
        end_address:'Union[str, Cell]'):
    """ 指定した範囲のCellの Iterator を取得する """
    begin_address = _get_address_a1_str(begin_address)
    end_address = _get_address_a1_str(end_address)
    rectangle = worksheet_val[begin_address:end_address]
    return rectangle

# def _get_rectangle(worksheet_val:Worksheet, begin_address:str, end_address:str):
#     """
#     ワークシートと開始・終了アドレスから対象の範囲の セルリスト list[list[Cell]] を取得する
#      WorkShhet['A1:B5']とすればよいので削除予定 
#     """
#     rectangle = worksheet_val[begin_address:end_address]
#     return rectangle

def _get_rectangle(worksheet_val:Worksheet, begin_address:'Union[str,Cell]', end_address:'Union[str,Cell]'):
    """
    ワークシートと開始・終了アドレスから対象の範囲の セルリスト list[list[Cell]] を取得する
     str, Cell に対応
    """
    begin_address = _get_cell_address(begin_address)
    end_address = _get_cell_address(end_address)
    rectangle = worksheet_val[begin_address:end_address]
    # rectangle = worksheet_val[begin_address + ':' + end_address]
    return rectangle

def _get_cells_by_search_keyword_in_rectangle(rectangle:list[list[Cell]] ,keyword:str):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列
    
    Returns:
        {list[openxl.cell.Cell]} : 見つかったCellのリスト
    """
    result_cells = []
    for col in rectangle:
        for cell in col:
            value = _get_cell_value(cell)
            if value == keyword:
                # cell_address = _get_cells_address(cell)
                # result.append(cell_address)
                result_cells.append(cell)
    return result_cells

def _search_keyword_in_rectangle(
        rectangle:list[list[Cell]] ,
        pattern:Union[str, list[str]],
        debug:bool=False):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列、または、パターン
    
    Returns:
        list[str] : A1形式のアドレス文字列のリスト
    """
    result = []
    cell:Cell=None
    for row_line in rectangle: #1列ごとにカウントアップしていく
        for cell in row_line:
            value = _get_cell_value(cell)
            if debug:
                print((cell.coordinate, cell.row, cell.column, value))
            if _is_match_patterns(pattern, value):
                cell_address = _get_cell_address(cell)
                result.append(cell_address)
    return result

# def _get_cells_iterator(
#         worksheet_val:Worksheet,
#         begin_address:'Union[str, Cell]',
#         end_address:'Union[str, Cell]'):
#     """ 指定した範囲のCellの Iterator を取得する """
#     begin_address = _get_address_a1_str(begin_address)
#     end_address = _get_address_a1_str(end_address)
#     rectangle = worksheet_val[begin_address:end_address]
#     return rectangle


def _get_cells_by_search_keyword_in_rectangle(rectangle:list[list[Cell]] ,keyword:str):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列
    
    Returns:
        {list[openxl.cell.Cell]} : 見つかったCellのリスト
    """
    result_cells = []
    for col in rectangle:
        for cell in col:
            value = _get_cell_value(cell)
            if value == keyword:
                # cell_address = _get_cells_address(cell)
                # result.append(cell_address)
                result_cells.append(cell)
    return result_cells

def search_rectangle_in_sheet(
    worksheet_val:Worksheet,
    begin_address:str,
    end_address:str,
    keyword:Union[str, list[str]],
    debug:bool=False):
    """
    指定したワークシートの、特定の範囲を検索
     (ワークシートと開始・終了アドレスを指定するVer)

    Args
        ws: worksheet object
        range_address: address (ex) 'A1:C5'
    """
    rectangle = _get_rectangle(worksheet_val,begin_address,end_address)
    result = _search_keyword_in_rectangle(rectangle, keyword, debug)
    return result

def search_entire_sheet(worksheet_val:Worksheet, keyword):
    """
    シート全体を検索する

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列、または、パターン
    """
    rectangle = worksheet_val.columns
    result = _search_keyword_in_rectangle(rectangle, keyword)
    return result

def _get_cells_of_entire_column_by_search(worksheet:Worksheet, keyword):
    """
    シート全体を検索して、マッチした列のlist[Cell]を取得する
    """
    rectangle = worksheet.columns
    # result = _search_keyword_in_rectangle(rectangle, keyword)
    result_cell = _get_cells_by_search_keyword_in_rectangle(rectangle, keyword)
    return result_cell

#########
def get_a1_address_from_cell(cell:Cell):
    """ A1形式のアドレスを取得する """
    return cell.coordinate

def get_row_and_col_from_any(address:Union[str, Cell]):
    """
    A1形式のアドレスを行と列に分割する(Cellでも対応可)

    Memo:
        'A1:B5' は A1(row=1,col=1) を返す
    Returns:
        int, int  :  row_number, column_number
    """
    return get_row_and_col_from_a1_address(address)

def get_row_and_col_from_a1_address(address:Union[str, Cell]):
    """
    A1形式のアドレスを行と列に分割する(Cellでも対応可)

    Memo:
        'A1:B5' は A1(row=1,col=1) を返す
    Returns:
        int, int  :  row_number, column_number
    """
    address_ = address
    if isinstance(address_, Cell):
        address_ = address_.coordinate
    if ':' in address_:
        address_ = address_[:address_.find(':')]
    # Memo 'A1:B5'は'AB15'と解釈される
    column_letter = ''.join(filter(str.isalpha, address_))
    row_number = ''.join(filter(str.isdigit, address_))
    # 列のアルファベットを数値に変換
    column_number = column_index_from_string(column_letter)
    # row, colを返す
    return int(row_number), int(column_number)

def get_a1_address_from_row_and_col(row:int ,col:int):
    """ R1C1形式のアドレスを行と列に変換する """
    # 列番号をアルファベットに変換
    column_letter = get_column_letter(col)
    # A1形式でアドレスを組み立て
    return f"{column_letter}{row}"

def get_cells(sheet:Worksheet, *args):
    """
    WorkSheetからCellを取得する

    Args:
        *args:
            第1引数がstrならA1形式
            第1引数、第2引数がintならR1C1形式
    
    Retruns:
        Cell
    """
    # 引数の最初が数値ならR1C1形式を採用、
    # この場合2つ目も読み取り、それぞれをrow,colとして扱う
    if isinstance(args[0], int):
        row = args[0]
        col = args[1]
        cell = sheet.cell(row=row, column=col)
    elif isinstance(args[0], str):
        if args[0].isdigit():
            row = args[0]
            col = args[1]
            cell = sheet.cell(row=row, column=col)
        else:
            a1 = args[0]
            cell:Cell = sheet[a1]
    else:
        cell = None
    return cell

def get_offset_cell(sheet:Worksheet, cell:Cell, row, col):
    """ 特定のセルからrow, col を移動したセルを取得する """
    now_row, now_col = get_row_and_col_from_a1_address(cell)
    new_row = now_row + row
    new_column = now_col + col
    return sheet.cell(row=new_row, column=new_column)

def get_offset_address(address:str, row, col):
    """
    特定のセルaddressからrow, col を移動したセルを取得する
     (Sheetは必要ない)
    """
    now_row, now_col = get_row_and_col_from_a1_address(address)
    new_row = now_row + row
    new_column = now_col + col
    ret_address = get_a1_address_from_row_and_col(new_row, new_column)
    return ret_address

def _value_is_blank(value):
    """ cellの値がNoneであるか判定する """
    if isinstance(value, Cell):
        value = value.value
    if value == '':
        is_blank = True
    else:
        is_blank = False
    if value == None:
        is_blank = True
    return is_blank

def reset_cell(cell:Cell):
    """セルをリセットする（文字を消して、書式を'標準'にする"""
    cell.value=None
    cell.style=ConstExcel.STYLE_NORMAL
## END Local Method
######################################################################
######################################################################
## Classes

class StrCell(str):
    """
    文字列として振る舞い、Cellをメンバに保持するクラス
    """
    def __new__(cls, value:Cell):
        if isinstance(value, Cell):
            _value = value.value
        else:
            _value = value
        obj = str.__new__(cls, str(_value))
        obj.cell = value
        return obj

    def __init__(self, value:Cell=None):
        # str クラスはイミュータブルであるため、__init__ での初期化は不要です。
        # cell のセットアップは __new__ で行います。
        self.cell = self.cell
        super().__init__()

# class StrCell(str):
#     """
#     文字列として振る舞い、Cellをメンバに保持するクラス
#      pandasなどで扱うときに、値以外にもCell内の書式などのデータもも扱いたいため、
#       このクラスを使用してCellとValueを扱う。
#     """
#     def __new__(cls, value):
#         #     self = object.__new__(cls)
#         # TypeError: object.__new__(StrCell) is not safe, use str.__new__()
#         if isinstance(value, Cell):
#             self = str.__new__(cls, str(value.value))
#         else:
#             self = str.__new__(cls, value)
#         # cls.cell = None
#         # self.__init__(value)
#         self.__set_cell(value)
#         pass

#     def __init__(self, cell) -> None:
#         self.cell = cell
#         if isinstance(cell, Cell):
#             self = StrCell(str(cell.value))
#             self.cell = cell
#         elif isinstance(cell, str):
#             super().__init__()
#             self = str(cell)
#             # self.cell = None
#         else:
#             self.cell = None
#         str
#     def __set_cell(self, value):
#         if isinstance(value, Cell):
#             self.cell = value
#         else:
#             self.cell = None
    #####
        
    # def __new__(cls, value):
    #     #     self = object.__new__(cls)
    #     # TypeError: object.__new__(StrCell) is not safe, use str.__new__()
    #     if isinstance(value, Cell):
    #         self = StrCell(str(value.value))
    #         self.cell = cell
    #     self = str.__new__(cls)
    #     # cls.cell = None
    #     self.__init__(value)
    #     pass
    # def __init__(self, cell) -> None:
    #     self.cell = cell
    #     if isinstance(cell, Cell):
    #         self = StrCell(str(cell.value))
    #         self.cell = cell
    #     elif isinstance(cell, str):
    #         super().__init__()
    #         self = str(cell)
    #         # self.cell = None
    #     else:
    #         self.cell = None
    #     str
    
    #####
    # def __str__(self):
    #     try:
    #         if self.cell == None:
    #             return ''
    #         self.cell.value
    #     except AttributeError:
    #         return ''

   
# class StrCell(str):
#     """
#     文字列として振る舞い、Cellをメンバに保持するクラス
#      pandasなどで扱うときに、値以外にもCell内の書式などのデータもも扱いたいため、
#       このクラスを使用してCellとValueを扱う。
#     """
#     def __init__(self, value:str, cell:Cell) -> None:
#         if isinstance(cell, Cell):
#             self = StrCell(str(value))
#             self.cell = cell
#         elif isinstance(cell, str):
#             super().__init__()
#             self = str(value)
#         else:
#             self.cell = None
#         ''
#         str

class SimpleCellInfo():
    """ セルの値を扱う（簡単な処理をするときに使用する） """
    def __init__(self) -> None:
        self.row = 0
        self.col = 0
        self.address = ''
    def set_row_and_col(self, row, col):
        self.row = row
        self.col = col
        self.address = get_a1_address_from_row_and_col(row, col)
    def set_address(self, a1_address):
        self.address = a1_address
        row, col = get_row_and_col_from_a1_address(a1_address)
        self.row = row
        self.col = col


class TwoCellsInfo():
    """ 2つのセルの値を扱う（主にBegin,Endの矩形アドレスを扱う想定） """
    def __init__(self) -> None:
        self.begin_cell = SimpleCellInfo()
        self.end_cell = SimpleCellInfo()
    
    def set_row_col(self, min_row, min_col, max_row, max_col):
        self.begin_cell.row = min_row
        self.begin_cell.col = min_col
        self.end_cell.row = max_row
        self.end_cell.col = max_col


class TwoCells():
    """ 2つのセルの値を扱う（主にBegin,Endの矩形アドレスを扱う想定） """
    def __init__(self) -> None:
        self.begin_cell:Cell = None
        self.end_cell:Cell = None
    
    def set_row_col(self, sheet:Worksheet, min_row, min_col, max_row, max_col):
        # self.begin_cell.set_row_and_col(min_row, min_col)
        # self.end_cell.set_row_and_col(max_row, max_col)
        self.begin_cell = get_cells(sheet, min_row, min_col)
        self.end_cell = get_cells(sheet, max_row, max_col)

class Direction():
    LEFT = 1 << 0  # 0000000001 in binary
    TOP = 1 << 1  # 0000000010 in binary
    UP = 1 << 1  # 0000000010 in binary #TOP==UP
    RIGHT = 1 << 2  # 0000000100 in binary
    BOTTOM = 1 << 3 # 0000001000 in binary
    DOWN = 1 << 3 # 0000001000 in binary #BOTTOM==DOWN
    @classmethod
    def _print_direction(cls, direction_flags):
        if Direction.RIGHT & direction_flags:
            if _DEBUG:
                print('flag = Direction.RIGHT', end=' | ')
        else:
            if _DEBUG:
                print('flag = Direction.LEFT', end=' | ')
        if Direction.DOWN & direction_flags:
            if _DEBUG:
                print('flag = Direction.DOWN')
        else:
            if _DEBUG:
                print('flag = Direction.UP')

### End Classes
######################################################################
### Local Method2
def _set_dirction(direction):
    """ 方向をセットする、何もない場合はRIGHT,BOTTOMとする """
    horizon = None
    if Direction.RIGHT & direction:
        horizon = Direction.RIGHT
    else:
        horizon = Direction.LEFT
    vertical = None
    if Direction.BOTTOM & direction:
        vertical = Direction.BOTTOM
    else:
        vertical = Direction.TOP
    return horizon | vertical

def _set_begin_col(begin_address, direction, loop_max_col):
    """ set_end_col """
    # target_row, target_col = get_row_and_col_from_a1_address(begin_address)
    direction = _set_dirction(direction)
    if Direction.LEFT & direction:
        end_col = 1
    else:
        #Direction.RIGHT & direction:
        end_col = loop_max_col
    return end_col

def _set_begin_row(begin_address, direction, loop_max_row):
    # target_row, target_col = get_row_and_col_from_a1_address(begin_address)
    if Direction.TOP & direction:
        end_row = 1
    else:
        #Direction.BOTTOM & direction:
        end_row = loop_max_row
    return end_row

def _get_for_range(begin_num, end_num):
    """
    get_end_address_to_end_horizon を実行するときの、rangeオブジェクトを取得する
      (end_num - begin_num) < 0 の時はStep-1としている
    """
    if begin_num == end_num:
        return range(begin_num, end_num+1)
        # return range(begin_num-1, end_num)
    if begin_num <= end_num:
        return range(begin_num, end_num + 1)
        # return range(begin_num-1, end_num)
    else:
        # return reversed(range(begin_num, end_num))
        return range(begin_num, end_num, -1)

def _get_for_range_b(begin_num, end_num):
    """
    get_end_address_to_end_horizon を実行するときの、rangeオブジェクトを取得する
      (end_num - begin_num) < 0 の時はStep-1としている
    """
    if begin_num <= end_num:
        # return range(begin_num, end_num + 1)
        return range(begin_num-1, end_num)
    else:
        # return reversed(range(begin_num, end_num))
        return range(begin_num, end_num, -1)

def _get_last_add_value(direction):
    """
    get_end_address_to_end_horizon,get_end_address_to_end_vertical
     を実行するときの、最後に追加する数値を取得する
      direction によって値が変更される
    """
    if Direction.LEFT & direction:
        add_col = +1
    else:
        #Direction.RIGHT & direction:
        add_col = -1
        # add_col = 0
    if Direction.BOTTOM & direction:
        add_row = -1
    else:
        #Direction.UP & direction:
        add_row = +1
    return add_row, add_col

def get_end_address_to_end_vertical(
        sheet:Worksheet,
        begin_address:str,
        direction,
        loop_max=_EXCEL_ROW_MAX,
        invalid_all_blank:bool=True,
        debug:bool=None):
    """
    連続した空白or入力値のセルの最終のアドレスを取得する 縦方向-垂直

    Args:
        invalid_all_blank: すべて空白の時はカウントを無効として、開始アドレスを返す
    """
    begin_row, begin_col = get_row_and_col_from_a1_address(begin_address)
    begin_val = sheet[begin_address]
    begin_is_blank = _value_is_blank(begin_val)
    direction_ = _set_dirction(direction)
    Direction._print_direction(direction)
    # 検索時などにMaxを検知すると、その前の行までしか範囲指定しないため、loop_max＋1をする
    # 最終行の次の行の判定ができないと、最終行マイナス1の範囲指定となってします
    # 最大値＋１でエラーとなることは未対応（発生時に対応する予定）
    loop_max += 1
    end_row = _set_begin_row(
        begin_address, direction_, loop_max)
    last_add_row, last_add_col = _get_last_add_value(direction)
    # for_range = _get_for_range(begin_row, end_row+1)
    for_range = _get_for_range(begin_row, end_row)
    for row in for_range:
        now_val = sheet.cell(row=row, column=begin_col).value
        now_is_blank = _value_is_blank(now_val)
        if debug:
            print((now_val, now_is_blank, get_a1_address_from_row_and_col(row, begin_col)))
        if begin_is_blank == now_is_blank:
            continue
        else:
            break
    else:
        pass
    # loopの最終行まで読み込んで、開始セルと終了せるが空白の時は、すべて空白となる。
    # 値がないときはフラグによってカウントを無効として、開始セルを返す
    if invalid_all_blank:
        if (end_row-1) <= row:
            if begin_is_blank and now_is_blank:
                return begin_address
    if _EXCEL_ROW_MAX <= row:
        row = _EXCEL_ROW_MAX
    else:
        # now_is_blank フラグが変更された前のセルが対象
        last_add_row, last_add_col = _get_last_add_value(direction)
        row += last_add_row
    end_address = get_a1_address_from_row_and_col(row, begin_col)
    return end_address

def get_end_cell_to_end_vertical(sheet:Worksheet, begin_address:'Union[str,Cell]', direction, loop_max=_MAX):
    """ 連続した空白or入力値のセルの最終のCellを取得する 縦方向-垂直"""
    end_address = get_end_address_to_end_vertical(
        sheet, begin_address, direction, loop_max)
    end_cell = get_cells(sheet, end_address)
    return end_cell

def get_end_address_to_end_horizon(
        sheet:Worksheet,
        begin_address:str,
        direction,
        loop_max=_EXCEL_COLUMN_MAX,
        invalid_all_blank:bool=True,
        debug:bool=False):
    """
    連続した空白or入力値のセルの最終のアドレスを取得する 横方向-水平

    Args:
        invalid_all_blank: すべて空白の時はカウントを無効として、開始アドレスを返す
    """
    begin_row, begin_col = get_row_and_col_from_a1_address(begin_address)
    begin_val = sheet[begin_address]
    begin_is_blank = _value_is_blank(begin_val)
    direction_ = _set_dirction(direction)
    
    # 検索時などにMaxを検知すると、その前の行までしか範囲指定しないため、loop_max＋1をする
    # 最終行の次の行の判定ができないと、最終行マイナス1の範囲指定となってします
    # 最大値＋１でエラーとなることは未対応（発生時に対応する予定）
    loop_max += 1
    end_col = _set_begin_col(
        begin_address, direction_, loop_max)
    
    last_add_row, last_add_col = _get_last_add_value(direction)
    for_range = _get_for_range(begin_col, end_col)
    for col in for_range:
        now_val = sheet.cell(row=begin_row, column=col).value
        now_is_blank = _value_is_blank(now_val)
        if debug:
            print((now_val, now_is_blank, get_a1_address_from_row_and_col(begin_row, col)), flush=True)
        if begin_is_blank == now_is_blank:
            continue
        else:
            break
    # loopの最終行まで読み込んで、開始セルと終了せるが空白の時は、すべて空白となる。
    # 値がないときはフラグによってカウントを無効として、開始セルを返す
    if invalid_all_blank:
        if (end_col-1) <= col:
            if begin_is_blank and now_is_blank:
                return begin_address
    if _EXCEL_COLUMN_MAX <= col:
        col = _EXCEL_COLUMN_MAX
    else:
        # now_is_blank フラグが変更された前のセルが対象
        last_add_row, last_add_col = _get_last_add_value(direction)
        col += last_add_col
        pass
    end_address = get_a1_address_from_row_and_col(begin_row, col)
    return end_address

def get_end_cell_to_end_horizon(sheet:Worksheet, begin_address:'Union[str,Cell]', direction, loop_max=_MAX):
    """ 連続した空白or入力値のセルの最終のCellを取得する 横方向-水平"""
    end_address = get_end_address_to_end_horizon(
        sheet, begin_address, direction, loop_max)
    end_cell = get_cells(sheet, end_address)
    return end_cell

def get_range_cell(sheet:Worksheet, value_list:'list[Union[str,Cell]]')->Cell:
    """ 複数のA1形式アドレスorCellのリストから矩形のCellを取得する """
    range_address = get_table_address(value_list)
    return get_cells(sheet, range_address)

def get_table_address(value_list:'list[Union[str,Cell]]'):
    """
    複数のA1形式アドレスのリストから矩形のアドレスを取得する

    Returns:
        {str} : address 'A1:B2'
    """
    min_row, min_col, max_row, max_col = None, None, None, None
    for value in value_list:
        address = _get_cell_address(value)
        row, col = get_row_and_col_from_a1_address(address)
        if min_row==None:
            min_row, min_col = row, col
            max_row, max_col = row, col
            continue
        min_row, max_row = _update_min_max_by_compare(min_row, max_row, row)
        min_col, max_col = _update_min_max_by_compare(min_col, max_col, col)
    begin_address = get_a1_address_from_row_and_col(min_row, min_col)
    end_address = get_a1_address_from_row_and_col(max_row, max_col)
    return begin_address + ':' + end_address


def get_max_address(value_list:'list[Union[str,Cell]]'):
    """ 引数のリストの中の最大行と最大列のセルを取得する """
    max_row, max_col = 0,0
    for value in value_list:
        row, col = get_row_and_col_from_any(value)
        if max_row <= row:
            max_row = row
        if max_col <= col:
            max_col = col
    ret = get_a1_address_from_row_and_col(max_row, max_col)
    return ret

def get_begin_and_end_cell_in_range(sheet:Worksheet, value_list:'list[Union[str,Cell]]'):
    """ 複数のA1形式アドレスのリストから矩形のアドレスを探して、そのbeginとendのCell取得する """
    begin_address, end_address = get_begin_and_end_address_in_range(value_list)
    begin_cell = get_cells(sheet, begin_address)
    end_cell = get_cells(sheet, end_address)
    return begin_cell, end_cell

def get_begin_and_end_address_in_range(value_list:'list[Union[str,Cell]]'):
    """ 複数のA1形式アドレスのリストから矩形のアドレスを探して、そのbeginとendをアドレスを文字列で取得する """
    min_row, min_col, max_row, max_col = None, None, None, None
    for value in value_list:
        address = _get_cell_address(value)
        row, col = get_row_and_col_from_a1_address(address)
        if min_row==None:
            min_row, min_col = row, col
            max_row, max_col = row, col
            continue
        min_row, max_row = _update_min_max_by_compare(min_row, max_row, row)
        min_col, max_col = _update_min_max_by_compare(min_col, max_col, col)
    begin_address = get_a1_address_from_row_and_col(min_row, min_col)
    end_address = get_a1_address_from_row_and_col(max_row, max_col)
    return begin_address, end_address

def get_range_address(address_list:'list[str]'):
    """
    複数のA1形式アドレスのリストから矩形のアドレスの始点と終点（左上と右下）を取得する
     Deprecated
    
    Returns:
        begin_address, end_address
    """
    min_row, min_col, max_row, max_col = None, None, None, None
    for address in address_list:
        row, col = get_row_and_col_from_a1_address(address)
        if min_row==None:
            min_row, min_col = row, col
            max_row, max_col = row, col
            continue
        min_row, max_row = _update_min_max_by_compare(min_row, max_row, row)
        min_col, max_col = _update_min_max_by_compare(min_col, max_col, col)
    begin_address = get_a1_address_from_row_and_col(min_row, min_col)
    end_address = get_a1_address_from_row_and_col(max_row, max_col)
    return begin_address, end_address

def _update_min_max_by_compare(min, max, value):
    """ value が min より小さければminを更新、maxより大きければmaxを更新して返す """
    if value < min:
        min = value
    if max < value:
        max = value
    return min, max


def _get_cell_from_sheet(sheet:Worksheet, a1_address:str)->Cell:
    """
    SheetからCellを取得する
     cell = sheet['A1']とすると cell.coordinate がVsCodeでリンクされないため、この関数で取得する
    """
    return sheet[a1_address]


def get_min_row_and_col(value_list:'Union[Cell,str]', max_row=_EXCEL_ROW_MAX, max_col=_EXCEL_COLUMN_MAX):
    min_row , min_col = max_row, max_col
    for value in value_list:
        row, col = get_row_and_col_from_a1_address(value)
        if row < min_row:
            min_row = row
        if col < min_col:
            min_col = col
    return min_row, min_col

def get_min_row_and_col_2d(value_list:'Union[Cell,str]', max_row=_EXCEL_ROW_MAX, max_col=_EXCEL_COLUMN_MAX):
    min_row , min_col = max_row, max_col
    for value_list_a in value_list:
        for value in value_list_a:
            row, col = get_row_and_col_from_a1_address(value)
            if row < min_row:
                min_row = row
            if col < min_col:
                min_col = col
    return min_row, min_col

def get_min_cell(sheet:Worksheet, value_list:'Union[Cell,str]', max_row=0, max_col=0):
    if max_row<1:
        max_row = sheet.max_row
    if max_col<1:
        max_col = sheet.max_column
    min_row, min_col = get_min_row_and_col(value_list, max_row, max_col)
    return Cell(sheet, min_row, min_col)

def get_min_cell_2d(sheet:Worksheet, value_list:'Union[Cell,str]', max_row=0, max_col=0):
    if max_row<1:
        max_row = sheet.max_row
    if max_col<1:
        max_col = sheet.max_column
    min_row, min_col = get_min_row_and_col_2d(value_list, max_row, max_col)
    return Cell(sheet, min_row, min_col)

def get_max_row_and_col(value_list:'Union[Cell,str]', min_row=1, min_col=1):
    max_row , max_col = min_row, min_col
    for value in value_list:
        row, col = get_row_and_col_from_a1_address(value)
        if max_row < row:
            max_row = row
        if max_col < col:
            max_col = col
    return max_row, max_col

def get_max_row_and_col_2d(value_list:'Union[Cell,str]', min_row=1, min_col=1):
    max_row , max_col = min_row, min_col
    for value_list_child in value_list:
        for value in value_list_child:
            row, col = get_row_and_col_from_a1_address(value)
            if max_row < row:
                max_row = row
            if max_col < col:
                max_col = col
    return max_row, max_col

def get_max_cell(sheet:Worksheet, value_list:'Union[Cell,str]', min_row=1, min_col=1):
    max_row, max_col = get_max_row_and_col(value_list, sheet.min_row, sheet.min_column)
    return Cell(sheet, max_row, max_col)

def get_max_cell_2d(sheet:Worksheet, value_list:'Union[Cell,str]', min_row=1, min_col=1):
    max_row, max_col = get_max_row_and_col_2d(value_list, sheet.min_row, sheet.min_column)
    return Cell(sheet, max_row, max_col)


def _align_row_col(begin_row, begin_col, end_row, end_col):
    """
    値の順番をそろえる
    
    end_row < begin_row などとなっているとループ処理の時に意図しない動作となるため
    """
    if end_row < begin_row:
        temp = end_row
        end_row = begin_row
        begin_row = temp
    if end_col < begin_col:
        temp = end_col
        end_col = begin_col
        begin_col = temp
    return begin_row, begin_col, end_row, end_col

def _align_row_col_with_direction(begin_row, begin_col, end_row, end_col, direction):
    """
    値の順番をそろえる
     方向（Driection）も考慮する
    
    end_row < begin_row などとなっているとループ処理の時に意図しない動作となるため
     Direction.UPだと end_col > begin_col となる
    Direction は 縦方向は(UP or DOWN)横方向は(RIGHT or LEFT)どちらかが選択される
     何も指定がない場合は（RIGHT,DOWN）がデフォルト
    """
    begin_row, begin_col, end_row, end_col = _align_row_col(
        begin_row, begin_col, end_row, end_col)
    # Direction
    # RIGHTはそのまま
    if Direction.LEFT & direction:
        temp = begin_row
        begin_row = end_row
        end_row = temp
    # DOWNはそのまま
    if Direction.UP & direction:
        temp = begin_col
        begin_col = end_col
        end_col = temp
    return begin_row, begin_col, end_row, end_col
        
def get_row_and_col_from_rect_address(table_address:str):
    """
    矩形のアドレスから、始点と終点（左上と右下）のアドレスを数値で取得する
      begin_row, begin_col, end_row, end_col の4つの値を取得する
    
    Note:
        セルからFor文で値を取得する時に使用する用
         'A1:B2' 形式の文字列を想定。（コロンで区切られたアルファベット＋数値）
    Returns:
        int, int, int, int:
         begin_row, begin_col, end_row, end_col
    """
    ex_data = ExcelSheetDataUtil(None, None)
    add_list = table_address.split(':')
    begin_address, end_address = get_range_address(add_list)
    begin_row, begin_col = ExcelSheetDataUtil._cnv_row_col_from_a1_address(begin_address)
    end_row, end_col = ExcelSheetDataUtil._cnv_row_col_from_a1_address(end_address)
    return begin_row, begin_col, end_row ,end_col

def get_values_from_range_address(book:Workbook, sheet:Worksheet, range_address:str):
    """
    矩形のセルからすべて値を取得する

    Args:
        table_address : 'A1:B2'
    """
    ex_data = ExcelSheetDataUtil(None, None)
    ex_data.set_workbook(book)
    ex_data.set_sheet(sheet)
    ex_data.range_address = range_address
    row, col = get_row_and_col_from_a1_address(range_address)
    # print('range_address info = '.format((range_address, row, col)))
    begin_row, begin_col, end_row, end_col = get_row_and_col_from_rect_address(range_address)
    cell_values_list = []
    for row in range(begin_row, end_row+1):
        buf_values = []
        for col in range(begin_col, end_col+1):
            value = ex_data.get_value_r1c1(row, col)
            buf_values.append(value)
        cell_values_list.append(buf_values)
    return cell_values_list


def get_values_from_range_address_np(
        book:Workbook,
        sheet:Worksheet,
        range_address:str=None,
        mode=ConstExcel.MODE_VALUE_STR_CELL):
    """
    矩形のセルからすべて値を取得する

    Args:
        range_address : 'A1:B2'
    """
    import numpy as np
    ex_data = ExcelSheetDataUtil(None, None)
    ex_data.set_workbook(book)
    ex_data.set_sheet(sheet)
    ex_data.range_address = range_address
    row, col = get_row_and_col_from_a1_address(range_address)
    # print('range_address info = '.format((range_address, row, col)))
    begin_row, begin_col, end_row, end_col = get_row_and_col_from_rect_address(range_address)

    col_amount = abs(end_col+1 - begin_col)
    cell_values_list = np.arange(0)
    for row in range(begin_row-1, end_row):
        row += 1
        data_rows = np.zeros(col_amount, dtype=object) #型指定しないと数値となる
        for i, col in enumerate(range(begin_col, end_col+1)):
            value_str = set_cell_value_r1c1(sheet, row, col)
            if mode==ConstExcel.MODE_VALUE_STR:
                value = value_str
            else:
                # デフォルトはStrCellで扱う
                # elif ConstExcel.MODE_VALUE_STR_CELL
                # value = StrCell(value_str)
                # value.cell = self.get_cell_r1c1(row, col)
                cell = get_cells(sheet, row, col)
                value = StrCell(cell)
            np.put(data_rows, [i], value)
        # data_rowsを cell_values_list に追加する
        if cell_values_list.size < 1:
            cell_values_list = data_rows
        else:
            cell_values_list = np.vstack((cell_values_list, data_rows))
    return cell_values_list


def reset_cells(sheet:Worksheet, begin_address:str, end_address:str):
    cells = sheet[begin_address : end_address]
    cell:Cell=None
    for rows in cells:
        for cell in rows:
            cell.value = None
            cell.style = ConstExcel.STYLE_NORMAL

def copy_cell(
        src_cell:Union[Cell,'ExcelSheetDataUtil'],
        dist_cell:Union[Cell,'ExcelSheetDataUtil'],
        style:bool=False):
    """
    引数src_cellから dist_cellにセルをコピーする（書式と値）

    Args:
        src_cell : コピー元のCell
            ExcelSheetDataUtil でも可能
        dist_cell : コピー先のCell
        style :
            True=書式をコピーする
        exists_only :
            True=値or書式が設定されているもののみコピーする
    """
    src_cell:Cell = _get_openxl_cell(src_cell)
    dist_cell:Cell = _get_openxl_cell(dist_cell)
    src_value = src_cell.value
    if src_value != '' and src_value!=None:
        dist_cell.value = src_value
    # src_style = src_cell.style
    # if src_style != ConstExcel.STYLE_NORMAL:
    #     # テーブルの書式は無視され’標準’となる
    #     if style:
    #         if src_cell.has_style :
    #             dist_cell.style = src_cell.style
    #             dist_cell.fill = src_cell.fill
    if style:
        dist_cell.style = src_cell.style
        dist_cell.number_format = src_cell.number_format
        # 例外が発生しました: TypeError unhashable type: 'StyleProxy'
        # コピーしないと上記エラーとなる
        dist_cell.fill = copy.copy(src_cell.fill)
        dist_cell.border = copy.copy(src_cell.border)

def _get_openxl_cell(value:Union[Cell,'ExcelSheetDataUtil']):
    if 'ExcelSheetDataUtil' in str(type(value)):
        ret = value.sheet[value.address]
    elif isinstance(value, Cell):
        ret = value
    else:
        ret = value
    return ret

# from excel_data import StrCell
def cnv_date_str_cell(excel_date_number_str_cell:StrCell, date_format='%y/%m/%d'):
    """
    エクセルから読み取った日付データ数値をYMD書式に変換する
        対象の型 StrCell
    Memo:
        "42523" > "2024/11/23"
    """
    buf = ExcelSheetDataUtil._cnv_datetime(excel_date_number_str_cell)
    if isinstance(buf , datetime):
        ret = StrCell(buf.strftime(date_format))
        ret.cell = excel_date_number_str_cell.cell
    else:
        ret = buf
    # print(' * ret type = {}, {}'.format(type(ret), ret))
    return ret

import pandas as pd
def cnv_date_str_yobi_cell(timestamp:pd.Timestamp, date_format='%Y/%m/%d(%aaa)'):
    """
    pd.date_range で作成した pandas.TimeStamp を文字列に変換する。
        対象の型 pandas.TimeStamp
    Memo:
        "2023-01-15 00:00:00" > "2023/01/15(日)"
         <class 'pandas._libs.tslibs.timestamps.Timestamp'> > str
    """
    _YOUBI_REPLACE_STR = '#__aaa__#'
    youbi=''
    if '%aaa' in date_format:
        # 曜日を日本語表記に変換する辞書
        day_of_week_dict = {
            0: '月',
            1: '火',
            2: '水',
            3: '木',
            4: '金',
            5: '土',
            6: '日'
        }
        youbi = day_of_week_dict[timestamp.day_of_week]
        date_format = date_format.replace('%aaa', _YOUBI_REPLACE_STR)
    # Timestampオブジェクトを指定された形式の文字列に変換する
    formatted_date = f"{timestamp.strftime(date_format)}"
    if _YOUBI_REPLACE_STR in date_format:
        formatted_date = formatted_date.replace(_YOUBI_REPLACE_STR, youbi)
    return formatted_date


def remove_youbi(timestamp_str:str, date_format='(%aaa)'):
    """
    文字列の日付データの曜日を消去する。
        対象の型 pandas.TimeStamp
    Memo:
        "2023/01/15(日)" > "2023/01/15"
         str > str
    """
    _YOUBI_REPLACE_STR = '#__aaa__#'
    youbi=''
    if '%aaa' in date_format:
        # 曜日を日本語表記に変換する辞書
        day_of_week_dict = {
            0: '月',
            1: '火',
            2: '水',
            3: '木',
            4: '金',
            5: '土',
            6: '日'
        }
        for buf in day_of_week_dict.values():
            rep_str = date_format.replace(date_format, buf)
            timestamp_str = timestamp_str.replace(rep_str, '')
    return timestamp_str


def cnv_date_str(excel_date_number_str):
    """
    エクセルから読み取った日付データ数値をYMD書式に変換する
        対象の型 str
    Memo:
        "42523" > "2024/11/23"
    """
    buf = ExcelSheetDataUtil._cnv_datetime(excel_date_number_str)
    if isinstance(buf , datetime):
        ret = buf.strftime('%y/%m/%d')
    else:
        ret = buf
    # print(' * ret type = {}, {}'.format(type(ret), ret))
    return ret

### End Local Method2
################################################################################
################################################################################
################################################################################
################################################################################
################################################################################

class ExcelSheetDataUtil():
    """
    エクセルのデータを扱うクラス
     主にCellのアドレス・値を扱う
      （1つのシートのみを対象とする）
    
    use modules:
        主に使用しているクラス・モジュールなど
        import openpyxl
        from openpyxl.cell import Cell
        from openpyxl.worksheet.worksheet import Worksheet
        from openpyxl.workbook.workbook import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.utils import column_index_from_string
    """
    class Direction(Direction):
        pass

    class ConstExcel(ConstExcel):
        pass

    def __init__(self, file_path:str, sheet_name:str, data_only:bool=True, debug:bool=None) -> None:
        self.__init_param(file_path, sheet_name, data_only, debug)
    
    def __init_param(self, file_path:str, sheet_name:str, data_only:bool, debug:bool=None):
        self.debug = debug
        self.set_workbook(file_path, data_only)
        self.set_sheet(sheet_name)
        self._init_param_after_set_sheet()
    
    def _init_param_after_set_sheet(self):
        # シートセット以降の初期化を行う
        self._update_valid_cell_in_sheet()
        self.address = 'A1'
        self.range_address = 'A1'
        # self.loop_max_row = _EXCEL_ROW_MAX
        # self.loop_max_col = _EXCEL_COLUMN_MAX
        # 有効なセル最大までにすると、すべて処理したときに時間がかかるので
        # 初期値Noneとして、sheet.min_rowなどの値を使用する
        self.loop_max_row = None
        self.loop_max_col = None
        # セルを連続で読み取るとき、探すときなどのとき、アクセスしたCellの情報をコンソールに出力する
        self.debug = False
    
    def copy_values_from_other_cell(self, other_cell:'ExcelSheetDataUtil'):
        """
        他の ExcelSheetDataUtil から アドレスとCellのみをコピーする
         （book,sheetの値はコピーされない）
        """
        self.address = other_cell.address
        self.set_cell(other_cell)
        self.range_address = other_cell.range_address
    
    def reset_book_sheet(self, data_only:bool=True):
        self.close()
        sheet_name = self.sheet.title
        self.set_workbook(self.file_path, data_only)
        self.set_sheet(sheet_name)

    def _update_valid_cell_in_sheet(self):
        """ WorkSheetの有効なセルの値を更新する """
        self.valid_cells = TwoCells()
        if self.sheet==None:
            return
        self.valid_cells.set_row_col(
            sheet=self.sheet,
            min_row=self.sheet.min_row,
            min_col=self.sheet.min_column,
            max_row=self.sheet.max_row,
            max_col=self.sheet.max_column)
    
    def get_max_cell(self, update=False):
        """
        WorkSheetの有効なセルを取得する
        """
        if update:
            self._update_valid_cell_in_sheet()
        
    
    def set_sheet(self, sheet_name:Union[str, Worksheet]):
        if isinstance(sheet_name, Worksheet):
            self.sheet = sheet_name
            return
        if sheet_name!=None:
            try:
                self.sheet = self.book[sheet_name]
            except KeyError as e:
                if 'does not exist.' in str(e):
                    # 主にWSheetがを存在しないときに発生する
                    # KeyError 'Worksheet Copy does not exist.'
                    # KeyError("Worksheet {0} does not exist.".format(key))
                    msg = str(e)
                    filename = Path(self.file_path).name
                    msg += '(file={}, sheet={})'.format(filename, sheet_name)
                    raise KeyError(msg)
                else:
                    raise e
        else:
            self.sheet = None
        self._update_valid_cell_in_sheet()
    
    def set_workbook(self, file_path:Union[str,Workbook], data_only:bool):
        """
        open excel file (openpyxl.load_workbook)
        """
        if isinstance(file_path, Workbook):
            self.book = file_path
            self.file_path = None
            return
        self.file_path = str(file_path)
        if file_path!=None:
            if self.debug:
                print('# read_file_path = {}'.format(file_path))
            if not Path(file_path).exists():
                raise FileNotFoundError(file_path)
            self.book = openpyxl.load_workbook(file_path, data_only=data_only)
            # 例外が発生しました: InvalidFileException
            #openpyxl does not support binary format .xlsb, please convert this file to .xlsx format if you want to open it with openpyxl
            # > xlsb ファイルはopenpyxlでは扱えない（他のパッケージでは取り扱えるものがある、未確認）240818
            # 例外が発生しました: PermissionError
            # [Errno 13] Permission denied: 'C:\\ZMyFolder\\after to base\\disk_240800\\~$__test_list_DISK0__HDD_VIDEO_MEDIA.xlsx'
            # > ファイルを開いていると、エラーとなる
        else:
            if self.debug:
                print('# read_file_path = {}'.format(None))
            self.book = None

    def set_workbook_with_pass(self, file_path, password, data_only:bool, out_put_file_path=None, debug:bool=None):
        """パスワード付きExcelファイルを読み込む"""
        decrypted = io.BytesIO()
        if self._get_debug(debug):
            print('* read_file_path = {}'.format(file_path))
            print('  data_only = {}'.format(data_only))
            print('  password = {}'.format(password))
        with open(str(file_path), 'rb') as fp:
            msfile = msoffcrypto.OfficeFile(fp)
            msfile.load_key(password=password)
            msfile.decrypt(decrypted)
        if out_put_file_path==None:
            if self._get_debug(debug):
                print('* out_put_file_path = {}'.format(out_put_file_path))
            file_name = Path(file_path).stem + '_unlock' + Path(file_path).suffix
            out_put_file_path = Path(file_path).parent.joinpath(file_name)
        # 暗号解除した後ファイルに保存
        with open(out_put_file_path, 'wb') as fp:
            fp.write(decrypted.getbuffer())
        # ファイルを開く
        self.set_workbook(out_put_file_path, data_only=data_only)

    def save_book(self, file_path:str=None, debug:bool=True):
        """
        WorkBookを保存する

        Caution:
            * 拡張子は'.xlsx'にすること。
             '.xlsm'で保存すると、開けなくなるので注意。
              この場合拡張子を'.xlsx'に変更すると開けるようになる。
            * ファイルを開いているとエラーが発生する 
                PermissionError [Errno 13] Permission denied: 'myworkbook.xlsx'
        """
        w_path = self._get_file_path(file_path)
        self.book.save(w_path)
        if debug:
            print('# write_excel_file = {}'.format(w_path))

    def create_new_file(
            self,
            file_path:str,
            sheet_name:str,
            set_self:bool=True,
            data_only:bool=False,
            debug:bool=None):
        """
        新規ファイルを作成する（xlsx）のみ対応
         （xlsmは非対応）
        
        Args:
            set_self : 新規作成した後、workbook,sheetなどをselfにセットする
        """
        workbook = openpyxl.Workbook() 
        # workbook.create_sheet()
        names = workbook.get_sheet_names()#この時シートの状態は['Sheet']となっている
        # sheet = workbook.active
        workbook.worksheets[0].title = sheet_name
        names_b = workbook.get_sheet_names()
        workbook.save(file_path)
        if set_self:
            self.__init_param(file_path, sheet_name, data_only, debug)

    def close(self):
        if self.book!=None:
            self.book.close()

    def _get_file_path(self, file_path:str):
        if file_path==None:
            file_path = self.file_path
        return file_path

    def _get_debug(self, arg_debug:bool):
        """ デバッグフラグをチェックする、引数のarg_debugを優先する """
        if arg_debug!=None:
            return arg_debug
        else:
            return self.debug

    def set_address_by_find(
            self, 
            keyword:str, 
            find_begin_address:str=None, 
            find_end_address:str=None, 
            found_number:int=0,
            debug:bool=None,
            check_after:bool=True):
        """
        文字列を検索して、存在したらセルのアドレスをセットする

        Args:
            keyword : regix pattern
            found_number : 複数発見したときのindex
        """
        if find_begin_address==None:
            find_begin_address = 'A1'
        if find_end_address==None:
            find_end_address = get_a1_address_from_row_and_col(
                self.valid_cells.end_cell.row, self.valid_cells.end_cell.col)
        # address_list = ExcelSheetDataUtil.find_value(
        #     self.sheet, keyword, find_begin_address, find_end_address, debug)
        address_list = CellUtil.get_address_list_by_find_keyword(
            self.sheet, keyword, find_begin_address, find_end_address, debug)
        if 0<len(address_list):
            address = address_list[found_number]
        else:
            address = None
            msg = 'keyword is not found(sheet={}, keyword={}, search_range={}:{})'.format(
                self.sheet.title, keyword, find_begin_address, find_end_address)
            raise Exception(msg)
        self.address = address
        cell = get_cells(self.sheet, address)
        self.set_cell(cell)
        if check_after:
            if not self.address_is_valid(self.address):
                msg = 'Not found value(address={})'.format(self.address)
                msg += '(sheet={}, keyword={}, begin={}, end={})'.format(self.sheet.title, keyword, find_begin_address, find_end_address)
                raise Exception(msg)
        # return address
        return cell

    def set_cell_by_find(self, keyword:str, find_begin_address:str=None, find_end_address:str=None, found_number:int=0):
        """
        文字列を検索して、存在したらCellをセットする

        Args:
            keyword : regix pattern
            found_number : 複数発見したときのindex
        """
        address = self.set_address_by_find(keyword, find_begin_address, find_end_address, found_number)
        cell = get_cells(self.sheet, address)
        self.set_cell(cell)
        return cell


    def set_cell(self, value:'Union[str, Cell, ExcelSheetDataUtil]'):
        """
        self.address, self.cell をCellによりセットする

        Args:
            value : str, Cell, ExcelSheetDataUtilに対応
        """
        if isinstance(value, Cell):
            self.cell = value
            # self.set_address_a1(self.cell.coordinate)
            self.address = self.cell.coordinate
        elif isinstance(value, str):
            # self.set_address_a1(value)
            if _is_a1_address(value):
                self.address = value
                self.cell = get_cells(self.sheet, value)
            else:
                msg = 'value is not a1 address(value={})'.format(value)
                raise ValueError(msg)
        elif isinstance(value, ExcelSheetDataUtil):
            self.cell = value.cell
            self.address = value.address
        else:
            msg = str(type(value))
            raise TypeError(msg)

    def set_address_a1(self, a1_address:str):
        """
        self.address, self.cell をA1形式の文字列によりセットする
        """
        cell = self.set_cell(a1_address)
        # self.address = a1_address
        # cell = get_cells(self.sheet, a1_address)
        # self.set_cell(cell)
        return cell

    def set_address_r1c1(self, row=None, col=None):
        """
        row,col によって self.addressをセットする
         row,colがNoneの場合は self.addressの値を使用する。
        """
        if row==None or col==None:
            self_row, self_col=self.get_row_and_col()
            if row==None:
                row = self_row
            if col==None:
                col = self_col
        address = self._cnv_a1_address_from_row_col(row, col)
        self.set_cell(address)

    @classmethod
    def address_is_valid(cls, address):
        ret = re.search('^[A-Z]{1,3}\d+$', str(address))
        if ret!=None:
            return True
        else:
            return False

    # rename find_value
    @classmethod
    def get_cells_by_find_keyword(
        cls,
        sheet:Worksheet,
        keyword:str,
        find_begin:'Union[str,Cell]'=None,
        find_end:'Union[str,Cell]'=None):
        """
        文字列を検索して、keywordが含まれるセルのリスト取得する

        Args:
            keyword : regix pattern
        """
        # TODO:
        # 発見したら終了させる
        # Rowを優先、Colを優先を実装
        if find_begin==None or find_end==None:
            # address = search_entire_sheet(sheet, keyword)
            cell_list = _get_cells_of_entire_column_by_search(sheet, keyword)
        else:
            # address = search_rectangle_in_sheet(
            #     sheet, find_begin_address, find_end_address, keyword)
            earch_cells = _get_cells_iterator(sheet, find_begin, find_end)
            cell_list = _get_cells_by_search_keyword_in_rectangle(
                sheet, find_begin, find_end, keyword)
        return cell_list

    # rename find_value
    @classmethod
    def get_address_list_by_find_keyword(
        cls,
        sheet:Worksheet,
        keyword:str,
        find_begin_address:str=None,
        find_end_address:str=None,
        debug:bool=None):
        """
        文字列を検索して、存在したらセルのアドレスをセットする

        Args:
            keyword : regix pattern
        """
        if find_begin_address==None or find_end_address==None:
            address_list = search_entire_sheet(sheet, keyword)
        else:
            address_list = search_rectangle_in_sheet(
                sheet, find_begin_address, find_end_address, keyword, debug)
        return address_list

    # >get_address_by_find_keyword
    @classmethod
    def find_value(
        cls,
        sheet:Worksheet,
        keyword:str,
        find_begin_address:str=None,
        find_end_address:str=None,
        debug:bool=False):
        """
        文字列を検索して、存在したらセルのアドレスをセットする
         deprecated method

        Args:
            keyword : regix pattern
        """
        if find_begin_address==None or find_end_address==None:
            address_list = search_entire_sheet(sheet, keyword)
        else:
            address_list = search_rectangle_in_sheet(
                sheet, find_begin_address, find_end_address, keyword, debug)
        return address_list
    
    def value_is_blank(self):
        row, col = get_row_and_col_from_a1_address(self.address)
        value = self.sheet.cell(row=row, column=col)
        return _value_is_blank(value)

    def get_end_cell_to_end_vertical(self, direction):
        """ 連続した空白or連続した入力済みのセルの最終Cellを取得する 縦方向-垂直"""
        max_row = self._get_loop_max_row()
        return get_end_cell_to_end_vertical(
            self.sheet, self.address, direction, max_row)

    def get_end_address_to_end_vertical(self, direction):
        """ 連続した空白or連続した入力済みのセルの最終のアドレスを取得する 縦方向-垂直"""
        max_row = self._get_loop_max_row()
        return get_end_address_to_end_vertical(
            self.sheet, self.address, direction, max_row, debug=self.debug)

    def get_end_cell_to_end_horizon(self,direction):
        """ 連続した空白or連続した入力済みのセルの最終Cellを取得する 横方向-水平"""
        max_col = self._get_loop_max_col()
        return get_end_cell_to_end_horizon(
            self.sheet, self.address, direction, max_col)

    def set_end_address_to_end_vertical(self, direction):
        """ 連続した空白or連続した入力済みのセルの最終のアドレスを取得する 縦方向-垂直"""
        address = self.get_end_address_to_end_vertical(direction)
        self.set_address_a1(address)
        return self.cell

    def get_end_address_to_end_horizon(self,direction):
        """ 連続した空白or連続した入力済みのセルの最終のアドレスを取得する 横方向-水平"""
        max_col = self._get_loop_max_col()
        return get_end_address_to_end_horizon(
            self.sheet, self.address, direction, max_col, debug=self.debug)
    
    def set_end_address_to_end_horizon(self,direction):
        """ 連続した空白or連続した入力済みのセルの最終のアドレスを取得する 横方向-水平"""
        self.address = self.get_end_address_to_end_horizon(direction)
        return self.address

    def _get_loop_max_col(self, update_info:bool=False):
        if self.loop_max_col!=None:
            return self.loop_max_col
        else:
            if update_info:
                self._update_valid_cell_in_sheet()
            return self.valid_cells.end_cell.column

    def _get_loop_max_row(self, update_info:bool=False):
        if self.loop_max_row!=None:
            return self.loop_max_row
        else:
            if update_info:
                self._update_valid_cell_in_sheet()
            return self.valid_cells.end_cell.row

    def get_range_address_with_find(self, keyword:str, direction=Direction.RIGHT|Direction.DOWN):
        """
        keywordを検索して、マッチしたaddress から Directionの方向に、連続したセルをチェックして、
         矩形のアドレスの始点と終点（左上と右下）を取得する

        Returns:
            range_address
        """
        cells = search_rectangle_in_sheet(
            self.sheet,
            self.valid_cells.begin_cell,
            self.valid_cells.end_cell,
            keyword)
        if len(cells)<1:
            return None
        self.set_cell(cells[0])
        horizon_end_address = self.get_end_address_to_end_horizon(direction)
        vertical_end_address = self.get_end_address_to_end_vertical(direction)
        begin_address, end_address = get_range_address([horizon_end_address, vertical_end_address])
        self.range_address = begin_address + ':' + end_address
        return self.range_address

    def get_range_address(self, direction=Direction.RIGHT|Direction.DOWN):
        """
        self.address から Directionの方向に、連続したセルをチェックして、
        矩形のアドレスの始点と終点（左上と右下）を取得する

        Returns:
            range_address
        """
        horizon_end_address = self.get_end_address_to_end_horizon(direction)
        vertical_end_address = self.get_end_address_to_end_vertical(direction)
        begin_address, end_address = get_range_address([horizon_end_address, vertical_end_address])
        self.range_address = begin_address + ':' + end_address
        return self.range_address
    
    def get_table_address(self, direction=Direction.RIGHT|Direction.DOWN):
        """
        現在のCellから縦横のDirection方向に連続した入力セルをたどり、矩形アドレス文字列を取得する
         deprecated
        """
        return get_range_address(Direction)

    def get_range_cells(self, direction=Direction.RIGHT|Direction.DOWN, check:bool=False):
        """ 現在のCellから縦横のDirection方向に連続した入力セルをたどり、矩形のCellを取得する """
        horizon_end_cell = self.get_end_cell_to_end_horizon(direction)
        vertical_end_cell = self.get_end_cell_to_end_vertical(direction)
        table_cells:'list[Cell]' = get_range_cell(self.sheet, [horizon_end_cell, vertical_end_cell])
        # if check:
        #     # リスト内のrowをすべてチェックして、Max,Minを取得する
        #     min_cell = get_min_cell(self.sheet, table_cells)
        #     max_cell = get_max_cell(self.sheet)
        # else:
        #     min_cell = table_cells[0]
        #     max_cell = table_cells[-1]
        return table_cells
    
    def get_range_address_in_cell_list(self, cell_list):
        # リスト内のrowをすべてチェックして、Max,Minを取得する
        # 2次元リスト対応
        # get_min_cell.__code__
        if self._cell_list_is_2d(cell_list):
            min_cell = get_min_cell_2d(self.sheet, cell_list)
            max_cell = get_max_cell_2d(self.sheet, cell_list)
        else:
            min_cell = get_min_cell(self.sheet, cell_list)
            max_cell = get_max_cell(self.sheet, cell_list)
        ret = str(min_cell.coordinate) + ':' + max_cell.coordinate
        return ret


    def _cell_list_is_2d(self, cell_list):
        if isinstance(cell_list, list):
            if 0<len(cell_list):
                if isinstance(cell_list[0], list):
                    return True
        elif isinstance(cell_list, tuple):
            if 0<len(cell_list):
                if isinstance(cell_list[0], tuple):
                    return True 
        return False

    ### CopyCells
    def copy_range_address(self,src_begin_cell_ex:'ExcelSheetDataUtil', src_cells:'Union[tuple, tuple[tuple]]', debug:bool=True):
        """
        sheet[address] で取得された src_cellsのでーたを、このクラスのアドレスを開始アドレスとして書き込む
        """
        debug = self._get_debug(debug)
        dist_begin_cell_ex = self
        src_cell_temp:Cell=None
        for crc_cell in src_cells:
            if isinstance(crc_cell, tuple):
                #2次元の場合
                for src_cell_b in crc_cell:
                    src_cell_temp = src_cell_b
                    offset_row, offset_col = src_begin_cell_ex.get_diff_row_and_col(
                        src_cell_temp.row, src_cell_temp.column)
                    dist_now_cell_ex = dist_begin_cell_ex.get_offset_cell_ex(offset_row, offset_col)
                    dist_now_cell_ex.copy_value(src_cell_temp, style=True)
                    ### log
                    value = src_cell_temp.value
                    if debug:
                        # if isinstance(value, datetime):
                        #     print()
                        print(' src({},{}) >> dist({},{})  [value={}]'.format(
                            src_cell_temp.row, src_cell_temp.column,
                            dist_now_cell_ex.cell.row, dist_now_cell_ex.cell.column, value))
            else:
                #1次元の場合
                src_cell_temp = crc_cell[0]
                offset_row, offset_col = src_begin_cell_ex.get_diff_row_and_col(
                    src_cell_temp.row, src_cell_temp.column)
                dist_now_cell_ex = dist_begin_cell_ex.get_offset_cell_ex(offset_row, offset_col)
                dist_now_cell_ex.copy_value(src_cell_temp, style=True)
                ### log
                value = src_cell_temp.value
                if debug:
                    print(' src({},{}) >> dist({},{})  [value={}]'.format(
                        src_cell_temp.row, src_cell_temp.column,
                        dist_now_cell_ex.cell.row, dist_now_cell_ex.cell.column, value))
            ###



    ### add for test4
    def get_begin_address_in_range(self):
        if not ':' in self.range_address:
            return None
        ret = self.range_address.split(':')[0]
        return ret
    def get_end_address_in_range(self):
        if not ':' in self.range_address:
            return None
        ret = self.range_address.split(':')[1]
        return ret

    def get_rows_range(self):
        begin_cell, end_cell = get_row_and_col_from_a1_address_range(
            self.address)
        return range(begin_cell.row, end_cell.row+1)
    
    def get_cols_range(self):
        begin_cell, end_cell = get_row_and_col_from_a1_address_range(
            self.address)
        return range(begin_cell.column, end_cell.column+1)
    
    def set_address_entire(self, opt):
        """
        列全体（or行全体）をself.addressにセットする

        Args:
            opt : row=1, col=2
        """
        if opt==ConstExcel.ROW:
            # row
            re_ret = re.search('[1-9]{1,8}', self.address)
            row_str = re_ret.group()
            begin_address = 'A' + row_str
            col_max = get_column_letter(self.valid_cells.end_cell.column)
            end_address = col_max + row_str
        elif opt==ConstExcel.COL:
            # col
            re_ret = re.search('[a-zA-Z]{1,3}', self.address)
            col_str = re_ret.group()
            begin_address = col_str + '1'
            end_address = col_str + str(self.valid_cells.end_cell.row)
        self.address = begin_address + ':' + end_address
        return self.address

    def set_range_address(self, address_value:'Union[str, list[str]]'):
        if isinstance(address_value, str):
            address_value = [address_value]
        elif isinstance(address_value, list):
            pass
        else:
            msg = 'address_value type is invalid. (type={})'.format(type(address_value))
            raise TypeError(msg)
        if self.address != '':
            address_value = [self.address] + address_value
        begin_address, end_address = get_range_address(address_value)
        self.range_address = begin_address + ':' + end_address

    def _address_is_loop_max_row(self, address:str):
        """ address がこのクラス内で処理する最大 "行" かを判定する """
        arg_row, arg_col = get_row_and_col_from_a1_address(address)
        if self.loop_max_row <= arg_row:
            return True
        else:
            return False

    def _address_is_loop_max_col(self, address:str):
        """ address がこのクラス内で処理する最大 "列" かを判定する """
        arg_row, arg_col = get_row_and_col_from_a1_address(address)
        if self.loop_max_row <= arg_row:
            return True
        else:
            return False

    def move_address(self, offset_row:int, offset_col:int):
        """ self.cell から offset_row, offset_col の分を移動する """
        # now_row, now_col = get_row_and_col_from_a1_address(self.address)
        # new_row = now_row + offset_row
        # new_column = now_col + offset_col
        # new_address = get_a1_address_from_row_and_col(new_row, new_column)
        # self.address = new_address
        # return new_address
        if (offset_row==0) and (offset_col==0):
            return self.cell
        cell = get_offset_cell(self.sheet, self.cell, offset_row, offset_col)
        self.set_cell(cell)
        return cell

    @classmethod
    def move_address_from(address:str, offset_row:int, offset_col:int):
        """
        addressから offset_row, offset_col の分を移動する

        Args:
            Address:A1形式のアドレスのみ対応
        """
        now_row, now_col = get_row_and_col_from_a1_address(address)
        new_row = now_row + offset_row
        new_column = now_col + offset_col
        new_address = get_a1_address_from_row_and_col(new_row, new_column)
        return new_address

    def get_value(self, address=None):
        if address==None:
            address = self.address
        return get_cell_value(self.sheet, address)

    def get_value_r1c1(self, row:int, col:int):
        # self.sheet.cell(row=row, column=col).value
        return get_cell_value_r1c1(self.sheet, row, col)

    def get_cell_r1c1(self, row:int, col:int)->Cell:
        # self.sheet.cell(row=row, column=col).value
        return self.sheet.cell(row=row, column=col)
    
    def get_cell(self):
        """
        self.address のセルを取得する
        """
        row, col = get_row_and_col_from_a1_address(self.address)
        return self.sheet.cell(row=row, column=col)
    
    @classmethod
    def _get_cell(cls, sheet:Worksheet, address:str):
        # self.sheet.cell(row=row, column=col).value
        row, col = get_row_and_col_from_a1_address(address)
        # return Cell(worksheet=sheet, row=row, column=col)
        return sheet.cell(row=row, column=col)
    
    @classmethod
    def _get_cell_not_sheet(cls, address:str):
        return cls._get_cell(None, address)

    def set_value(self, value:str, address:str=None):
        if address==None:
            address = self.address
        return set_cell_value(self.sheet, address, value)

    def set_value_offset(self, value:str, offset_row:int=0, offset_col:int=0):
        cell = self.get_offset_cell(offset_row, offset_col)
        return self.set_value(value, cell.coordinate)

    def copy_value(self, src_cell:Union[Cell,'ExcelSheetDataUtil'], style:bool=False, exists_only:bool=True):
        """
        引数src_cellから self.addressに値をコピーする

        Args:
            src_cell : コピー元のCell
                ExcelSheetDataUtil でも可能
            style : True=書式をコピーする
            exists_only : True=値or書式が設定されているもののみコピーする
        """
        copy_cell(
            src_cell,
            self.cell,
            style=style)
        # # dist_cell:Cell = self.sheet[self.address]
        # dist_cell = self._get_openxl_cell(self)
        # src_cell:Cell = self._get_openxl_cell(src_cell)
        # src_value = src_cell.value
        # if src_value != '':
        #     dist_cell.value = src_value
        # src_style = src_cell.style
        # if src_style != self.ConstExcel.STYLE_NORMAL:
        #     # テーブルの書式は無視され’標準’となる
        #     if style:
        #         if src_cell.has_style :
        #             dist_cell.style = src_cell.style
        #             dist_cell.fill = src_cell.fill
        # if style:
        #     dist_cell.style = src_cell.style
        #     # 例外が発生しました: TypeError unhashable type: 'StyleProxy'
        #     # コピーしないと上記エラーとなる
        #     dist_cell.number_format = src_cell.number_format
        #     dist_cell.fill = copy.copy(src_cell.fill)
        #     dist_cell.border = copy.copy(src_cell.border)
    
    def _get_openxl_cell(self, value:Union[Cell,'ExcelSheetDataUtil']):
        if 'ExcelSheetDataUtil' in str(type(value)):
            ret = value.sheet[value.address]
        elif isinstance(value, Cell):
            ret = value
        else:
            ret = value
        return ret

    def set_value_r1c1(self, value:str, row:int, col:int):
        # self.sheet.cell(row=row, column=col).value
        return get_cell_value_r1c1(self.sheet, value, row, col)

    def get_diff_row_and_col(self, row:int , col:int):
        """
        引数のアドレス[row,col]から、self.address(self.cell)のアドレスを引いたrow,colを取得する
         ※cellコピーをするときにコピー元のbegin_cellとnowの差分を算出するとき使用する
        """
        now_row, now_col = get_row_and_col_from_a1_address(self.address)
        return row - now_row, col - now_col

    def get_offset_row_and_col(self, offset_row:int, offset_col:int):
        """
        self.address(self.cell)から、offset[row,col]を足したrow,colを取得する
         ※cellコピーをするときにコピー先のbegin_cellに足して今のコピー先を算出するのに使用する
        """
        row, col = get_row_and_col_from_a1_address(self.address)
        return row + offset_row , col + offset_col

    def get_offset_cell(self, offset_row:int, offset_col:int)->Cell:
        """
        self.address(self.cell)から、offset[row,col]を足した "Cell" を取得する
         ※cellコピーをするときにコピー先のbegin_cellに足して今のコピー先を算出するのに使用する
        """
        return self.get_offset_cell_ex(offset_row, offset_col).cell

    def get_offset_cell_ex(self, offset_row:int, offset_col:int)->'ExcelSheetDataUtil':
        """
        self.address(self.cell)から、offset[row,col]を足した "ExcelSheetDataUtil" を取得する
         ※cellコピーをするときにコピー先のbegin_cellに足して今のコピー先を算出するのに使用する
        """
        cell_ex = self.copy_self()
        cell_ex.move_address(offset_row, offset_col)
        return cell_ex


    def get_row_and_col(self):
        """
        self.address(A1形式のアドレス)から、行と列を取得する

        Memo:
            'A1:B5' は A1(row=1,col=1) を返す
        Returns:
            int, int  :  row_number, column_number
        """
        row, col = get_row_and_col_from_a1_address(self.address)
        return row, col
    
    def get_row_and_col_begin(self):
        address = self.get_begin_address()
        return get_row_and_col_from_a1_address(address)
    
    def get_row_and_col_end(self):
        address = self.get_end_address()
        return get_row_and_col_from_a1_address(address)
    
    def get_begin_address(self):
        if ':' in self.address:
            buf = self.address.split(':')
            return buf[0]
        else:
            return self.address

    def get_end_address(self):
        if ':' in self.address:
            buf = self.address.split(':')
            return buf[-1]
        else:
            return self.address
        
    def _get_end_address(self, address:str):
        if ':' in address:
            buf = address.split(':')
            return buf[-1]
        else:
            return address
    
    def get_row(self):
        row, col = get_row_and_col_from_a1_address(self.address)
        return row

    def get_col(self):
        row, col = get_row_and_col_from_a1_address(self.address)
        return col
    
    def get_column_letter(self):
        return get_column_letter(self.cell.column)
    
    @classmethod
    def _cnv_a1_address_from_cell(cls, cell:Cell):
        return get_a1_address_from_cell(cell)
        return _get_cells_address(cell)

    @classmethod
    def _cnv_row_col_from_a1_address(cls, a1_address):
        """ A1 > (1,1)  """
        return get_row_and_col_from_a1_address(a1_address)

    @classmethod
    def get_row_and_col_from_a1_address(cls, a1_address):
        """ A1 > (1,1)  """
        return get_row_and_col_from_a1_address(a1_address)
    
    @classmethod
    def _cnv_a1_address_from_row_col(cls, row, col):
        """ (1,1) > A1 """
        return get_a1_address_from_row_and_col(row, col)

    @classmethod
    def get_a1_address_from_row_and_col(cls, a1_address):
        """ A1 > (1,1)  """
        return get_a1_address_from_row_and_col(a1_address)

    @classmethod
    def _get_cells_address(cls, cell:Cell):
        """
        Cell からセルの番地を取得する(A1形式)
        """
        return _get_cell_address(cell)

    def get_values_from_range_address_np(
            self,
            range_address:str=None,
            mode=ConstExcel.MODE_VALUE_STR_CELL):
        """
        矩形のセルからすべて値を取得する

        Args:
            range_address :
                対象の範囲
                 例）'A1:B2'
            mode :
                値を取得する型を指定する
                    ConstExcel.MODE_VALUE_STR_CELL : StrCell型で取得する（メンバcellを持つstr）
                    ConstExcel.MODE_VALUE_STR : 通常のstrで取得する

        Returns:
            numpy.Array
        """
        import numpy as np
        if range_address==None:
            range_address = self.range_address
        row, col = get_row_and_col_from_a1_address(range_address)
        # print('range_address info = '.format((range_address, row, col)))
        begin_row, begin_col, end_row, end_col = get_row_and_col_from_rect_address(range_address)

        col_amount = abs(end_col+1 - begin_col)
        cell_values_list = np.arange(0)
        for row in range(begin_row-1, end_row):
            row += 1
            data_rows = np.zeros(col_amount, dtype=object) #型指定しないと数値となる
            for i, col in enumerate(range(begin_col, end_col+1)):
                # if get_a1_address_from_row_and_col(row, col) == 'J42':
                #     print()
                value_str = self.get_value_r1c1(row, col)
                if mode==ConstExcel.MODE_VALUE_STR:
                    value = value_str
                else:
                    # デフォルトはStrCellで扱う
                    # elif ConstExcel.MODE_VALUE_STR_CELL
                    # value = StrCell(value_str)
                    # value.cell = self.get_cell_r1c1(row, col)
                    cell = self.get_cell_r1c1(row, col)
                    value = StrCell(cell)
                np.put(data_rows, [i], value)
            # data_rowsを cell_values_list に追加する
            if cell_values_list.size < 1:
                cell_values_list = data_rows
            else:
                cell_values_list = np.vstack((cell_values_list, data_rows))
        return cell_values_list


    def get_values_from_range_address_pd_temp(
            self, range_address:str=None,
            columns:int=0,
            index:int=None,
            mode=ConstExcel.MODE_VALUE_STR_CELL):
        """
        矩形のセルからすべて値を取得する

        Args:
            range_address : 'A1:B2'
            columns:
                df.columnsに指定する行（この次の行からdfとして扱う）
            mode:
                エクセルから読み取ったときにdfの中の値の型を指定する
                    ConstExcelで指定する
        Returns:
            pandas.DataFrame
        """
        import pandas as pd
        cell_values_np = self.get_values_from_range_address_np(range_address, mode=mode)
        if columns!=None and 0<=columns:
            columns_np = cell_values_np[columns]
            cell_values_np = cell_values_np[columns+1:]
        else:
            columns_np = None
        if index!=None and index!=0:
            index_np = cell_values_np[:, 0]
            cell_values_np = cell_values_np[:, 1:]
        else:
            index_np = None
        df = pd.DataFrame(
            cell_values_np,
            columns=columns_np, index=index_np)
        if columns!=None and 0<=columns:
            # first_row_as_list = df.iloc[columns_np].astype(str).tolist()
            # first_row_as_list = columns_np.astype(str).tolist()
            first_row_as_list = df.iloc[columns].astype(str).tolist()
            df.columns = first_row_as_list
        return df

    def get_values_from_range_address_pd(
            self, range_address:str=None,
            columns:int=0,
            index:int=None,
            mode=ConstExcel.MODE_VALUE_STR_CELL):
        """
        矩形のセルからすべて値を取得する

        Args:
            range_address : 'A1:B2'
            columns:
                df.columnsに指定する行（この次の行からdfとして扱う）
                0以上：その数値の行をcolumnとする
                None or (-1以下):columnを設定しない
            mode:
                エクセルから読み取ったときにdfの中の値の型を指定する
                    ConstExcelで指定する
        Returns:
            pandas.DataFrame
        """
        import pandas as pd
        cell_values_np = self.get_values_from_range_address_np(
            range_address, mode=mode)
        if columns==None:
            columns = -1
        if 0<=columns:
            columns_np = cell_values_np[0]
            cell_values_np = cell_values_np[1:]
        elif columns<0:
            columns_np = []
            col_amount = cell_values_np.shape[0]
            # columns_np = ['' for x in range(col_amount)]
            columns_np = np.full(col_amount, '', dtype='<U1')
            # 240818 
            # 暫定修正
            # エクセルの途中の行読み取り時、columns=Noneだと、以下の形状不一致エラーとなる            
            # ValueError: Shape of passed values is (5, 1), indices imply (5, 0)
            # columnsを指定しない場合、空の配列を作成して、
            # いったん無理やり形状を合わせて処理を進める
            #行方向に連結する
            cell_values_np = np.vstack((columns_np, cell_values_np))
            columns_np = cell_values_np[0]
            cell_values_np = cell_values_np[1:]
        else:
            columns_np = None
        if index!=None and index!=0:
            index_np = cell_values_np[:, 0]
            cell_values_np = cell_values_np[:, 1:]
        else:
            index_np = None
        df = pd.DataFrame(
            cell_values_np,
            columns=columns_np, index=index_np)
        if columns!=None and 0<=columns:
            # first_row_as_list = df.iloc[columns_np].astype(str).tolist()
            # first_row_as_list = columns_np.astype(str).tolist()
            first_row_as_list = df.iloc[columns].astype(str).tolist()
            df.columns = first_row_as_list
        return df

    def write_to_cell_by_pd(self):
        raise NotImplementedError()

    @classmethod
    def _cnv_datetime(cls, value):
        """ str(value)が isdiit=True の時intに変換する（str StrCellに対応） """
        if isinstance(value, str):
            if not value.isdigit():
                return value
            else:
                value = int(value)
        elif isinstance(value, StrCell):
            if not str(value).isdigit():
                return value
            else:
                value = int(value)
        elif isinstance(value, int):
            pass
        else:
            return value
        # return(datetime(1899, 12, 30) + timedelta(days=value))
        # return(datetime(1899, 12, 30) + timedelta(days=str(value))) #TypeError: unsupported type for timedelta days component: str
        return(datetime(1899, 12, 30) + timedelta(days=value))
    
    def copy_self(self, cell_or_address_value:'Union[str,Cell]'=None)->'ExcelSheetDataUtil':
        """
        ExcelSheetDataUtil クラスをコピーして取得する
         Sheet,Book,path などのメンバ変数を引き継ぎたいときに使用する
        
        Args:
            cell_or_address_value : セットするCell,アドレスを指定する
             （オブジェクトをコピーした後このセル/アドレスがセットされる）
        """
        ret = copy.copy(self)
        if cell_or_address_value!=None:
            ret.set_cell(cell_or_address_value)
        return ret

    def find_entire_row_in_range(self, index_row:int, keyrowd, debug=None):
        """ 
        self.addredd の行を指定してkeywordに合致するCellを取得する

        Memo:
            1行目はindex_row=0とする。
        """
        debug = self._get_debug(debug)
        row, begin_col = self.get_row_and_col_begin()
        row += index_row
        # if index_row==0 and 1 < row:
        #     begin_row -= 1
        begin_row = row - 1
        _, end_col = self.get_row_and_col_end()
        direction = Direction.DOWN + Direction.RIGHT
        cells_2d = self._get_iter_cells_by_r1c1(
            self.sheet,
            begin_row, begin_col, row, end_col,
            direction)
        cell:Cell=None
        for cell_list in cells_2d:
            for cell in cell_list:
                val = _get_cell_value(cell)
                # # print('cell[{}, {}]'.format(cell.row, cell.column))
                if debug:
                    add = get_a1_address_from_row_and_col(cell.row, cell.column)
                    print('cell {} [{}]={}'.format(cell, add, val))
                if keyrowd in val:
                    return cell
        return None

    def get_entire_col_in_range_by_cell(self, cell:Cell):
        """ 
        self.addredd の中にcellと一致するcolがあれば、その列を取得する
        """
        col = cell.column
        begin_row, begin_col = self.get_row_and_col_begin()
        end_row, end_col = self.get_row_and_col_end()
        if not(begin_col <= col  and col <= end_col):
            # col_index = col - begin_col
            return None
        cells_2d = self._get_iter_cells_by_r1c1(
            self.sheet,
            begin_row, col, end_row, col,
            Direction.DOWN & Direction.RIGHT, row_first=False)
        return cells_2d

    # def is_include_address()

    def find_entire_col_in_range(self, index_col:int, keyrowd):
        """ 
        self.addredd の行を指定してkeywordに合致するCellを取得する
            最初に一致したCellのみを返却する
        """
        begin_row, col = self.get_row_and_col_begin()
        # col += index_col
        # 240818
        col = index_col
        end_row, _ = self.get_row_and_col_end()
        cells_2d = self._get_iter_cells_by_r1c1(
            self.sheet,
            begin_row, col, end_row, col,
            Direction.DOWN & Direction.RIGHT)
        cell:Cell=None
        for i, cell_list in enumerate(cells_2d):
            for j, cell in enumerate(cell_list):
                val = _get_cell_value(cell)
                # print('cell[{}, {}]'.format(cell.row, cell.column))
                add = get_a1_address_from_row_and_col(cell.row, cell.column)
                # print('cell[{}]={}'.format(add, val))
                if keyrowd in val:
                    return cell
        return None

    @classmethod
    def _get_iter_cells_by_r1c1(
        cls, worksheet: Worksheet, 
        begin_row, begin_col, end_row, end_col,
        direction:int=Direction.DOWN, row_first:bool=True, debug:bool=None):
        """
        Cellの配列を取得する
         numpyの2次元配列（type=Cell）を取得する
        """
        begin_row, begin_col, end_row, end_col = _align_row_col_with_direction(
            begin_row, begin_col, end_row, end_col, direction)
        # if begin_col == end_col: #_get_for_rangeで対応される
        #     begin_col -= 1
        #     end_col += 1
        # ###
        # if direction & Direction.DOWN:
        #     begin_row -= 1
        #     end_row -= 1
        # else:
        #     begin_row += 1
        #     end_row += 1
        # ###
        # row_range = _get_for_range_b(begin_row, end_row)
        row_range = _get_for_range(begin_row, end_row)
        col_range = _get_for_range(begin_col, end_col) # test13
        # col_range = range(begin_col, end_col)
        # 1列ごとに取得していく（1列の行をすべて取得して、次の列の行を･･･となる）
        if row_first:
            # col_amount = abs(end_col+1 - begin_col)
            col_amount = len(col_range)
            cell_list = np.arange(0)
            for row in row_range:
                row += 1
                cell_list_row = np.zeros(col_amount, dtype=object) #型指定しないと数値となる
                for i, col in enumerate(col_range):
                    cell = worksheet.cell(row=row, column=col)
                    if debug:
                        print('** {}'.format(cell.coordinate))
                    np.put(cell_list_row, [i], cell)
                # data_rowsを cell_list に追加する
                if cell_list.size < 1:
                    cell_list = cell_list_row
                    # cell_list = np.vstack((cell_list, cell_list_row))
                else:
                    cell_list = np.vstack((cell_list, cell_list_row))
        else:
            # row_amount = abs(begin_row+1 - end_row)
            # row_amount = abs(begin_row - end_row)
            row_amount = len(row_range)
            # col_amount = abs(begin_col+1 - end_col)
            cell_list = np.arange(0)
            for col in col_range:
                # col += 1
                cell_list_col = np.zeros(row_amount, dtype=object) #型指定しないと数値となる
                for i, row in enumerate(row_range):
                    cell = Cell(worksheet=worksheet, row=row, column=col)
                    if debug:
                        print('** {}'.format(cell.coordinate))
                    np.put(cell_list_col, [i], cell)
                # data_rowsを cell_list に追加する
                if cell_list.size < 1:
                    cell_list = cell_list_col
                    # cell_list = np.vstack((cell_list, cell_list_col))
                else:
                    cell_list = np.vstack((cell_list, cell_list_col))
        # 1次元と2次元の時がある
        if cell_list.ndim == 1:
            cell_list = [cell_list]
        return cell_list

    def add_sheet(self, add_sheet_name:str):
        names = self.book.get_sheet_names()
        self.book.create_sheet(add_sheet_name)
        names_b = self.book.get_sheet_names()
        ''#debug用

    def get_str_cell_obj(self):
        """ 保持しているself.cell を StrCell class で取得する """
        return StrCell(self.cell)

    def reset_valid_cells(self):
        """
        シートの有効なセルをすべてリセットする(value=None, style='標準')

        Returns : {str} : クリアした範囲のアドレス文字列 ”A1:B2”
        """
        begin_address = self.valid_cells.begin_cell.coordinate
        end_address = self.valid_cells.end_cell.coordinate
        reset_cells(self.sheet, begin_address, end_address)
        return begin_address + ':' + end_address


### End ExcelSheetDataUtil class
##########################################################################################
##########################################################################################
##########################################################################################
##########################################################################################


def main_old():
    filename = 'FileIO.xlsm'
    ex_data = ExcelSheetDataUtil(filename, 'TestData')
    # ex_data.loop_max_col = 100
    # ex_data.loop_max_row = 10000
    begin_address  = ex_data.set_address_by_find('■TableA','A1', 'F38')
    # begin_address  = ex_data.set_address_by_find('■TableB','A1', 'K10')
    # ex_data.address = 'U10'
    print(ex_data.get_value())
    print('begin_address = {}'.format(begin_address))
    begin_address_b = ex_data.move_address(2,0)
    print('begin_address_b = {}'.format(begin_address_b))
    range_address = ex_data.get_range_address()
    # table_begin_address, table_end_address = range_address.split(':')
    print('range_address = {}'.format(range_address))
    ###########
    # cell_values = get_values_from_range_address(ex_data.book, ex_data.sheet, range_address)
    # cell_values = get_values_from_range_address_np(ex_data.book, ex_data.sheet, range_address)
    # cell_values = ex_data.get_values_from_range_address_np()
    cell_values = ex_data.get_values_from_range_address_pd(columns=1)
    print('cell_values = ')
    import pprint
    pprint.pprint(cell_values)
    # pprint.pprint(cell_values.values)
    print('=====')
    df = cell_values[cell_values['Enable'] == '●']
    print(df)
    print('=====')
    item_names = cell_values['ItemName']
    print(type(item_names))
    print(item_names)
    print('# to list')
    buf = list(df['ColA'])
    print(buf)
    print(type(buf))
        

def main():
    filename = 'FileIO.xlsm'
    ex_data = ExcelSheetDataUtil(filename, 'TestData')
    begin_address  = ex_data.set_address_by_find('■TableA','A1', 'F38')
    # begin_address  = ex_data.set_address_by_find('■TableB','A1', 'K10')
    print(ex_data.get_value())
    print('begin_address = {}'.format(begin_address))
    begin_address_b = ex_data.move_address(2,0)
    print('begin_address_b = {}'.format(begin_address_b))
    range_address = ex_data.get_range_address()
    print('range_address = {}'.format(range_address))
    ###########
    cells_df = ex_data.get_values_from_range_address_pd(columns=1)
    print('cell_values = ')
    import pprint
    pprint.pprint(cells_df)
    # pprint.pprint(cell_values.values)
    print('=====')
    df = cells_df[cells_df['Enable']=='●']
    print(df)
    print('=====')
    item_names = cells_df['ItemName']
    print(type(item_names))
    print(item_names)

    print('# to list')
    buf = list(df['ItemName'])
    print(buf)
    print(type(buf))

class CellUtil(ExcelSheetDataUtil):
    pass

if __name__ == '__main__':
    print('# Excute Excel_data.py\n')
    # main()

