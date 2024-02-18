"""
Excelの行列から特定の文字列を検索し、セル番地を返す
"""
# https://ramunememo.hatenablog.com/entry/2021/09/23/182202
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from typing import Union

def _get_cell_value(cell_val:Cell):
    """
    セルのデータを取得（文字列）
    """
    try:
        return str(cell_val.value)
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ（未対応）
        # 何のエラーが発生するか確認してから対応する
        raise e

def _get_cells_address(cell_val:Cell):
    """
    セルの番地を取得
    """
    cell_address = get_column_letter(cell_val.column) +  str(cell_val.row)
    return cell_address


def _search_cell_list(cell_list:'list[Cell]', keyword:str):
    """
    セルのリストの中からキーワードに合致するアドレスを取得する

    get_address_list_to_match_keyword
    """
    result = []
    for cell in cell_list:
        value = _get_cell_value(cell)
        if value == keyword:
            cell_address = _get_cells_address(cell)
            result.append(cell_address)
    return result


# # 特定の列を検索 >>> _search_cell_list
# def search_column(column, keyword):
#     result = []
#     for cell in column:
#         value = _get_cell_value(cell)
#         if value == keyword:
#             cell_address = _get_cells_address(cell)
#             result.append(cell_address)
#     return result

# # 特定の行を検索 >>> _search_cell_list
# def search_row(row, keyword):
#     result = []
#     for cell in row:
#         # セルのデータを文字列に変換
#         try:
#             value = str(cell.value)
#         # 文字列に変換できないデータはスキップ
#         except:
#             continue
#         # キーワードに一致するセルの番地を取得
#         if value == keyword:
#             cell_address = openpyxl.utils.get_column_letter(cell.column) +  str(cell.row)
#             result.append(cell_address)
            
#     return result

def _get_address_a1_str(value:'Union[str, Cell]'):
    """
    A1アドレスを取得する
     文字列の場合はそのまま、Cellの場合はCell.coordinateを取得する
    """
    if isinstance(value, str):
        return value
    elif isinstance(value, Cell):
        return str(value.coordinate)
    else:
        return value


def _get_cells_iterator(
        worksheet_val:Worksheet,
        begin_address:'Union[str, Cell]',
        end_address:'Union[str, Cell]'):
    """ 指定した範囲のCellの Iterator を取得する """
    begin_address = _get_address_a1_str(begin_address)
    end_address = _get_address_a1_str(end_address)
    rectangle = worksheet_val[begin_address:end_address]
    return rectangle

def _get_rectangle(worksheet_val:Worksheet, begin_address:str, end_address:str):
    rectangle = worksheet_val[begin_address:end_address]
    return rectangle

def _get_cells_by_search_keyword_in_rectangle(rectangle:list[list[Cell]] ,keyword:str):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列
    
    Returns:
        {list[openxl.cell.Cell]} : 見つかったCellのリスト
    """
    result_cells = []
    for col in rectangle:
        for cell in col:
            value = _get_cell_value(cell)
            if value == keyword:
                # cell_address = _get_cells_address(cell)
                # result.append(cell_address)
                result_cells.append(cell)
    return result_cells

def _get_rectangle(worksheet_val:Worksheet, begin_address:str, end_address:str):
    """ ワークシートと開始・終了アドレスから対象の範囲の セルリスト list[list[Cell]] を取得する """
    rectangle = worksheet_val[begin_address:end_address]
    return rectangle

def _search_keyword_in_rectangle(rectangle:list[list[Cell]] ,keyword:str):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列
    """
    result = []
    for col in rectangle:
        for cell in col:
            value = _get_cell_value(cell)
            if value == keyword:
                cell_address = _get_cells_address(cell)
                result.append(cell_address)
    return result

# 特定の範囲を検索
def search_rectangle_in_sheet(worksheet_val:Worksheet, begin_address:str, end_address:str ,keyword:str):
    """
    指定したワークシートの、特定の範囲を検索
     (ワークシートと開始・終了アドレスを指定するVer)

    Args
        ws: worksheet object
        range_address: address (ex) 'A1:C5'
    """
    rectangle = _get_rectangle(worksheet_val,begin_address,end_address)
    result = _search_keyword_in_rectangle(rectangle, keyword)
    return result

# 特定の範囲を検索
def search_rectangle(rectangle ,keyword:str):
    """
    特定の範囲を検索（セルのリストは取得済みであること）
     (以前のものと互換性を保つため)
    
    Args
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列
    """
    result = _search_keyword_in_rectangle(rectangle, keyword)
    return result

def search_entire_sheet(worksheet_val:Worksheet, keyword):
    """
    シート全体を検索
    """
    rectangle = worksheet_val.columns
    result = _search_keyword_in_rectangle(rectangle, keyword)
    return result

def _get_cells_of_entire_column_by_search(worksheet_val:Worksheet, keyword):
    """
    シート全体を検索して、マッチした列のlist[Cell]を取得する
    """
    rectangle = worksheet_val.columns
    # result = _search_keyword_in_rectangle(rectangle, keyword)
    result_cell = _get_cells_by_search_keyword_in_rectangle(rectangle, keyword)
    return result_cell

def main():
    # filename = 'FileIO.xlsb'
    """
    例外が発生しました: InvalidFileException
openpyxl does not support binary format .xlsb, please convert this file to .xlsx format if you want to open it with openpyxl
    """
    # filename = 'FileIO.xlsm'
    filename = 'myworkbook.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['TestData']

    result = _search_keyword_in_rectangle(ws['A1':'F38'], '■TableA')
    print(result)

    search_cells = _get_cells_iterator(ws, 'A1', 'F38')
    result_cells = _get_cells_by_search_keyword_in_rectangle(search_cells, '■TableA')
    print(result_cells)

    result = search_entire_sheet(ws, '■TableA')
    print(result)
    
    result_cell = _get_cells_of_entire_column_by_search(ws, '■TableA')
    print(result_cell)

if __name__ == '__main__':
    main()