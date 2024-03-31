"""
シートの有効範囲をすべて削除する
"""


from excel_data import ExcelSheetDataUtil
from pathlib import Path


print('*テーブルにデータ入力')
file_name = 'myworkbook.xlsx'
sheet_name = 'Write'
ex_data = ExcelSheetDataUtil(file_name, sheet_name, data_only=True)

# keyword = '■DF抽出_and_書式コピー表A'
# ex_data.set_address_by_find(keyword, 'A35', 'I50', debug=False,)
# ex_data.move_address(1,0)
begin_address = ex_data.valid_cells.begin_cell.coordinate
print('begin_address = {}'.format(begin_address))
# buf_right_address = ex_data.get_end_address_to_end_horizon(ex_data.Direction.RIGHT)
end_address = ex_data.valid_cells.end_cell.coordinate
print('end_address = {}'.format(end_address))
# buf_bottom_address = ex_data.get_end_address_to_end_vertical(ex_data.Direction.DOWN)
# print('buf_bottom_address = {}'.format(buf_bottom_address))
ex_data.set_range_address([begin_address, end_address])
print('range_address = {}'.format(ex_data.range_address))
range_address_a = ex_data.range_address

# df_a = ex_data.get_values_from_range_address_pd(ex_data.range_address, columns=1)


# print('df_a = ')
# print(df_a.columns)
# print(df_a.values)


# keyword = '■TestData'
# ex_data.set_address_by_find(keyword, debug=False,)
# ex_data.move_address(1,0)
# print('begin_address = {}'.format(ex_data.address))
# buf_right_address = ex_data.get_end_address_to_end_horizon(ex_data.Direction.RIGHT)
# print('buf_right_address = {}'.format(buf_right_address))
# buf_bottom_address = ex_data.get_end_address_to_end_vertical(ex_data.Direction.DOWN)
# print('buf_bottom_address = {}'.format(buf_bottom_address))
# ex_data.set_range_address([buf_right_address, buf_bottom_address])
# print('range_address = {}'.format(ex_data.range_address))

# df_b = ex_data.get_values_from_range_address_pd(ex_data.range_address, columns=1)

# print('df_b = ')
# print(df_b.columns)
# print(df_b.values)

# col_index_b = list(df_b.columns).index('Result')
# col_index_a = list(df_a.columns).index('Result')
# for i, key in enumerate(df_b['ID']):
#     # val = df_b['ID' == key]['Result']
#     # df_buf = df_b.query('ID == "{}"'.format(key))
#     if i==0:
#         df_buf = df_b.loc[df_b.ID == key]['Result']
#         print(type(df_buf))
#         # print(df_buf.info())
#         print(df_buf.shape)
#         print(df_buf.values[0])
#         # val = df_buf['Result'].T
#         pass
#     # val = df_buf.iloc[0,2]
#     # val = df_b.at[key, 'Result'] #xx
#     val = df_b.iloc[i, col_index_b]
#     val = df_b.loc[df_b.ID == key]['Result'].values[0]
#     print('ID[{}] = {}'.format(key, val))

# import pandas as pd
# ####
# # df_aとdf_bをID列に基づいて結合
# merged_df = pd.merge(df_a, df_b[['ID', 'Result']], on='ID', how='left')
# # df_aのResult列をdf_bのResult列で更新
# # df_a['Result'] = merged_df['Result']
# df_a['Result'] = merged_df['Result_y']
# ####

# # 結果を表示
# print('df_a = ')
# print(df_a)

# 書き込みのためにシートを読み直す（data_only=False)
ex_data.reset_book_sheet(data_only=False)



print('*CellのStyleクラスについて')
from openpyxl.cell import Cell
from openpyxl.styles.styleable import StyleableObject
src_cell:Cell = ex_data.sheet['A1']
print('## style = ')
print('* src_cell.style =')
print(src_cell.style)
# 標準
print('* src_cell.fill =')
print(src_cell.fill)
# <openpyxl.styles.fills.PatternFill object>
# patternType=None, fgColor=<openpyxl.styles.colors.Color object>
# rgb='00000000', indexed=None, auto=None, theme=None, tint=0.0, type='rgb', bgColor=<openpyxl.styles.colors.Color object>
# rgb='00000000', indexed=None, auto=None, theme=None, tint=0.0, type='rgb'
from openpyxl.styles.fills import PatternFill
fill = PatternFill()
print('* src_cell.fill.fgColor =')
fgColor = src_cell.fill.fgColor
print(fgColor)
# <openpyxl.styles.colors.Color object>
# rgb='00000000', indexed=None, auto=None, theme=None, tint=0.0, type='rgb'
print('* src_cell.fill.bgColor =')
bgColor = src_cell.fill.bgColor
print(bgColor)
print('* src_cell.number_format =')
print(src_cell.number_format)
print('* src_cell.border =')
print(src_cell.border)
"""
## style = 
* src_cell.style =
標準
* src_cell.fill =
<openpyxl.styles.fills.PatternFill object>
Parameters:
patternType=None, fgColor=<openpyxl.styles.colors.Color object>
Parameters:
rgb='00000000', indexed=None, auto=None, theme=None, tint=0.0, type='rgb', bgColor=<openpyxl.styles.colors.Color object>
Parameters:
rgb='00000000', indexed=None, auto=None, theme=None, tint=0.0, type='rgb'
* src_cell.number_format =
General
* src_cell.border =
<openpyxl.styles.borders.Border object>
Parameters:
outline=True, diagonalUp=False, diagonalDown=False, start=None, end=None, left=<openpyxl.styles.borders.Side object>
Parameters:
style=None, color=None, right=<openpyxl.styles.borders.Side object>
Parameters:
style=None, color=None, top=<openpyxl.styles.borders.Side object>
Parameters:
style=None, color=None, bottom=<openpyxl.styles.borders.Side object>
Parameters:
style=None, color=None, diagonal=<openpyxl.styles.borders.Side object>
Parameters:
style=None, color=None, vertical=None, horizontal=None
##style =
<openpyxl.styles.styleable.StyleableObject object at 0x0000024DB3648790>
"""

# ex_data.copy_value(src_cell, style=True, exists_only=False)
cell_b:Cell = ex_data.sheet['C3']
style = StyleableObject(ex_data.sheet)
print('##style = ')
print(style)
# cell_b.s

# def get_default_cell_format():
#     """ Cellのデフォルト書式設定を取得する """
#     pass

# style = get_default_cell_format
# ex_data.sheet['A1'].style = style


from openpyxl.styles import NamedStyle, Font, Border, Side

def get_default_cell_format():
    """Cellのデフォルト書式設定を取得する"""
    # フォントの設定
    font = Font(name='ＭＳ Ｐゴシック', size=11)
    # 'thin'
    # 'none'xx
    # 境界線の設定
    thin_border = Border(left=Side(style=None), 
                         right=Side(style=None), 
                         top=Side(style=None), 
                         bottom=Side(style=None))
    # NamedStyleオブジェクトの作成
    style = NamedStyle(name="user_style", font=font, border=thin_border)
    return style


print('*セルのスタイルを変更するには、Styleを作成して、bookに名前を登録した後、登録した名前をセットするだけでよい')
# Worksheetは取得済みの想定
# styleの適用
style = get_default_cell_format()
# openpyxlのバージョンによっては、NamedStyleをWorkbookに登録する必要があります
wb = ex_data.book
sheet = ex_data.sheet
if style.name not in wb.named_styles:
    wb.add_named_style(style)
# sheet['C3'].style = style.name
    
print('*Styleオブジェクトを作成して、直接入れてもよい')
# sheet['C3'].border = style.border

print('*セルのスタイルをリセットするには、名前を変えるだけでよい')
sheet['C4'].style = '標準'


# range_begin_address , range_end_address = range_address_a.split(':')
# print('range_address_a = {}'.format(range_address_a))
# ex_data.address = range_address_a
# ex_data.set_cell(range_address_a)
# address_list = ex_data.find_value(
#     ex_data.sheet,
#     keyword='Result',
#     find_begin_address=range_begin_address,
#     find_end_address=range_end_address)
# if len(address_list)<1:
#     msg = 'Result is nothing(sheet={})'.format()
#     raise Exception(msg)
# begin_address = address_list[0]
# begin_cell = ex_data._get_cell(begin_address)
# end_cell = ex_data._get_cell(range_end_address)
# begin_cell_buf = ex_data.find_entire_row_in_range(0, keyrowd='Result', debug=None)
# print('begin_cell_buf = {}'.format(begin_cell_buf))
# # print('******')
# # begin_cell_buf = ex_data.find_entire_col_in_range(0, keyrowd='A005')
# # print(begin_cell_buf)
# # ex_data.reset_book_sheet(data_only=False)
# cells_2d = ex_data.get_entire_col_in_range_by_cell(begin_cell_buf)
# # cells_2d:list[numpy.ndarray]
# print('格納先のCellをリスト[cells_2d]にすべて取得しておいて、そこにdfのfor文で直接Cell.Valueに入れていく')
# from openpyxl.cell import Cell
# cell:Cell=None
# df = df_a['Result']
# for num , row_cells in enumerate(cells_2d):
#     # print('### {}'.format(i))
#     # for (data, cell) in zip(df, row_cells):
#     for j, cell in enumerate(row_cells):
#         # if j==0:
#         #     continue
#         if j!=0:
#             j-=1
#             val = df.iloc[j]
#             print('{}, {} = {}'.format(cell,j, val))
#             cell.value = val
#             row, col = ex_data._cnv_row_col_from_a1_address(cell.coordinate)
#             ex_data.sheet.cell(row, col, val)
#     # for cell in enumerate(row_cells):
#     #     print(cell.coordinate)
#     # for data in df_a['Result']:
#     #     print(data)
# print()


try:
    ex_data.save_book()
except PermissionError as e:
    msg = '[!]ERROR: ファイルアクセスエラーです。ファイルが開いていたらファイルを閉じてください。'
    print(msg)
    # raise e