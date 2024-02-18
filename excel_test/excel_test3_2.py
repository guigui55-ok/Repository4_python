"""
Excelの行列から特定の文字列を検索し、そこから入力連続したセルを縦・横に取得して、矩形アドレスを取得する
クラス化
"""
# https://ramunememo.hatenablog.com/entry/2021/09/23/182202
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
import re
from typing import Union

import excel_test3_2_3 as ex_mod
from excel_test3_2_3 import CellUtil, Direction

def main():
    # filename = 'FileIO.xlsm'
    filename = 'myworkbook.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['TestData']

    cellu = CellUtil(ws)
    # result = search_rectangle_in_sheet(ws, 'A1', 'F38', '■TableA')
    result  = cellu.set_address_by_find('■TableA','A1', 'F38')
    print(result)
    # result = search_entire_sheet(ws, '■TableA')
    result  = cellu.set_address_by_find('■TableA')
    print(result)
    #### 下方向に向かって連続した空白か入力済みセルの最終アドレスを取得する
    sheet = ws
    # begin_address = result[0]
    begin_address = cellu.address
    _MAX = 32767
    begin_row, begin_col = ex_mod.get_row_and_col_from_a1_address(begin_address)
    begin_val = sheet[begin_address]
    begin_is_blank = ex_mod._value_is_blank(begin_val)
    end_row = begin_row
    for offset_row in range(_MAX):
        row = begin_row + offset_row
        now_val = sheet.cell(row=row, column=begin_col).value
        now_is_blank = ex_mod._value_is_blank(now_val)
        # print((now_val, now_is_blank))
        if begin_is_blank == now_is_blank:
            continue
        else:
            end_row = begin_row + offset_row - 1
            
            break
    end_address = ex_mod.get_a1_address_from_row_and_col(end_row, begin_col)
    print(end_address)
    val = sheet[end_address].value
    print(val)
    ###
    print('###')
    #
    # end_address = get_end_address_to_end_vertical(
    #     sheet, begin_address, Direction.TOP, _MAX)
    end_address = cellu.get_end_address_to_end_vertical(Direction.UP)
    print('begin to UP = {}'.format(end_address))
    val = ex_mod.get_cell_value(sheet, end_address)
    print(val)
    #
    # end_address = get_end_address_to_end_vertical(
    #     sheet, begin_address, Direction.BOTTOM, _MAX)
    end_address = cellu.get_end_address_to_end_vertical(Direction.DOWN)
    print('begin to DOWN = {}'.format(end_address))
    val = ex_mod.get_cell_value(sheet, end_address)
    print(val)
    #
    # begin_address_b = get_offset_address(begin_address,2,0)
    begin_address_b = cellu.move_address(2,0)
    table_begin_address = begin_address_b
    #
    # end_address = get_end_address_to_end_horizon(
    #     sheet, begin_address_b, Direction.LEFT, _MAX)
    end_address = cellu.get_end_address_to_end_horizon(Direction.LEFT)
    print('begin to LEFT = {}'.format(end_address))
    val = ex_mod.get_cell_value(sheet, end_address)
    print(val)
    #
    # end_address = get_end_address_to_end_horizon(
    #     sheet, begin_address_b, Direction.RIGHT, _MAX)
    end_address = cellu.get_end_address_to_end_horizon(Direction.RIGHT)
    print('begin to RIGHT = {}'.format(end_address))
    val = ex_mod.get_cell_value(sheet, end_address)
    print(val)
    table_end_address = end_address
    #
    print('***')
    table_address = ex_mod.get_table_address([table_begin_address, table_end_address])
    print('table = {}'.format(table_address))
    # 
    print('=========')
    address = 'A1:B5'
    row, col = ex_mod.get_row_and_col_from_a1_address(address)
    print((address, row, col))

if __name__ == '__main__':
    main()