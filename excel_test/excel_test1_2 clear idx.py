# https://kirinote.com/python-value-clear/

import openpyxl

wb = openpyxl.load_workbook('C:/Users/xxx/Desktop/test/test.xlsx')
ws = wb['Sheet1']

for row in ws.iter_rows():
    for cell in row:
        if cell.col_idx == 2:
            cell.value = None

#別名で保存
wb.save('C:/Users/xxx/Desktop/test/test値クリア.xlsx')


# 列の値をクリア（特定列以降）
# 以下のコードを実行すると、特定の列以降の値をクリアします。

import openpyxl

wb = openpyxl.load_workbook('C:/Users/xxx/Desktop/test/test.xlsx')
ws = wb['Sheet1']

for row in ws.iter_rows(min_col=3):
  for cell in row:
      cell.value = None

#別名で保存
wb.save('C:/Users/xxx/Desktop/test/test値クリア.xlsx')