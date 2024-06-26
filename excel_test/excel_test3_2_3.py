"""
Excelの行列から特定の文字列を検索し、そこから入力連続したセルを縦・横に取得して、矩形アドレスを取得する
クラス化
"""
# https://ramunememo.hatenablog.com/entry/2021/09/23/182202
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
import re
import copy
from datetime import datetime, timedelta
import numpy as np
from typing import Union
from pathlib import Path

_DEBUG = False
# https://office-hack.com/excel/maximum-number-of-lines/
"""
Excel 2007
最大行数	最大列数
1,048,576	16,384
Excel 2010
最大行数	最大列数
1,048,576	16,384
Excel 2013
最大行数	最大列数
1,048,576	16,384
Excel 2016
最大行数	最大列数
1,048,576	16,384
Excel 2019
最大行数	最大列数
1,048,576	16,384
"""

_MAX = 32767
_EXCEL_ROW_MAX = 1048576
_EXCEL_COLUMN_MAX = 16384
def get_cell_value(sheet:Worksheet, value:'Union[str,Cell]'):
    """
    セルのデータを取得（文字列）
    """
    try:
        if isinstance(value, Cell):
            # val = value.value
            # valueがsheet=Noneでも取得できるようにしている
            val = sheet[value.coordinate]
        else:
            val = sheet[value].value
        if val == None:
            val = ''
        return val
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        # address==Noneの時は TypeError が発生する
        # TypeError: expected string or bytes-like object
        raise e

def get_cell_value_r1c1(sheet:Worksheet, row:int, col:int):
    """
    セルのデータを取得（R1C1形式）
    """
    try:
        val = sheet.cell(row=row, column=col).value
        if val == None:
            val = ''
        return val
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        raise e

def set_cell_value(sheet:Worksheet, address, value):
    """
    セルのデータを入力（文字列）
    """
    try:
        sheet[address].value = value
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        # address==Noneの時は TypeError が発生する
        # TypeError: expected string or bytes-like object
        raise e

def set_cell_value_r1c1(sheet:Worksheet, row:int, col:int, value):
    """
    セルのデータを入力（R1C1形式）
    """
    try:
        sheet.cell(row=row, column=col).value = value
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        raise e

def _get_cell_value(cell_val:Cell):
    """
    セルのデータを取得（文字列）

    Memo:
        セルの値が空のときはNoneが返る
    """
    try:
        return str(cell_val.value)
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        raise e

def _get_address(value:'Union[str,Cell]'):
    """
    セルの番地をA1形式で取得する
    Args:
        value {str,Cell}
    """
    return _get_cell_address(value)

def _get_cell_address(value:'Union[str,Cell]'):
    """
    セルの番地をA1形式で取得する
    Args:
        value {str,Cell}
    """
    if isinstance(value, Cell):
        # cell_address = get_column_letter(cell_val.column) +  str(cell_val.row)
        address_str = value.coordinate
    elif isinstance(value, str):
        if re.match(r'[a-zA-Z]{1,3}[1-9]{1,8}', value)!=None:
            address_str = value
        else:
            raise ValueError('value is not Address')
    else:
        raise TypeError(type(value))
    return address_str

def _is_match_patterns(patterns:Union[str, list[str]], value:str):
    if not isinstance(patterns, list):
        patterns = [str(patterns)]
    for pattern in patterns:
        ret = re.search(pattern, value)
        if ret!=None:
            return True
    return False

def _is_include_address(cell_a:'Union[Cell, str]', cell_b:'Union[Cell,str]'):
    """
    cell_a と cell_b の範囲が重なっている部分があるか判定する

    Returns: {bool}
    """
    a_begin_cell, a_end_cell = get_row_and_col_from_a1_address_range(
        cell_a)
    b_begin_cell, b_end_cell = get_row_and_col_from_a1_address_range(
        cell_b)
    if a_begin_cell.row <= b_begin_cell.row <= a_end_cell.row\
        or a_begin_cell.row <= b_end_cell.row <= a_end_cell.row:
        if a_begin_cell.column <= b_begin_cell.column <= a_end_cell.column\
            or a_begin_cell.column <= b_end_cell.column <= a_end_cell.column:
            return True
    return False


def get_row_and_col_from_a1_address_range(cell_address):
    """
    A1形式のアドレスを行と列に分割する 範囲用 （例:'A1:B2'）

    Memo:
      未対応(対応予定) 'A1:B2, C3, D4:E5:F6'
        ['A1:B2', 'C3', 'D4:E5:F6']
    
    Returns:
        list[SimpleCellInfo, SimpleCellInfo]  :  
          list[begin_cell, bend_cell]
    """
    cell_address = _get_cell_address(cell_address)
    if ':' in cell_address:
        buf = cell_address.split(':')
        begin_address = buf[0]
        end_address = buf[1]
    else:
        begin_address = cell_address
        end_address = cell_address
    row, col = get_row_and_col_from_a1_address(begin_address)
    # begin_cell = SimpleCellInfo()
    begin_cell = Cell(worksheet=None, row=row, column=col)
    # openxl.~.Cell は インスタンス引数にworksheetがあるので、上記クラスで扱う
    row, col = get_row_and_col_from_a1_address(end_address)
    end_cell = Cell(worksheet=None, row=row, column=col)
    return begin_cell, end_cell

def _is_a1_address(value):
    ret = re.search(r'[a-zA-Z]{1,3}[0-9]{1,8}', value)
    if ret!=None:
        return True
    else:
        return False

def _is_single_cell(cell_address:str):
    cell_address = _get_cell_address(cell_address)
    if not ':' in cell_address:
        if not ',' in cell_address:
            return True
    
    row, col = get_row_and_col_from_a1_address(cell_address)
    if row==1 and col==1:
        return True
    else:
        return False

def _is_match_patterns(patterns:Union[str, list[str]], value:str):
    if not isinstance(patterns, list):
        patterns = [str(patterns)]
    for pattern in patterns:
        ret = re.search(pattern, value)
        if ret!=None:
            return True
    return False


def _search_cell_list(cell_list:'list[Cell]', keyword:str):
    """
    セルのリストの中からキーワードに合致するアドレスを取得する

    get_address_list_to_match_keyword
    """
    result = []
    for cell in cell_list:
        value = _get_cell_value(cell)
        if value == keyword:
            cell_address = _get_cell_address(cell)
            result.append(cell_address)
    return result

def _get_address_a1_str(value:'Union[str, Cell]'):
    """
    A1アドレスを取得する
     文字列の場合はそのまま、Cellの場合はCell.coordinateを取得する
    """
    if isinstance(value, str):
        return value
    elif isinstance(value, Cell):
        return str(value.coordinate)
    else:
        return value

def _get_cells_iterator(
        worksheet_val:Worksheet,
        begin_address:'Union[str, Cell]',
        end_address:'Union[str, Cell]'):
    """ 指定した範囲のCellの Iterator を取得する """
    begin_address = _get_address_a1_str(begin_address)
    end_address = _get_address_a1_str(end_address)
    rectangle = worksheet_val[begin_address:end_address]
    return rectangle

# def _get_rectangle(worksheet_val:Worksheet, begin_address:str, end_address:str):
#     """
#     ワークシートと開始・終了アドレスから対象の範囲の セルリスト list[list[Cell]] を取得する
#      WorkShhet['A1:B5']とすればよいので削除予定 
#     """
#     rectangle = worksheet_val[begin_address:end_address]
#     return rectangle

def _get_rectangle(worksheet_val:Worksheet, begin_address:'Union[str,Cell]', end_address:'Union[str,Cell]'):
    """
    ワークシートと開始・終了アドレスから対象の範囲の セルリスト list[list[Cell]] を取得する
     str, Cell に対応
    """
    begin_address = _get_cell_address(begin_address)
    end_address = _get_cell_address(end_address)
    rectangle = worksheet_val[begin_address:end_address]
    return rectangle

def _get_cells_by_search_keyword_in_rectangle(rectangle:list[list[Cell]] ,keyword:str):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列
    
    Returns:
        {list[openxl.cell.Cell]} : 見つかったCellのリスト
    """
    result_cells = []
    for col in rectangle:
        for cell in col:
            value = _get_cell_value(cell)
            if value == keyword:
                # cell_address = _get_cells_address(cell)
                # result.append(cell_address)
                result_cells.append(cell)
    return result_cells

def _search_keyword_in_rectangle(
        rectangle:list[list[Cell]] ,
        pattern:Union[str, list[str]],
        debug:bool=False):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列、または、パターン
    
    Returns:
        list[str] : A1形式のアドレス文字列のリスト
    """
    result = []
    cell:Cell=None
    for col in rectangle:
        for cell in col:
            value = _get_cell_value(cell)
            if debug:
                print((cell.row, cell.column, value))
            if _is_match_patterns(pattern, value):
                cell_address = _get_cell_address(cell)
                result.append(cell_address)
    return result

# def _get_cells_iterator(
#         worksheet_val:Worksheet,
#         begin_address:'Union[str, Cell]',
#         end_address:'Union[str, Cell]'):
#     """ 指定した範囲のCellの Iterator を取得する """
#     begin_address = _get_address_a1_str(begin_address)
#     end_address = _get_address_a1_str(end_address)
#     rectangle = worksheet_val[begin_address:end_address]
#     return rectangle


def _get_cells_by_search_keyword_in_rectangle(rectangle:list[list[Cell]] ,keyword:str):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列
    
    Returns:
        {list[openxl.cell.Cell]} : 見つかったCellのリスト
    """
    result_cells = []
    for col in rectangle:
        for cell in col:
            value = _get_cell_value(cell)
            if value == keyword:
                # cell_address = _get_cells_address(cell)
                # result.append(cell_address)
                result_cells.append(cell)
    return result_cells

def search_rectangle_in_sheet(
    worksheet_val:Worksheet,
    begin_address:str,
    end_address:str,
    keyword:Union[str, list[str]],
    debug:bool=False):
    """
    指定したワークシートの、特定の範囲を検索
     (ワークシートと開始・終了アドレスを指定するVer)

    Args
        ws: worksheet object
        range_address: address (ex) 'A1:C5'
    """
    rectangle = _get_rectangle(worksheet_val,begin_address,end_address)
    result = _search_keyword_in_rectangle(rectangle, keyword, debug)
    return result

def search_entire_sheet(worksheet_val:Worksheet, keyword):
    """
    シート全体を検索する

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列、または、パターン
    """
    rectangle = worksheet_val.columns
    result = _search_keyword_in_rectangle(rectangle, keyword)
    return result

def _get_cells_of_entire_column_by_search(worksheet:Worksheet, keyword):
    """
    シート全体を検索して、マッチした列のlist[Cell]を取得する
    """
    rectangle = worksheet.columns
    # result = _search_keyword_in_rectangle(rectangle, keyword)
    result_cell = _get_cells_by_search_keyword_in_rectangle(rectangle, keyword)
    return result_cell

#########
def get_a1_address_from_cell(cell:Cell):
    """ A1形式のアドレスを取得する """
    return cell.coordinate

def get_row_and_col_from_a1_address(address:Union[str, Cell]):
    """
    A1形式のアドレスを行と列に分割する

    Memo:
        'A1:B5' は A1(row=1,col=1) を返す
    Returns:
        int, int  :  row_number, column_number
    """
    address_ = address
    if isinstance(address_, Cell):
        address_ = address_.coordinate
    if ':' in address_:
        address_ = address_[:address_.find(':')]
    # Memo 'A1:B5'は'AB15'と解釈される
    column_letter = ''.join(filter(str.isalpha, address_))
    row_number = ''.join(filter(str.isdigit, address_))
    # 列のアルファベットを数値に変換
    column_number = column_index_from_string(column_letter)
    # row, colを返す
    return int(row_number), int(column_number)

def get_a1_address_from_row_and_col(row:int ,col:int):
    """ R1C1形式のアドレスを行と列に変換する """
    # 列番号をアルファベットに変換
    column_letter = get_column_letter(col)
    # A1形式でアドレスを組み立て
    return f"{column_letter}{row}"

def get_cells(sheet:Worksheet, *args):
    """
    WorkSheetからCellを取得する

    Args:
        *args:
            第1引数がstrならA1形式
            第1引数、第2引数がintならR1C1形式
    """
    # 引数の最初が数値ならR1C1形式を採用、
    # この場合2つ目も読み取り、それぞれをrow,colとして扱う
    if isinstance(args[0], int):
        row = args[0]
        col = args[1]
        cell = sheet.cell(row=row, column=col)
    elif isinstance(args[0], str):
        if args[0].isdigit():
            row = args[0]
            col = args[1]
            cell = sheet.cell(row=row, column=col)
        else:
            a1 = args[0]
            cell:Cell = sheet[a1]
    else:
        cell = None
    return cell

def get_offset_cell(sheet:Worksheet, cell:Cell, row, col):
    """ 特定のセルからrow, col を移動したセルを取得する """
    now_row, now_col = get_row_and_col_from_a1_address(cell)
    new_row = now_row + row
    new_column = now_col + col
    return sheet.cell(row=new_row, column=new_column)

def get_offset_address(address:str, row, col):
    """
    特定のセルaddressからrow, col を移動したセルを取得する
     (Sheetは必要ない)
    """
    now_row, now_col = get_row_and_col_from_a1_address(address)
    new_row = now_row + row
    new_column = now_col + col
    ret_address = get_a1_address_from_row_and_col(new_row, new_column)
    return ret_address

def _value_is_blank(value):
    """ cellの値がNoneであるか判定する """
    if isinstance(value, Cell):
        value = value.value
    if value == '':
        is_blank = True
    else:
        is_blank = False
    if value == None:
        is_blank = True
    return is_blank

######################################################################
######################################################################
class SimpleCellInfo():
    """ セルの値を扱う（簡単な処理をするときに使用する） """
    def __init__(self) -> None:
        self.row = 0
        self.col = 0
        self.address = ''


class TwoCellsInfo():
    """ 2つのセルの値を扱う（主にBegin,Endの矩形アドレスを扱う想定） """
    def __init__(self) -> None:
        self.begin_cell = SimpleCellInfo()
        self.end_cell = SimpleCellInfo()
    
    def set_row_col(self, min_row, min_col, max_row, max_col):
        self.begin_cell.row = min_row
        self.begin_cell.col = min_col
        self.end_cell.row = max_row
        self.end_cell.col = max_col


class TwoCellsInfo():
    """ 2つのセルの値を扱う（主にBegin,Endの矩形アドレスを扱う想定） """
    def __init__(self) -> None:
        self.begin_cell = SimpleCellInfo()
        self.end_cell = SimpleCellInfo()
    
    def set_row_col(self, min_row, min_col, max_row, max_col):
        self.begin_cell.row = min_row
        self.begin_cell.col = min_col
        self.end_cell.row = max_row
        self.end_cell.col = max_col

class ConstExcel():
    ROW = 1
    COLUMN = 2
    COL = COLUMN
    STYLE_NORMAL = '標準'

class Direction():
    LEFT = 1 << 0  # 0000000001 in binary
    TOP = 1 << 1  # 0000000010 in binary
    UP = 1 << 1  # 0000000010 in binary #TOP==UP
    RIGHT = 1 << 2  # 0000000100 in binary
    BOTTOM = 1 << 3 # 0000001000 in binary
    DOWN = 1 << 3 # 0000001000 in binary #BOTTOM==DOWN
    @classmethod
    def _print_direction(cls, direction_flags):
        if Direction.RIGHT & direction_flags:
            if _DEBUG:
                print('flag = Direction.RIGHT', end=' | ')
        else:
            if _DEBUG:
                print('flag = Direction.LEFT', end=' | ')
        if Direction.DOWN & direction_flags:
            if _DEBUG:
                print('flag = Direction.DOWN')
        else:
            if _DEBUG:
                print('flag = Direction.UP')

def _set_dirction(direction):
    """ 方向をセットする、何もない場合はRIGHT,BOTTOMとする """
    horizon = None
    if Direction.RIGHT & direction:
        horizon = Direction.RIGHT
    else:
        horizon = Direction.LEFT
    vertical = None
    if Direction.BOTTOM & direction:
        vertical = Direction.BOTTOM
    else:
        vertical = Direction.TOP
    return horizon | vertical

def _set_begin_col(begin_address, direction, loop_max_col):
    # target_row, target_col = get_row_and_col_from_a1_address(begin_address)
    direction = _set_dirction(direction)
    if Direction.LEFT & direction:
        end_col = 1
    else:
        #Direction.RIGHT & direction:
        end_col = loop_max_col
    return end_col

def _set_begin_row(begin_address, direction, loop_max_row):
    # target_row, target_col = get_row_and_col_from_a1_address(begin_address)
    if Direction.TOP & direction:
        end_row = 1
    else:
        #Direction.BOTTOM & direction:
        end_row = loop_max_row
    return end_row

def _get_for_range(begin_num, end_num):
    """
    get_end_address_to_end_horizon を実行するときの、rangeオブジェクトを取得する
      (end_num - begin_num) < 0 の時はStep-1としている
    """
    if begin_num <= end_num:
        return range(begin_num, end_num + 1)
        # return range(begin_num-1, end_num)
    else:
        # return reversed(range(begin_num, end_num))
        return range(begin_num, end_num, -1)

def _get_for_range_b(begin_num, end_num):
    """
    get_end_address_to_end_horizon を実行するときの、rangeオブジェクトを取得する
      (end_num - begin_num) < 0 の時はStep-1としている
    """
    if begin_num <= end_num:
        # return range(begin_num, end_num + 1)
        return range(begin_num-1, end_num)
    else:
        # return reversed(range(begin_num, end_num))
        return range(begin_num, end_num, -1)

def _get_last_add_value(direction):
    """
    get_end_address_to_end_horizon,get_end_address_to_end_vertical
     を実行するときの、最後に追加する数値を取得する
      direction によって値が変更される
    """
    if Direction.LEFT & direction:
        add_col = +1
    else:
        #Direction.RIGHT & direction:
        # add_col = -1
        add_col = 0
    if Direction.BOTTOM & direction:
        add_row = -1
    else:
        #Direction.UP & direction:
        add_row = +1
    return add_row, add_col

def get_end_address_to_end_vertical(
        sheet:Worksheet,
        begin_address:str,
        direction,
        loop_max=_EXCEL_ROW_MAX,
        invalid_all_blank:bool=True):
    """
    連続した空白or入力値のセルの最終のアドレスを取得する 縦方向-垂直

    Args:
        invalid_all_blank: すべて空白の時はカウントを無効として、開始アドレスを返す
    """
    begin_row, begin_col = get_row_and_col_from_a1_address(begin_address)
    begin_val = sheet[begin_address]
    begin_is_blank = _value_is_blank(begin_val)
    direction_ = _set_dirction(direction)
    Direction._print_direction(direction)
    end_row = _set_begin_row(
        begin_address, direction_, loop_max)
    last_add_row, last_add_col = _get_last_add_value(direction)
    # for_range = _get_for_range(begin_row, end_row+1)
    for_range = _get_for_range(begin_row, end_row)
    for row in for_range:
        now_val = sheet.cell(row=row, column=begin_col).value
        now_is_blank = _value_is_blank(now_val)
        # print((now_val, now_is_blank, get_a1_address_from_row_and_col(row, begin_col)))
        if begin_is_blank == now_is_blank:
            continue
        else:
            break
    else:
        pass
    # loopの最終行まで読み込んで、開始セルと終了せるが空白の時は、すべて空白となる。
    # 値がないときはフラグによってカウントを無効として、開始セルを返す
    if invalid_all_blank:
        if (end_row-1) <= row:
            if begin_is_blank and now_is_blank:
                return begin_address
    if _EXCEL_ROW_MAX <= row:
        row = _EXCEL_ROW_MAX
    else:
        # now_is_blank フラグが変更された前のセルが対象
        last_add_row, last_add_col = _get_last_add_value(direction)
        row += last_add_row
    end_address = get_a1_address_from_row_and_col(row, begin_col)
    return end_address

def get_end_cell_to_end_vertical(sheet:Worksheet, begin_address:'Union[str,Cell]', direction, loop_max=_MAX):
    """ 連続した空白or入力値のセルの最終のCellを取得する 縦方向-垂直"""
    end_address = get_end_address_to_end_vertical(
        sheet, begin_address, direction, loop_max)
    end_cell = get_cells(sheet, end_address)
    return end_cell

def get_end_address_to_end_horizon(
        sheet:Worksheet,
        begin_address:str,
        direction,
        loop_max=_EXCEL_COLUMN_MAX,
        invalid_all_blank:bool=True):
    """
    連続した空白or入力値のセルの最終のアドレスを取得する 横方向-水平

    Args:
        invalid_all_blank: すべて空白の時はカウントを無効として、開始アドレスを返す
    """
    begin_row, begin_col = get_row_and_col_from_a1_address(begin_address)
    begin_val = sheet[begin_address]
    begin_is_blank = _value_is_blank(begin_val)
    direction_ = _set_dirction(direction)
    end_col = _set_begin_col(
        begin_address, direction_, loop_max)
    last_add_row, last_add_col = _get_last_add_value(direction)
    for_range = _get_for_range(begin_col, end_col)
    for col in for_range:
        now_val = sheet.cell(row=begin_row, column=col).value
        now_is_blank = _value_is_blank(now_val)
        # print((now_val, now_is_blank, get_a1_address_from_row_and_col(begin_row, col)))
        if begin_is_blank == now_is_blank:
            continue
        else:
            break
    # loopの最終行まで読み込んで、開始セルと終了せるが空白の時は、すべて空白となる。
    # 値がないときはフラグによってカウントを無効として、開始セルを返す
    if invalid_all_blank:
        if (end_col-1) <= col:
            if begin_is_blank and now_is_blank:
                return begin_address
    if _EXCEL_COLUMN_MAX <= col:
        col = _EXCEL_COLUMN_MAX
    else:
        # now_is_blank フラグが変更された前のセルが対象
        last_add_row, last_add_col = _get_last_add_value(direction)
        col += last_add_col
        pass
    end_address = get_a1_address_from_row_and_col(begin_row, col)
    return end_address

def get_end_cell_to_end_horizon(sheet:Worksheet, begin_address:'Union[str,Cell]', direction, loop_max=_MAX):
    """ 連続した空白or入力値のセルの最終のCellを取得する 横方向-水平"""
    end_address = get_end_address_to_end_horizon(
        sheet, begin_address, direction, loop_max)
    end_cell = get_cells(sheet, end_address)
    return end_cell

def get_range_cell(sheet:Worksheet, value_list:'list[Union[str,Cell]]')->Cell:
    """ 複数のA1形式アドレスorCellのリストから矩形のCellを取得する """
    range_address = get_table_address(value_list)
    return get_cells(sheet, range_address)

def get_table_address(value_list:'list[Union[str,Cell]]'):
    """
    複数のA1形式アドレスのリストから矩形のアドレスを取得する

    Returns:
        {str} : address 'A1:B2'
    """
    min_row, min_col, max_row, max_col = None, None, None, None
    for value in value_list:
        address = _get_cell_address(value)
        row, col = get_row_and_col_from_a1_address(address)
        if min_row==None:
            min_row, min_col = row, col
            max_row, max_col = row, col
            continue
        min_row, max_row = _update_min_max_by_compare(min_row, max_row, row)
        min_col, max_col = _update_min_max_by_compare(min_col, max_col, col)
    begin_address = get_a1_address_from_row_and_col(min_row, min_col)
    end_address = get_a1_address_from_row_and_col(max_row, max_col)
    return begin_address + ':' + end_address

def get_begin_and_end_cell_in_range(sheet:Worksheet, value_list:'list[Union[str,Cell]]'):
    """ 複数のA1形式アドレスのリストから矩形のアドレスを探して、そのbeginとendのCell取得する """
    begin_address, end_address = get_begin_and_end_address_in_range(value_list)
    begin_cell = get_cells(sheet, begin_address)
    end_cell = get_cells(sheet, end_address)
    return begin_cell, end_cell

def get_begin_and_end_address_in_range(value_list:'list[Union[str,Cell]]'):
    """ 複数のA1形式アドレスのリストから矩形のアドレスを探して、そのbeginとendをアドレスを文字列で取得する """
    min_row, min_col, max_row, max_col = None, None, None, None
    for value in value_list:
        address = _get_cell_address(value)
        row, col = get_row_and_col_from_a1_address(address)
        if min_row==None:
            min_row, min_col = row, col
            max_row, max_col = row, col
            continue
        min_row, max_row = _update_min_max_by_compare(min_row, max_row, row)
        min_col, max_col = _update_min_max_by_compare(min_col, max_col, col)
    begin_address = get_a1_address_from_row_and_col(min_row, min_col)
    end_address = get_a1_address_from_row_and_col(max_row, max_col)
    return begin_address, end_address

def get_range_address(address_list:'list[str]'):
    """
    複数のA1形式アドレスのリストから矩形のアドレスの始点と終点（左上と右下）を取得する
     Deprecated
    
    Returns:
        begin_address, end_address
    """
    min_row, min_col, max_row, max_col = None, None, None, None
    for address in address_list:
        row, col = get_row_and_col_from_a1_address(address)
        if min_row==None:
            min_row, min_col = row, col
            max_row, max_col = row, col
            continue
        min_row, max_row = _update_min_max_by_compare(min_row, max_row, row)
        min_col, max_col = _update_min_max_by_compare(min_col, max_col, col)
    begin_address = get_a1_address_from_row_and_col(min_row, min_col)
    end_address = get_a1_address_from_row_and_col(max_row, max_col)
    return begin_address, end_address

def _update_min_max_by_compare(min, max, value):
    """ value が min より小さければminを更新、maxより大きければmaxを更新して返す """
    if value < min:
        min = value
    if max < value:
        max = value
    return min, max


def _get_cell_from_sheet(sheet:Worksheet, a1_address:str)->Cell:
    """
    SheetからCellを取得する
     cell = sheet['A1']とすると cell.coordinate がVsCodeでリンクされないため、この関数で取得する
    """
    return sheet[a1_address]


def get_min_row_and_col(value_list:'Union[Cell,str]', max_row=_EXCEL_ROW_MAX, max_col=_EXCEL_COLUMN_MAX):
    min_row , min_col = max_row, max_col
    for value in value_list:
        row, col = get_row_and_col_from_a1_address(value)
        if row < min_row:
            min_row = row
        if col < min_col:
            min_col = col
    return min_row, min_col

def get_min_row_and_col_2d(value_list:'Union[Cell,str]', max_row=_EXCEL_ROW_MAX, max_col=_EXCEL_COLUMN_MAX):
    min_row , min_col = max_row, max_col
    for value_list_a in value_list:
        for value in value_list_a:
            row, col = get_row_and_col_from_a1_address(value)
            if row < min_row:
                min_row = row
            if col < min_col:
                min_col = col
    return min_row, min_col

def get_min_cell(sheet:Worksheet, value_list:'Union[Cell,str]', max_row=0, max_col=0):
    if max_row<1:
        max_row = sheet.max_row
    if max_col<1:
        max_col = sheet.max_column
    min_row, min_col = get_min_row_and_col(value_list, max_row, max_col)
    return Cell(sheet, min_row, min_col)

def get_min_cell_2d(sheet:Worksheet, value_list:'Union[Cell,str]', max_row=0, max_col=0):
    if max_row<1:
        max_row = sheet.max_row
    if max_col<1:
        max_col = sheet.max_column
    min_row, min_col = get_min_row_and_col_2d(value_list, max_row, max_col)
    return Cell(sheet, min_row, min_col)

def get_max_row_and_col(value_list:'Union[Cell,str]', min_row=1, min_col=1):
    max_row , max_col = min_row, min_col
    for value in value_list:
        row, col = get_row_and_col_from_a1_address(value)
        if max_row < row:
            max_row = row
        if max_col < col:
            max_col = col
    return max_row, max_col

def get_max_row_and_col_2d(value_list:'Union[Cell,str]', min_row=1, min_col=1):
    max_row , max_col = min_row, min_col
    for value_list_child in value_list:
        for value in value_list_child:
            row, col = get_row_and_col_from_a1_address(value)
            if max_row < row:
                max_row = row
            if max_col < col:
                max_col = col
    return max_row, max_col

def get_max_cell(sheet:Worksheet, value_list:'Union[Cell,str]', min_row=1, min_col=1):
    max_row, max_col = get_max_row_and_col(value_list, sheet.min_row, sheet.min_column)
    return Cell(sheet, max_row, max_col)

def get_max_cell_2d(sheet:Worksheet, value_list:'Union[Cell,str]', min_row=1, min_col=1):
    max_row, max_col = get_max_row_and_col_2d(value_list, sheet.min_row, sheet.min_column)
    return Cell(sheet, max_row, max_col)


def _align_row_col(begin_row, begin_col, end_row, end_col):
    """
    値の順番をそろえる
    
    end_row < begin_row などとなっているとループ処理の時に意図しない動作となるため
    """
    if end_row < begin_row:
        temp = end_row
        end_row = begin_row
        begin_row = temp
    if end_col < begin_col:
        temp = end_col
        end_col = begin_col
        begin_col = temp
    return begin_row, begin_col, end_row, end_col

def _align_row_col_with_direction(begin_row, begin_col, end_row, end_col, direction):
    """
    値の順番をそろえる
     方向（Driection）も考慮する
    
    end_row < begin_row などとなっているとループ処理の時に意図しない動作となるため
     Direction.UPだと end_col > begin_col となる
    Direction は 縦方向は(UP or DOWN)横方向は(RIGHT or LEFT)どちらかが選択される
     何も指定がない場合は（RIGHT,DOWN）がデフォルト
    """
    begin_row, begin_col, end_row, end_col = _align_row_col(
        begin_row, begin_col, end_row, end_col)
    # Direction
    # RIGHTはそのまま
    if Direction.LEFT & direction:
        temp = begin_row
        begin_row = end_row
        end_row = temp
    # DOWNはそのまま
    if Direction.UP & direction:
        temp = begin_col
        begin_col = end_col
        end_col = temp
    return begin_row, begin_col, end_row, end_col
        
def get_row_and_col_from_rect_address(table_address:str):
    """
    矩形のアドレスから、始点と終点（左上と右下）のアドレスを数値で取得する
      begin_row, begin_col, end_row, end_col の4つの値を取得する
    
    Note:
        セルからFor文で値を取得する時に使用する用
         'A1:B2' 形式の文字列を想定。（コロンで区切られたアルファベット＋数値）
    Returns:
        int, int, int, int:
         begin_row, begin_col, end_row, end_col
    """
    ex_data = ExcelSheetDataUtil(None, None)
    add_list = table_address.split(':')
    begin_address, end_address = get_range_address(add_list)
    begin_row, begin_col = ExcelSheetDataUtil._cnv_row_col_from_a1_address(begin_address)
    end_row, end_col = ExcelSheetDataUtil._cnv_row_col_from_a1_address(end_address)
    return begin_row, begin_col, end_row ,end_col

def get_values_from_range_address(book:Workbook, sheet:Worksheet, range_address:str):
    """
    矩形のセルからすべて値を取得する

    Args:
        table_address : 'A1:B2'
    """
    ex_data = ExcelSheetDataUtil(None, None)
    ex_data.set_workbook(book)
    ex_data.set_sheet(sheet)
    ex_data.range_address = range_address
    row, col = get_row_and_col_from_a1_address(range_address)
    # print('range_address info = '.format((range_address, row, col)))
    begin_row, begin_col, end_row, end_col = get_row_and_col_from_rect_address(range_address)
    cell_values_list = []
    for row in range(begin_row, end_row+1):
        buf_values = []
        for col in range(begin_col, end_col+1):
            value = ex_data.get_value_r1c1(row, col)
            buf_values.append(value)
        cell_values_list.append(buf_values)
    return cell_values_list


def get_values_from_range_address_np(book:Workbook, sheet:Worksheet, range_address:str):
    """
    矩形のセルからすべて値を取得する

    Args:
        range_address : 'A1:B2'
    """
    import numpy as np
    ex_data = ExcelSheetDataUtil(None, None)
    ex_data.set_workbook(book)
    ex_data.set_sheet(sheet)
    ex_data.range_address = range_address
    row, col = get_row_and_col_from_a1_address(range_address)
    # print('range_address info = '.format((range_address, row, col)))
    begin_row, begin_col, end_row, end_col = get_row_and_col_from_rect_address(range_address)

    col_amount = abs(end_col+1 - begin_col)
    cell_values_list = np.arange(0)
    for row in range(begin_row-1, end_row):
        row += 1
        data_rows = np.zeros(col_amount, dtype=object) #型指定しないと数値となる
        for i, col in enumerate(range(begin_col, end_col+1)):
            value = ex_data.get_value_r1c1(row, col)
            np.put(data_rows, [i], value)
        # data_rowsを cell_values_list に追加する
        if cell_values_list.size < 1:
            cell_values_list = data_rows
        else:
            cell_values_list = np.vstack((cell_values_list, data_rows))
    return cell_values_list


################################################################################
################################################################################
################################################################################
################################################################################
################################################################################

class ExcelSheetDataUtil():
    """
    エクセルのデータを扱うクラス
     主にCellのアドレス・値を扱う
      （1つのシートのみを対象とする）
    
    use modules:
        主に使用しているクラス・モジュールなど
        import openpyxl
        from openpyxl.cell import Cell
        from openpyxl.worksheet.worksheet import Worksheet
        from openpyxl.workbook.workbook import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.utils import column_index_from_string
    """
    class Direction(Direction):
        pass

    class ConstExcel(ConstExcel):
        pass

    def __init__(self, file_path:str, sheet_name:str, data_only:bool=True) -> None:
        self.__init_param(file_path, sheet_name, data_only)
    
    def __init_param(self, file_path:str, sheet_name:str, data_only:bool):
        self.set_workbook(file_path, data_only)
        self.set_sheet(sheet_name)
        self._update_valid_cell_in_sheet()
        self.address = 'A1'
        self.range_address = 'A1'
        # self.loop_max_row = _EXCEL_ROW_MAX
        # self.loop_max_col = _EXCEL_COLUMN_MAX
        # 有効なセル最大までにすると、すべて処理したときに時間がかかるので
        # 初期値Noneとして、sheet.min_rowなどの値を使用する
        self.loop_max_row = None
        self.loop_max_col = None
    
    def copy_values_from_other_cell(self, other_cell:'ExcelSheetDataUtil'):
        """
        他の ExcelSheetDataUtil から アドレスとCellのみをコピーする
         （book,sheetの値はコピーされない）
        """
        self.address = other_cell.address
        self.cell = other_cell.cell
        self.range_address = other_cell.range_address
    
    def reset_book_sheet(self, data_only:bool=True):
        self.close()
        sheet_name = self.sheet.title
        self.set_workbook(self.file_path, data_only)
        self.set_sheet(sheet_name)

    def _update_valid_cell_in_sheet(self):
        """ WorkSheetの有効なセルの値を更新する """
        self.valid_cells = TwoCellsInfo()
        if self.sheet==None:
            return
        self.valid_cells.set_row_col(
            min_row=self.sheet.min_row,
            min_col=self.sheet.min_column,
            max_row=self.sheet.max_row,
            max_col=self.sheet.max_column)
    
    def set_sheet(self, sheet_name:Union[str, Worksheet]):
        if isinstance(sheet_name, Worksheet):
            self.sheet = sheet_name
            return
        if sheet_name!=None:
            try:
                self.sheet = self.book[sheet_name]
            except KeyError as e:
                if 'does not exist.' in str(e):
                    # 主にWSheetがを存在しないときに発生する
                    # KeyError 'Worksheet Copy does not exist.'
                    # KeyError("Worksheet {0} does not exist.".format(key))
                    msg = str(e)
                    filename = Path(self.file_path).name
                    msg += '(file={}, sheet={})'.format(filename, sheet_name)
                    raise KeyError(msg)
                else:
                    raise e
        else:
            self.sheet = None
    
    def set_workbook(self, file_path:Union[str,Workbook], data_only:bool):
        
        if isinstance(file_path, Workbook):
            self.book = file_path
            self.file_path = None
            return
        self.file_path = str(file_path)
        if file_path!=None:
            if not Path(file_path).exists():
                raise FileNotFoundError(file_path)
            self.book = openpyxl.load_workbook(file_path, data_only=data_only)
        else:
            self.book = None

    def save_book(self, file_path:str=None):
        """
        WorkBookを保存する

        Caution:
            * 拡張子は'.xlsx'にすること。
             '.xlsm'で保存すると、開けなくなるので注意。
              この場合拡張子を'.xlsx'に変更すると開けるようになる。
            * ファイルを開いているとエラーが発生する 
                PermissionError [Errno 13] Permission denied: 'myworkbook.xlsx'
        """
        self.book.save(self._get_file_path(file_path))

    def close(self):
        if self.book!=None:
            self.book.close()

    def _get_file_path(self, file_path:str):
        if file_path==None:
            file_path = self.file_path
        return file_path

    def set_address_by_find(
            self, 
            keyword:str, 
            find_begin_address:str=None, 
            find_end_address:str=None, 
            found_number:int=0,
            debug:bool=True,
            check_after:bool=True):
        """
        文字列を検索して、存在したらセルのアドレスをセットする

        Args:
            keyword : regix pattern
            found_number : 複数発見したときのindex
        """
        if find_begin_address==None:
            find_begin_address = 'A1'
        if find_end_address==None:
            find_end_address = get_a1_address_from_row_and_col(
                self.valid_cells.end_cell.row, self.valid_cells.end_cell.col)
        # address_list = ExcelSheetDataUtil.find_value(
        #     self.sheet, keyword, find_begin_address, find_end_address, debug)
        address_list = CellUtil.get_address_list_by_find_keyword(
            self.sheet, keyword, find_begin_address, find_end_address, debug)
        if 0<len(address_list):
            address = address_list[found_number]
        else:
            address = None
        self.address = address
        self.cell = get_cells(self.sheet, address)
        if check_after:
            if not self.address_is_valid(self.address):
                msg = 'Not found value(address={})'.format(self.address)
                msg += '(sheet={}, keyword={}, begin={}, end={})'.format(self.sheet.title, keyword, find_begin_address, find_end_address)
                raise Exception(msg)
        # return address
        return self.cell

    def set_cell_by_find(self, keyword:str, find_begin_address:str=None, find_end_address:str=None, found_number:int=0):
        """
        文字列を検索して、存在したらCellをセットする

        Args:
            keyword : regix pattern
            found_number : 複数発見したときのindex
        """
        address = self.set_address_by_find(keyword, find_begin_address, find_end_address, found_number)
        self.cell = get_cells(self.sheet, address)
        return self.cell

    def set_cell(self, value):
        if isinstance(value, Cell):
            self.cell = value
        elif isinstance(value, str):
            self.set_address_a1(value)
        else:
            msg = str(type(value))
            raise TypeError(msg)

    def set_address_a1(self, a1_address:str):
        """
        self.address, self.cell をA1形式の文字列によりセットする
        """
        self.address = a1_address
        self.cell = get_cells(self.sheet, a1_address)
        return self.cell

    def set_address_r1c1(self, row=None, col=None):
        """
        row,col によって self.addressをセットする
         row,colがNoneの場合は self.addressの値を使用する。
        """
        if row==None or col==None:
            self_row, self_col=self.get_row_and_col()
            if row==None:
                row = self_row
            if col==None:
                col = self_col
        address = self._cnv_a1_address_from_row_col(row, col)
        self.set_address_a1(address)

    @classmethod
    def address_is_valid(cls, address):
        ret = re.search('^[A-Z]{1,3}\d+$', str(address))
        if ret!=None:
            return True
        else:
            return False

    # rename find_value
    @classmethod
    def get_cells_by_find_keyword(
        cls,
        sheet:Worksheet,
        keyword:str,
        find_begin:'Union[str,Cell]'=None,
        find_end:'Union[str,Cell]'=None):
        """
        文字列を検索して、keywordが含まれるセルのリスト取得する

        Args:
            keyword : regix pattern
        """
        # TODO:
        # 発見したら終了させる
        # Rowを優先、Colを優先を実装
        if find_begin==None or find_end==None:
            # address = search_entire_sheet(sheet, keyword)
            cell_list = _get_cells_of_entire_column_by_search(sheet, keyword)
        else:
            # address = search_rectangle_in_sheet(
            #     sheet, find_begin_address, find_end_address, keyword)
            earch_cells = _get_cells_iterator(sheet, find_begin, find_end)
            cell_list = _get_cells_by_search_keyword_in_rectangle(
                sheet, find_begin, find_end, keyword)
        return cell_list

    # rename find_value
    @classmethod
    def get_address_list_by_find_keyword(
        cls,
        sheet:Worksheet,
        keyword:str,
        find_begin_address:str=None,
        find_end_address:str=None,
        debug:bool=False):
        """
        文字列を検索して、存在したらセルのアドレスをセットする

        Args:
            keyword : regix pattern
        """
        if find_begin_address==None or find_end_address==None:
            address_list = search_entire_sheet(sheet, keyword)
        else:
            address_list = search_rectangle_in_sheet(
                sheet, find_begin_address, find_end_address, keyword)
        return address_list

    # >get_address_by_find_keyword
    @classmethod
    def find_value(
        cls,
        sheet:Worksheet,
        keyword:str,
        find_begin_address:str=None,
        find_end_address:str=None,
        debug:bool=False):
        """
        文字列を検索して、存在したらセルのアドレスをセットする
         deprecated method

        Args:
            keyword : regix pattern
        """
        if find_begin_address==None or find_end_address==None:
            address_list = search_entire_sheet(sheet, keyword)
        else:
            address_list = search_rectangle_in_sheet(
                sheet, find_begin_address, find_end_address, keyword, debug)
        return address_list
    
    def value_is_blank(self):
        row, col = get_row_and_col_from_a1_address(self.address)
        value = self.sheet.cell(row=row, column=col)
        return _value_is_blank(value)

    def get_end_cell_to_end_vertical(self, direction):
        """ 連続した空白or連続した入力済みのセルの最終Cellを取得する 縦方向-垂直"""
        return get_end_cell_to_end_vertical(
            self.sheet, self.address, direction, self.loop_max_row)

    def get_end_address_to_end_vertical(self, direction):
        """ 連続した空白or連続した入力済みのセルの最終のアドレスを取得する 縦方向-垂直"""
        max_row = self._get_loop_max_row()
        return get_end_address_to_end_vertical(
            self.sheet, self.address, direction, max_row)

    def get_end_cell_to_end_horizon(self,direction):
        """ 連続した空白or連続した入力済みのセルの最終Cellを取得する 横方向-水平"""
        return get_end_cell_to_end_horizon(
            self.sheet, self.address, direction, self.loop_max_col)

    def set_end_address_to_end_vertical(self, direction):
        """ 連続した空白or連続した入力済みのセルの最終のアドレスを取得する 縦方向-垂直"""
        address = self.get_end_address_to_end_vertical(direction)
        self.set_address_a1(address)
        return self.cell

    def get_end_address_to_end_horizon(self,direction):
        """ 連続した空白or連続した入力済みのセルの最終のアドレスを取得する 横方向-水平"""
        max_col = self._get_loop_max_col()
        return get_end_address_to_end_horizon(
            self.sheet, self.address, direction, max_col)
    
    def set_end_address_to_end_horizon(self,direction):
        """ 連続した空白or連続した入力済みのセルの最終のアドレスを取得する 横方向-水平"""
        self.address = self.get_end_address_to_end_horizon(direction)
        return self.address

    def _get_loop_max_col(self, update_info:bool=False):
        if self.loop_max_col!=None:
            return self.loop_max_col
        else:
            if update_info:
                self._update_valid_cell_in_sheet()
            return self.valid_cells.end_cell.col

    def _get_loop_max_row(self, update_info:bool=False):
        if self.loop_max_row!=None:
            return self.loop_max_row
        else:
            if update_info:
                self._update_valid_cell_in_sheet()
            return self.valid_cells.end_cell.row

    def get_range_address(self, direction=Direction.RIGHT|Direction.DOWN):
        """
        self.address から Directionの方向に、連続したセルをチェックして、
        矩形のアドレスの始点と終点（左上と右下）を取得する

        Returns:
            range_address
        """
        horizon_end_address = self.get_end_address_to_end_horizon(direction)
        vertical_end_address = self.get_end_address_to_end_vertical(direction)
        begin_address, end_address = get_range_address([horizon_end_address, vertical_end_address])
        self.range_address = begin_address + ':' + end_address
        return self.range_address
    
    def get_table_address(self, direction=Direction.RIGHT|Direction.DOWN):
        """
        現在のCellから縦横のDirection方向に連続した入力セルをたどり、矩形アドレス文字列を取得する
         deprecated
        """
        return get_range_address(Direction)

    def get_range_cells(self, direction=Direction.RIGHT|Direction.DOWN, check:bool=False):
        """ 現在のCellから縦横のDirection方向に連続した入力セルをたどり、矩形のCellを取得する """
        horizon_end_cell = self.get_end_cell_to_end_horizon(direction)
        vertical_end_cell = self.get_end_cell_to_end_vertical(direction)
        table_cells:'list[Cell]' = get_range_cell(self.sheet, [horizon_end_cell, vertical_end_cell])
        # if check:
        #     # リスト内のrowをすべてチェックして、Max,Minを取得する
        #     min_cell = get_min_cell(self.sheet, table_cells)
        #     max_cell = get_max_cell(self.sheet)
        # else:
        #     min_cell = table_cells[0]
        #     max_cell = table_cells[-1]
        return table_cells
    
    def get_range_address_in_cell_list(self, cell_list):
        """
        リスト内のcellをすべてチェックして、A1形式のアドレス"A1:B2"に変換する
         (2次元リストにも対応)
        """
        # リスト内のrowをすべてチェックして、Max,Minを取得する
        # 2次元リスト対応
        # get_min_cell.__code__
        if self._cell_list_is_2d(cell_list):
            min_cell = get_min_cell_2d(self.sheet, cell_list)
            max_cell = get_max_cell_2d(self.sheet, cell_list)
        else:
            min_cell = get_min_cell(self.sheet, cell_list)
            max_cell = get_max_cell(self.sheet, cell_list)
        ret = str(min_cell.coordinate) + ':' + max_cell.coordinate
        return ret


    def _cell_list_is_2d(self, cell_list):
        if isinstance(cell_list, list):
            if 0<len(cell_list):
                if isinstance(cell_list[0], list):
                    return True
        elif isinstance(cell_list, tuple):
            if 0<len(cell_list):
                if isinstance(cell_list[0], tuple):
                    return True 
        return False

    ### add for test4
    def get_begin_address_in_range(self):
        if not ':' in self.range_address:
            return None
        ret = self.range_address.split(':')[0]
        return ret
    def get_end_address_in_range(self):
        if not ':' in self.range_address:
            return None
        ret = self.range_address.split(':')[1]
        return ret

    def get_rows_range(self):
        begin_cell, end_cell = get_row_and_col_from_a1_address_range(
            self.address)
        return range(begin_cell.row, end_cell.row+1)
    
    def get_cols_range(self):
        begin_cell, end_cell = get_row_and_col_from_a1_address_range(
            self.address)
        return range(begin_cell.column, end_cell.column+1)
    
    def set_address_entire(self, opt):
        """
        列全体（or行全体）をself.addressにセットする

        Args:
            opt : row=1, col=2
        """
        if opt==ConstExcel.ROW:
            # row
            re_ret = re.search('[1-9]{1,8}', self.address)
            row_str = re_ret.group()
            begin_address = 'A' + row_str
            col_max = get_column_letter(self.valid_cells.end_cell.col)
            end_address = col_max + row_str
        elif opt==ConstExcel.COL:
            # col
            re_ret = re.search('[a-zA-Z]{1,3}', self.address)
            col_str = re_ret.group()
            begin_address = col_str + '1'
            end_address = col_str + str(self.valid_cells.end_cell.row)
        self.address = begin_address + ':' + end_address
        return self.address

    def set_range_address(self, address_value:'Union[str, list[str]]'):
        if isinstance(address_value, str):
            address_value = [address_value]
        elif isinstance(address_value, list):
            pass
        else:
            msg = 'address_value type is invalid. (type={})'.format(type(address_value))
            raise TypeError(msg)
        if self.address != '':
            address_value = [self.address] + address_value
        begin_address, end_address = get_range_address(address_value)
        self.range_address = begin_address + ':' + end_address

    def _address_is_loop_max_row(self, address:str):
        """ address がこのクラス内で処理する最大 "行" かを判定する """
        arg_row, arg_col = get_row_and_col_from_a1_address(address)
        if self.loop_max_row <= arg_row:
            return True
        else:
            return False

    def _address_is_loop_max_col(self, address:str):
        """ address がこのクラス内で処理する最大 "列" かを判定する """
        arg_row, arg_col = get_row_and_col_from_a1_address(address)
        if self.loop_max_row <= arg_row:
            return True
        else:
            return False

    def move_address(self, offset_row:int, offset_col:int):
        """ self.cell から offset_row, offset_col の分を移動する """
        # now_row, now_col = get_row_and_col_from_a1_address(self.address)
        # new_row = now_row + offset_row
        # new_column = now_col + offset_col
        # new_address = get_a1_address_from_row_and_col(new_row, new_column)
        # self.address = new_address
        # return new_address
        cell = get_offset_cell(self.sheet, self.cell, offset_row, offset_col)
        self.set_cell(cell)
        return cell

    @classmethod
    def move_address_from(address:str, offset_row:int, offset_col:int):
        """
        addressから offset_row, offset_col の分を移動する

        Args:
            Address:A1形式のアドレスのみ対応
        """
        now_row, now_col = get_row_and_col_from_a1_address(address)
        new_row = now_row + offset_row
        new_column = now_col + offset_col
        new_address = get_a1_address_from_row_and_col(new_row, new_column)
        return new_address

    def get_value(self, address=None):
        if address==None:
            address = self.address
        return get_cell_value(self.sheet, address)

    def get_value_r1c1(self, row:int, col:int):
        # self.sheet.cell(row=row, column=col).value
        return get_cell_value_r1c1(self.sheet, row, col)

    def get_cell_r1c1(self, row:int, col:int):
        # self.sheet.cell(row=row, column=col).value
        return self.sheet.cell(row=row, column=col)
    
    def get_cell(self):
        # self.sheet.cell(row=row, column=col).value
        row, col = get_row_and_col_from_a1_address(self.address)
        return self.sheet.cell(row=row, column=col)
    
    @classmethod
    def _get_cell(cls, address:str):
        # self.sheet.cell(row=row, column=col).value
        row, col = get_row_and_col_from_a1_address(address)
        return Cell(worksheet=None, row=row, column=col)

    def set_value(self, value:str, address:str=None):
        if address==None:
            address = self.address
        return set_cell_value(self.sheet, address, value)

    def copy_value(self, src_cell:Union[Cell,'ExcelSheetDataUtil'], style:bool=False, exists_only:bool=True):
        """
        引数src_cellから self.addressに値をコピーする

        Args:
            src_cell : コピー元のCell
                ExcelSheetDataUtil でも可能
            style : True=書式をコピーする
            exists_only : True=値or書式が設定されているもののみコピーする
        """
        # dist_cell:Cell = self.sheet[self.address]
        dist_cell = self._get_openxl_cell(self)
        src_cell:Cell = self._get_openxl_cell(src_cell)
        src_value = src_cell.value
        if src_value != '':
            dist_cell.value = src_value
        src_style = src_cell.style
        if src_style != self.ConstExcel.STYLE_NORMAL:
            # テーブルの書式は無視され’標準’となる
            if style:
                if src_cell.has_style :
                    dist_cell.style = src_cell.style
                    dist_cell.fill = src_cell.fill
        if style:
            dist_cell.style = src_cell.style
            # 例外が発生しました: TypeError unhashable type: 'StyleProxy'
            # コピーしないと上記エラーとなる
            dist_cell.fill = copy.copy(src_cell.fill)
            dist_cell.border = copy.copy(src_cell.border)
    
    def _get_openxl_cell(self, value:Union[Cell,'ExcelSheetDataUtil']):
        if 'ExcelSheetDataUtil' in str(type(value)):
            ret = value.sheet[value.address]
        elif isinstance(value, Cell):
            ret = value
        else:
            ret = value
        return ret

    def set_value_r1c1(self, value:str, row:int, col:int):
        # self.sheet.cell(row=row, column=col).value
        return get_cell_value_r1c1(self.sheet, value, row, col)

    def get_diff_row_and_col(self, row:int , col:int):
        now_row, now_col = get_row_and_col_from_a1_address(self.address)
        return row - now_row, col - now_col

    def get_offset_row_and_col(self, offset_row:int, offset_col:int):
        row, col = get_row_and_col_from_a1_address(self.address)
        return row + offset_row , col + offset_col

    def get_row_and_col(self):
        """
        self.address(A1形式のアドレス)から、行と列を取得する

        Memo:
            'A1:B5' は A1(row=1,col=1) を返す
        Returns:
            int, int  :  row_number, column_number
        """
        row, col = get_row_and_col_from_a1_address(self.address)
        return row, col
    
    def get_row_and_col_begin(self):
        address = self.get_begin_address()
        return get_row_and_col_from_a1_address(address)
    
    def get_row_and_col_end(self):
        address = self.get_end_address()
        return get_row_and_col_from_a1_address(address)
    
    def get_begin_address(self):
        if ':' in self.address:
            buf = self.address.split(':')
            return buf[0]
        else:
            return self.address

    def get_end_address(self):
        if ':' in self.address:
            buf = self.address.split(':')
            return buf[-1]
        else:
            return self.address
    
    def get_row(self):
        row, col = get_row_and_col_from_a1_address(self.address)
        return row

    def get_col(self):
        row, col = get_row_and_col_from_a1_address(self.address)
        return col
    
    @classmethod
    def _cnv_a1_address_from_cell(cls, cell:Cell):
        return get_a1_address_from_cell(cell)
        return _get_cells_address(cell)

    @classmethod
    def _cnv_row_col_from_a1_address(cls, a1_address):
        """ A1 > (1,1)  """
        return get_row_and_col_from_a1_address(a1_address)

    @classmethod
    def get_row_and_col_from_a1_address(cls, a1_address):
        """ A1 > (1,1)  """
        return get_row_and_col_from_a1_address(a1_address)
    
    @classmethod
    def _cnv_a1_address_from_row_col(cls, row, col):
        """ (1,1) > A1 """
        return get_a1_address_from_row_and_col(row, col)

    @classmethod
    def get_a1_address_from_row_and_col(cls, a1_address):
        """ A1 > (1,1)  """
        return get_a1_address_from_row_and_col(a1_address)

    @classmethod
    def _get_cells_address(cls, cell:Cell):
        """
        Cell からセルの番地を取得する(A1形式)
        """
        return _get_cell_address(cell)

    def get_values_from_range_address_np(self, range_address:str=None):
        """
        矩形のセルからすべて値を取得する

        Args:
            range_address : 'A1:B2'
        Returns:
            numpy.Array
        """
        import numpy as np
        if range_address==None:
            range_address = self.range_address
        row, col = get_row_and_col_from_a1_address(range_address)
        # print('range_address info = '.format((range_address, row, col)))
        begin_row, begin_col, end_row, end_col = get_row_and_col_from_rect_address(range_address)

        col_amount = abs(end_col+1 - begin_col)
        cell_values_list = np.arange(0)
        for row in range(begin_row-1, end_row):
            row += 1
            data_rows = np.zeros(col_amount, dtype=object) #型指定しないと数値となる
            for i, col in enumerate(range(begin_col, end_col+1)):
                value = self.get_value_r1c1(row, col)
                np.put(data_rows, [i], value)
            # data_rowsを cell_values_list に追加する
            if cell_values_list.size < 1:
                cell_values_list = data_rows
            else:
                cell_values_list = np.vstack((cell_values_list, data_rows))
        return cell_values_list



    def get_values_from_range_address_pd(
            self, range_address:str=None, columns:int=None, index:int=None):
        """
        矩形のセルからすべて値を取得する

        Args:
            range_address : 'A1:B2'
        Returns:
            pandas.DataFrame
        """
        import pandas as pd
        cell_values_np = self.get_values_from_range_address_np(range_address)
        if columns!=None:
            columns = cell_values_np[0]
            cell_values_np = cell_values_np[1:]
        if index!=None:
            index = cell_values_np[:, 0]
            cell_values_np = cell_values_np[:, 1:]
        df = pd.DataFrame(
            cell_values_np,
            columns=columns, index=index)
        return df

    @classmethod
    def _cnv_datetime(cls, value:str):
        if isinstance(value, str):
            if not value.isdigit():
                return value
        elif isinstance(value, int):
            pass
        else:
            return value
        return(datetime(1899, 12, 30) + timedelta(days=value))
    
    def copy_self(self):
        return copy.copy(self)

    def find_entire_row_in_range(self, index_row:int, keyrowd):
        """ 
        self.addredd の行を指定してkeywordに合致するCellを取得する
        """
        row, begin_col = self.get_row_and_col_begin()
        row += index_row
        # if index_row==0 and 1 < row:
        #     begin_row -= 1
        begin_row = row - 1
        _, end_col = self.get_row_and_col_end()
        direction = Direction.DOWN + Direction.RIGHT
        cells_2d = self._get_iter_cells_by_r1c1(
            self.sheet,
            begin_row, begin_col, row, end_col,
            direction)
        cell:Cell=None
        for cell_list in cells_2d:
            for cell in cell_list:
                val = _get_cell_value(cell)
                # # print('cell[{}, {}]'.format(cell.row, cell.column))
                # add = get_a1_address_from_row_and_col(cell.row, cell.column)
                # print('cell[{}]={}'.format(add, val))
                if keyrowd in val:
                    return cell
        return None

    def get_entire_col_in_range_by_cell(self, cell:Cell):
        """ 
        self.addredd の中にcellと一致するcolがあれば、その列を取得する
        """
        begin_row, begin_col = self.get_row_and_col_begin()
        end_row, end_col = self.get_row_and_col_end()
        col = cell.column
        if not(begin_col <= col  and col <= end_col):
            # col_index = col - begin_col
            return None
        cells_2d = self._get_iter_cells_by_r1c1(
            self.sheet,
            begin_row, col, end_row, col,
            Direction.DOWN & Direction.RIGHT, row_first=False)
        return cells_2d

    # def is_include_address()

    def find_entire_col_in_range(self, index_col:int, keyrowd):
        """ 
        self.addredd の行を指定してkeywordに合致するCellを取得する
        """
        begin_row, col = self.get_row_and_col_begin()
        col += index_col
        end_row, _ = self.get_row_and_col_end()
        cells_2d = self._get_iter_cells_by_r1c1(
            self.sheet,
            begin_row, col, end_row, col,
            Direction.DOWN & Direction.RIGHT)
        cell:Cell=None
        for cell_list in cells_2d:
            for cell in cell_list:
                val = _get_cell_value(cell)
                # print('cell[{}, {}]'.format(cell.row, cell.column))
                add = get_a1_address_from_row_and_col(cell.row, cell.column)
                print('cell[{}]={}'.format(add, val))
                if keyrowd in val:
                    return cell
        return None

    @classmethod
    def _get_iter_cells_by_r1c1(
        cls, worksheet: Worksheet, 
        begin_row, begin_col, end_row, end_col,
        direction:int=Direction.DOWN, row_first:bool=True):
        """
        Cellの配列を取得する
         numpyの2次元配列（type=Cell）を取得する
        """
        begin_row, begin_col, end_row, end_col = _align_row_col_with_direction(
            begin_row, begin_col, end_row, end_col, direction)
        if begin_col == end_col:
            begin_col -= 1
        # ###
        # if direction & Direction.DOWN:
        #     begin_row -= 1
        #     end_row -= 1
        # else:
        #     begin_row += 1
        #     end_row += 1
        # ###
        # row_range = _get_for_range_b(begin_row, end_row)
        row_range = _get_for_range(begin_row, end_row)
        # col_range = _get_for_range(begin_col, end_col)
        col_range = range(begin_col, end_col)
        # 1列ごとに取得していく（1列の行をすべて取得して、次の列の行を･･･となる）
        if row_first:
            # col_amount = abs(end_col+1 - begin_col)
            col_amount = len(col_range)
            cell_list = np.arange(0)
            for row in row_range:
                row += 1
                cell_list_row = np.zeros(col_amount, dtype=object) #型指定しないと数値となる
                for i, col in enumerate(col_range):
                    value = worksheet.cell(row=row, column=col)
                    np.put(cell_list_row, [i], value)
                # data_rowsを cell_list に追加する
                if cell_list.size < 1:
                    cell_list = cell_list_row
                    # cell_list = np.vstack((cell_list, cell_list_row))
                else:
                    cell_list = np.vstack((cell_list, cell_list_row))
        else:
            # row_amount = abs(begin_row+1 - end_row)
            # row_amount = abs(begin_row - end_row)
            row_amount = len(row_range)
            # col_amount = abs(begin_col+1 - end_col)
            cell_list = np.arange(0)
            for col in col_range:
                col += 1
                cell_list_col = np.zeros(row_amount, dtype=object) #型指定しないと数値となる
                for i, row in enumerate(row_range):
                    value = Cell(worksheet=worksheet, row=row, column=col)
                    print('** {}'.format(value.coordinate))
                    np.put(cell_list_col, [i], value)
                # data_rowsを cell_list に追加する
                if cell_list.size < 1:
                    cell_list = cell_list_col
                    # cell_list = np.vstack((cell_list, cell_list_col))
                else:
                    cell_list = np.vstack((cell_list, cell_list_col))
        # 1次元と2次元の時がある
        if cell_list.ndim == 1:
            cell_list = [cell_list]
        return cell_list

##########################################################################################
##########################################################################################
##########################################################################################
##########################################################################################



def main():
    # filename = 'FileIO.xlsm'
    filename = 'myworkbook.xlsx'
    wb = openpyxl.load_workbook(filename)
    ws = wb['TestData']

    cellu = CellUtil(ws)
    # result = search_rectangle_in_sheet(ws, 'A1', 'F38', '■TableA')
    result  = cellu.set_address_by_find('■TableA','A1', 'F38')
    print(result)
    # result = search_entire_sheet(ws, '■TableA')
    result  = cellu.set_address_by_find('■TableA')
    print(result)

    result  = cellu.set_cell_by_find('■TableA')
    print(result)
    #### 下方向に向かって連続した空白か入力済みセルの最終アドレスを取得する
    sheet = ws
    # begin_address = result[0]
    begin_address = cellu.address
    _MAX = 32767
    begin_row, begin_col = get_row_and_col_from_a1_address(begin_address)
    begin_val = sheet[begin_address]
    begin_is_blank = _value_is_blank(begin_val)
    end_row = begin_row
    for offset_row in range(_MAX):
        row = begin_row + offset_row
        now_val = sheet.cell(row=row, column=begin_col).value
        now_is_blank = _value_is_blank(now_val)
        # print((now_val, now_is_blank))
        if begin_is_blank == now_is_blank:
            continue
        else:
            end_row = begin_row + offset_row - 1
            
            break
    end_address = get_a1_address_from_row_and_col(end_row, begin_col)
    print(end_address)
    val = sheet[end_address].value
    print(val)
    ###
    print('###')
    #
    # end_address = get_end_address_to_end_vertical(
    #     sheet, begin_address, Direction.TOP, _MAX)
    end_address = cellu.get_end_address_to_end_vertical(Direction.UP)
    print('begin to UP = {}'.format(end_address))
    val = get_cell_value(sheet, end_address)
    print(val)
    ##
    end_cell = cellu.get_end_cell_to_end_vertical(Direction.UP)
    print('begin to UP cell = {}'.format(end_address))
    val = get_cell_value(sheet, end_cell)
    print(val)
    #
    # end_address = get_end_address_to_end_vertical(
    #     sheet, begin_address, Direction.BOTTOM, _MAX)
    end_address = cellu.get_end_address_to_end_vertical(Direction.DOWN)
    print('begin to DOWN = {}'.format(end_address))
    val = get_cell_value(sheet, end_address)
    print(val)
    ##
    end_cell = cellu.get_end_cell_to_end_vertical(Direction.DOWN)
    print('begin to DOWN cell= {}'.format(end_cell))
    val = get_cell_value(sheet, end_cell)
    print(val)
    #
    # begin_address_b = get_offset_address(begin_address,2,0)
    begin_cell_b = cellu.move_address(2,0)
    begin_address_b = begin_cell_b.coordinate
    table_begin_address = begin_address_b
    table_beginn_cell = begin_cell_b
    #
    # end_address = get_end_address_to_end_horizon(
    #     sheet, begin_address_b, Direction.LEFT, _MAX)
    end_address = cellu.get_end_address_to_end_horizon(Direction.LEFT)
    print('begin to LEFT = {}'.format(end_address))
    val = get_cell_value(sheet, end_address)
    print(val)
    ##
    end_cell = cellu.get_end_cell_to_end_horizon(Direction.LEFT)
    print('begin to LEFT cell = {}'.format(end_cell))
    val = get_cell_value(sheet, end_cell)
    print(val)
    #
    # end_address = get_end_address_to_end_horizon(
    #     sheet, begin_address_b, Direction.RIGHT, _MAX)
    end_address = cellu.get_end_address_to_end_horizon(Direction.RIGHT)
    print('begin to RIGHT = {}'.format(end_address))
    val = get_cell_value(sheet, end_address)
    print(val)
    table_end_address = end_address
    ##
    end_cell = cellu.get_end_cell_to_end_horizon(Direction.RIGHT)
    print('begin to RIGHT cell = {}'.format(end_cell))
    val = get_cell_value(sheet, end_cell)
    print(val)
    table_end_cell = end_cell
    #
    print('***')
    table_address = get_table_address([table_begin_address, table_end_address])
    print('table = {}'.format(table_address))
    #
    print('***')
    table_cell = get_range_cell(cellu.sheet, [table_beginn_cell, table_end_cell])
    print('table cell = {}'.format(table_cell))
    # 
    print('=========')
    address = 'A1:B5'
    row, col = get_row_and_col_from_a1_address(address)
    print((address, row, col))

class CellUtil(ExcelSheetDataUtil):
    pass


if __name__ == '__main__':
    main()