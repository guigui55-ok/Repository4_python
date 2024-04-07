"""
セルのスタイルを変更する

表を読み込んで、特定の行の値が、特定の値だったら、行すべてのスタイルを変更する
"""


import excel_data
from excel_data import ExcelSheetDataUtil
from pathlib import Path
import pandas as pd


# print('*書き込みシートをすべてクリア')
file_name = 'myworkbook.xlsx'
sheet_name = 'my sheet1'
ex_data = ExcelSheetDataUtil(file_name, sheet_name, data_only=True)

print('データの用意')
keyword = 'ItemL'
ex_data.set_address_by_find(keyword, 'A1', 'P13', debug=False,)
print('begin_address = {}'.format(ex_data.address))
ex_data.range_address = ex_data.get_range_address()
print('range_address = {}'.format(ex_data.range_address))

df_a = ex_data.get_values_from_range_address_pd(
    ex_data.range_address, columns=None)
# first_row_as_list = df_a.iloc[0].astype(str).tolist()
# df_a.columns = first_row_as_list
print('df_a = ')
print(df_a.columns)
print(df_a.shape)
print(df_a)

df:pd.DataFrame = df_a
from openpyxl.cell import Cell
print('*Styleを変更する')
IS_SET_COLOR = False
IS_SET_COLOR = True
# for i , column_name in enumerate(df.columns):
#     print('Processing {}'.format(i))
#     for j, data_series_value in enumerate(df[column_name]):
#         value = data_series_value
#         cell:Cell = ex_data.get_offset_cell(j, i)
#         ex_data.set_value(value, cell.coordinate)
#         print('[i ,j] = [{}, {}] = ({}){}'.format(i, j, cell.coordinate, value))
#         cell.value = value
#     print('-------')
# print()

# https://qiita.com/github-nakasho/items/3f861395227e5645cce7
# patternTypeには‘darkDown’, ‘darkGrid’, ‘lightGrid’, ‘gray0625’, ‘lightGray’, ‘mediumGray’, ‘darkGray’, ‘darkHorizontal’, ‘solid’, ‘darkVertical’, ‘lightUp’, ‘lightTrellis’, ‘darkUp’, ‘lightDown’, ‘lightVertical’, ‘gray125’, ‘darkTrellis’, ‘lightHorizontal’を指定することができます。

print('書き込みのためにシートを読み直す（data_only=False)')
ex_data.reset_book_sheet(data_only=False)

import copy
from openpyxl.styles import PatternFill
fill = PatternFill(patternType='lightHorizontal', fgColor='d3d3d3', bgColor='ff1493')
fill = PatternFill(patternType='darkDown', fgColor='d3d3d3')
fill = PatternFill(fill_type="solid", fgColor="D3D3D3")
from excel_data import StrCell
# ItemL を 値で絞込
col_name ='ItemL'
col_name =0
target_list = ['L7','ItemL']
str_cell:StrCell = None
for target in target_list:
    filtered_values = df[df[col_name] == target]
    print(type(filtered_values)) #DataFrame,shpe=(1, 3)
    print(filtered_values.shape)
    for index, row in filtered_values.iterrows():
        for column, str_cell in row.items():
            str_cell.cell.fill = copy.copy(fill)
            print('{}[{}] ,patternType={}, fgColor={}, gbColor={}'.format(
                str_cell.cell.coordinate, str_cell.cell.value, fill.patternType,
                fill.fgColor.rgb, fill.bgColor.rgb))
        pass

print()

# from openpyxl.styles.fonts import Font
# from openpyxl.styles.colors import Color
# from openpyxl.styles.borders import Side
# # from openpyxl.styles.borders import _SideStyle
# print('*セルを取得して値を出力する')
# for address in address_list:
#     buf_cell = ex_data._get_cell(ex_data.sheet, address)
#     print('#############################')
#     print('address = {}'.format(buf_cell.coordinate))
#     # フォント属性
#     print('============ font')
#     font:Font = buf_cell.font
#     print('buf_cell.font.name = {}'.format(font.name))
#     print('buf_cell.font.size = {}'.format(font.size))
#     print('buf_cell.font.bold = {}'.format(font.bold))
#     print('buf_cell.font.italic = {}'.format(font.italic))
#     print('buf_cell.font.underline = {}'.format(font.underline))
#     print('buf_cell.font.strike = {}'.format(font.strike))
#     print('============ color')
#     color:Color = font.color
#     print('buf_cell.font.color = {}'.format(color))
#     print('color.rgb = {}'.format(color.rgb))
#     print('color.indexed = {}'.format(color.indexed))
#     print('color.type = {}'.format(color.type))
#     print('color.auto = {}'.format(color.auto))
#     print('color.theme = {}'.format(color.theme))
#     print('color.tint = {}'.format(color.tint))

#     print('============ fill')
#     # 塗りつぶし（背景色）
#     print('buf_cell.fill.fill_type = {}'.format(buf_cell.fill))
#     print('buf_cell.fill.fill_type = {}'.format(buf_cell.fill.fill_type))
#     print('buf_cell.fill.start_color = {}'.format(buf_cell.fill.start_color))
#     print('buf_cell.fill.end_color = {}'.format(buf_cell.fill.end_color))

#     print('============ border')
#     # 境界線
#     print('buf_cell.border.left = {}'.format(buf_cell.border.left))
#     side:Side = buf_cell.border.left
#     # print('side.border_style = {}'.format(side.border_style))
#     print('side.color = {}'.format(side.color))
#     print('side.color = {}'.format(side.style))
#     print('buf_cell.border.right = {}'.format(buf_cell.border.right))
#     print('buf_cell.border.top = {}'.format(buf_cell.border.top))
#     print('buf_cell.border.bottom = {}'.format(buf_cell.border.bottom))

#     print('============ etc')
#     # 数値フォーマット
#     print('buf_cell.number_format = {}'.format(buf_cell.number_format))

#     # セルの水平方向の配置
#     print('buf_cell.alignment.horizontal = {}'.format(buf_cell.alignment.horizontal))

#     # セルの垂直方向の配置
#     print('buf_cell.alignment.vertical = {}'.format(buf_cell.alignment.vertical))

#     # セルのテキストの折り返し
#     print('buf_cell.alignment.wrap_text = {}'.format(buf_cell.alignment.wrap_text))

#     # セルの縮小して全体を表示
#     print('buf_cell.alignment.shrink_to_fit = {}'.format(buf_cell.alignment.shrink_to_fit))

#     # セルのインデント
#     print('buf_cell.alignment.indent = {}'.format(buf_cell.alignment.indent))

#     # セルのテキストの向き
#     print('buf_cell.alignment.text_rotation = {}'.format(buf_cell.alignment.text_rotation))
#     print('#############################')





try:
    ex_data.save_book()
except PermissionError as e:
    msg = '[!]ERROR: ファイルアクセスエラーです。ファイルが開いていたらファイルを閉じてください。'
    print(msg)
    # raise e