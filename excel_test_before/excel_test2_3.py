"""
Excelの行列から特定の文字列を検索し、そこから入力連続したセルを縦・横に取得して、矩形アドレスを取得する
"""
# https://ramunememo.hatenablog.com/entry/2021/09/23/182202
import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import re
from typing import Union

def _get_cell_value(cell_val:Cell):
    """
    セルのデータを取得（文字列）
    """
    try:
        return str(cell_val.value)
    except Exception as e:
        # 文字列に変換できないエラーの時はスキップ 予定（未対応）
        # 何のエラーが発生するか確認してから対応する
        raise e

def _get_cells_address(cell_val:Cell):
    """
    セルの番地を取得
    """
    cell_address = get_column_letter(cell_val.column) +  str(cell_val.row)
    return cell_address

def _is_match_patterns(patterns:Union[str, list[str]], value:str):
    if not isinstance(patterns, list):
        patterns = [str(patterns)]
    for pattern in patterns:
        ret = re.search(pattern, value)
        if ret!=None:
            return True
    return False

def _get_rectangle(worksheet_val:Worksheet, begin_address:str, end_address:str):
    """ ワークシートと開始・終了アドレスから対象の範囲の セルリスト list[list[Cell]] を取得する """
    rectangle = worksheet_val[begin_address:end_address]
    return rectangle

def _search_keyword_in_rectangle(rectangle:list[list[Cell]] ,pattern:Union[str, list[str]]):
    """
    特定の範囲を検索（セルのリストは取得済みであること）

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列、または、パターン
    """
    result = []
    for col in rectangle:
        for cell in col:
            value = _get_cell_value(cell)
            if _is_match_patterns(pattern, value):
                cell_address = _get_cells_address(cell)
                result.append(cell_address)
    return result

def search_rectangle_in_sheet(
    worksheet_val:Worksheet,
    begin_address:str,
    end_address:str,
    keyword:Union[str, list[str]]):
    """
    指定したワークシートの、特定の範囲を検索
     (ワークシートと開始・終了アドレスを指定するVer)

    Args
        ws: worksheet object
        range_address: address (ex) 'A1:C5'
    """
    rectangle = _get_rectangle(worksheet_val,begin_address,end_address)
    result = _search_keyword_in_rectangle(rectangle, keyword)
    return result

def search_entire_sheet(worksheet_val:Worksheet, keyword):
    """
    シート全体を検索する

    Args:
        rectangle: セルのリスト Worksheet["A1":"B2"]で取得されたもの
        keyword: 検索対象の文字列、または、パターン
    """
    rectangle = worksheet_val.columns
    result = _search_keyword_in_rectangle(rectangle, keyword)
    return result

def main():
    filename = 'FileIO.xlsm'
    wb = openpyxl.load_workbook(filename)
    ws = wb['TestData']

    result = search_rectangle_in_sheet(ws, 'A1', 'F38', '■TableA')
    print(result)
    result = search_entire_sheet(ws, '■TableA')
    print(result)

if __name__ == '__main__':
    main()