import openpyxl

import pathlib,os
dir_path_obj = pathlib.Path(__file__).parent
read_file_path = str(dir_path_obj.joinpath('test13.xlsx'))
result_file_path = str(dir_path_obj.joinpath('test13_result.xlsx'))

wb1 = openpyxl.load_workbook(read_file_path)
wb2 = openpyxl.load_workbook(result_file_path)

sheet = []
ws2 = wb2['Sheet1']
invoice_gyo = []
co = []
for l in range(2,6):#アンケのデータの行
    co.append(l)
for j in range(10,35):#入力先のコマ数の欄
    invoice_gyo.append(j)
for i in range(1):#請求書のシート(名前)のデータ
    print(wb1.worksheets[i])
    sheet.append(wb1.worksheets[i])
#請求書のリストの名前とアンケートの名前を照合する

for name in sheet:
    for col in co:   #アンケート側の名前を一回ずつ認証し値を入れる　#←②この操作にいってほしい
        buf = '<Worksheet'+' '+ '"'+str(ws2.cell(row=col,column=4).value)+'"'+'>'
        print(buf)
        if str(name) == buf:
           for k in invoice_gyo:
                print('sheet={}  , gyo={}'.format(name, k))
                buf2 = name.cell(row = k, column = 3).value
                if buf2 is None: #←①この操作を一回したら
                    # buf3 = ws2.cell(row=col,column=5).value
                    buf3 = ws2.cell(row=k,column=5).value
                    print('buf3 = {}'.format(buf3))
                    name.cell(row = k, column = 3).value =  buf3
wb1.save(dir_path_obj.joinpath('test13_jikken.xlsx'))